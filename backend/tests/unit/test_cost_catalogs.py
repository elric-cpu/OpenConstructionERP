"""Unit tests for user-owned cost catalogs.

Covers:
    - Catalog CRUD: create with required currency, list with item counts,
      rename, and the currency-change guard (rejected while the catalog
      has items, allowed when empty).
    - File import into a catalog: inline catalog creation, catalog_currency
      requirement when the file has no currency column, currency inheritance
      for rows without one, and the mixed_currency_count warning.
    - Required-mapping enforcement: import rejects with 422 when the mapping
      does not cover description and rate.
    - Excel export of one catalog: header + rows round-trip, catalog-currency
      fallback, formula-injection neutralisation via _excel_safe.
    - Delete modes: keep_items detaches rows, delete_items soft-deletes them.

Handlers run end-to-end against a transaction-isolated PostgreSQL session,
the same stack the production endpoints use.
"""

from __future__ import annotations

import io
import uuid

import pytest
import pytest_asyncio
from fastapi import HTTPException, UploadFile
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.modules.costs.models import CostCatalog, CostItem
from app.modules.costs.router import (
    create_cost_catalog,
    create_cost_item,
    delete_cost_catalog,
    export_cost_catalog_excel,
    import_cost_file,
    list_cost_catalogs,
    update_cost_catalog,
)
from app.modules.costs.schemas import (
    CostCatalogCreate,
    CostCatalogUpdate,
    CostItemCreate,
)
from app.modules.costs.service import CostCatalogService, CostItemService
from tests._pg import transactional_session

_USER_ID = str(uuid.uuid4())
# JWT-claims payload for the endpoints that now take ``CurrentUserPayload``
# (the owner-scoped catalog list + file import read ``sub`` and ``role`` off
# the dict). Endpoints still taking ``CurrentUserId`` keep the bare ``_USER_ID``.
_USER_PAYLOAD = {"sub": _USER_ID, "role": "admin"}


@pytest_asyncio.fixture
async def session() -> AsyncSession:
    async with transactional_session() as s:
        yield s


def _upload(content: bytes, filename: str = "rates.csv") -> UploadFile:
    return UploadFile(file=io.BytesIO(content), filename=filename)


async def _import_file(
    session: AsyncSession,
    content: bytes,
    *,
    column_map: str | None = None,
    catalog_id: str | None = None,
    catalog_name: str | None = None,
    catalog_currency: str | None = None,
) -> dict:
    """Call the import handler directly with every Form default made explicit.

    Calling the route function outside FastAPI leaves unset parameters as
    ``Form(...)`` sentinel objects, so this wrapper always passes real values.
    """
    return await import_cost_file(
        _USER_PAYLOAD,
        file=_upload(content),
        column_map=column_map,
        catalog_id=catalog_id,
        catalog_name=catalog_name,
        catalog_currency=catalog_currency,
        service=CostItemService(session),
        catalog_service=CostCatalogService(session),
    )


async def _read_streaming_body(response: object) -> bytes:
    chunks: list[bytes] = []
    async for chunk in response.body_iterator:  # type: ignore[attr-defined]
        chunks.append(chunk if isinstance(chunk, bytes) else chunk.encode("utf-8"))
    return b"".join(chunks)


# ── Catalog CRUD ───────────────────────────────────────────────────────────


async def test_create_and_list_catalogs(session: AsyncSession) -> None:
    catalog_service = CostCatalogService(session)

    created = await create_cost_catalog(
        CostCatalogCreate(name="My Rates", currency="eur", description="House catalog"),
        _USER_ID,
        service=catalog_service,
    )
    assert created.name == "My Rates"
    assert created.currency == "EUR"  # normalised to uppercase
    assert created.source == "manual"
    assert created.item_count == 0
    assert created.created_by == uuid.UUID(_USER_ID)

    # Attach one active and one inactive item: count must reflect ACTIVE only.
    item_service = CostItemService(session)
    await item_service.create_cost_item(
        CostItemCreate(code="C-1", description="Concrete", unit="m3", rate=100, catalog_id=created.id)
    )
    inactive = await item_service.create_cost_item(
        CostItemCreate(code="C-2", description="Steel", unit="kg", rate=2, catalog_id=created.id)
    )
    await item_service.delete_cost_item(inactive.id)

    listed = await list_cost_catalogs(_USER_PAYLOAD, service=catalog_service)
    assert len(listed) == 1
    assert listed[0].id == created.id
    assert listed[0].item_count == 1


async def test_catalog_currency_required_and_validated(session: AsyncSession) -> None:
    with pytest.raises(ValueError):
        CostCatalogCreate(name="No currency", currency="EURO")
    with pytest.raises(ValueError):
        CostCatalogCreate(name="No currency", currency="1$")


async def test_manual_item_inherits_catalog_currency(session: AsyncSession) -> None:
    catalog_service = CostCatalogService(session)
    catalog = await catalog_service.create_catalog(CostCatalogCreate(name="CHF book", currency="CHF"))

    item_service = CostItemService(session)
    response = await create_cost_item(
        CostItemCreate(code="M-1", description="Masonry", unit="m2", rate=55, catalog_id=catalog.id),
        _USER_ID,
        service=item_service,
    )
    assert response.currency == "CHF"
    assert response.catalog_id == catalog.id

    # An explicit row currency is never overwritten.
    response2 = await create_cost_item(
        CostItemCreate(
            code="M-2", description="Masonry imported", unit="m2", rate=60, currency="USD", catalog_id=catalog.id
        ),
        _USER_ID,
        service=item_service,
    )
    assert response2.currency == "USD"


async def test_manual_item_unknown_catalog_rejected(session: AsyncSession) -> None:
    item_service = CostItemService(session)
    with pytest.raises(HTTPException) as exc_info:
        await item_service.create_cost_item(
            CostItemCreate(code="X-1", description="Orphan", unit="m", rate=1, catalog_id=uuid.uuid4())
        )
    assert exc_info.value.status_code == 422


async def test_update_catalog_rejects_currency_change_with_items(session: AsyncSession) -> None:
    catalog_service = CostCatalogService(session)
    catalog = await catalog_service.create_catalog(CostCatalogCreate(name="Guarded", currency="EUR"))

    item_service = CostItemService(session)
    await item_service.create_cost_item(
        CostItemCreate(code="G-1", description="Grout", unit="kg", rate=4, catalog_id=catalog.id)
    )

    # Name change is fine.
    updated = await update_cost_catalog(
        catalog.id,
        CostCatalogUpdate(name="Guarded v2"),
        _USER_PAYLOAD,
        service=catalog_service,
    )
    assert updated.name == "Guarded v2"
    assert updated.item_count == 1

    # Currency change with items present is rejected with a clear message.
    with pytest.raises(HTTPException) as exc_info:
        await update_cost_catalog(
            catalog.id,
            CostCatalogUpdate(currency="USD"),
            _USER_PAYLOAD,
            service=catalog_service,
        )
    assert exc_info.value.status_code == 409
    assert "currency" in str(exc_info.value.detail).lower()

    # Once the catalog is empty the change goes through.
    empty = await catalog_service.create_catalog(CostCatalogCreate(name="Empty", currency="EUR"))
    changed = await update_cost_catalog(
        empty.id,
        CostCatalogUpdate(currency="USD"),
        _USER_PAYLOAD,
        service=catalog_service,
    )
    assert changed.currency == "USD"


# ── Import into a catalog ──────────────────────────────────────────────────


async def test_import_creates_catalog_and_inherits_currency(session: AsyncSession) -> None:
    csv_bytes = b"code,description,unit,rate\nP-001,Concrete C30,m3,120.50\nP-002,Rebar,kg,2.40\n"
    result = await _import_file(
        session,
        csv_bytes,
        catalog_name="Site rates 2026",
        catalog_currency="PLN",
    )

    assert result["imported"] == 2
    assert result["catalog_currency"] == "PLN"
    assert result["mixed_currency_count"] == 0
    catalog_id = uuid.UUID(result["catalog_id"])

    catalog = await session.get(CostCatalog, catalog_id)
    assert catalog is not None
    assert catalog.name == "Site rates 2026"
    assert catalog.source == "import"

    items = (await session.execute(select(CostItem).where(CostItem.catalog_id == catalog_id))).scalars().all()
    assert len(items) == 2
    # No currency column in the file: every row inherits the catalog currency.
    assert {item.currency for item in items} == {"PLN"}


async def test_import_requires_catalog_currency_without_currency_column(session: AsyncSession) -> None:
    csv_bytes = b"code,description,unit,rate\nP-001,Concrete,m3,100\n"
    with pytest.raises(HTTPException) as exc_info:
        await _import_file(session, csv_bytes, catalog_name="No currency anywhere")
    assert exc_info.value.status_code == 422
    assert "catalog_currency" in str(exc_info.value.detail)


async def test_import_into_existing_catalog_counts_mixed_currencies(session: AsyncSession) -> None:
    catalog_service = CostCatalogService(session)
    catalog = await catalog_service.create_catalog(CostCatalogCreate(name="EUR book", currency="EUR"))

    csv_bytes = (
        b"code,description,unit,rate,currency\n"
        b"R-001,Excavation,m3,18.00,\n"  # empty -> inherits EUR
        b"R-002,Imported pump,pcs,950.00,USD\n"  # differs -> counted, kept
        b"R-003,Formwork,m2,32.00,EUR\n"  # matches -> not counted
    )
    result = await _import_file(session, csv_bytes, catalog_id=str(catalog.id))

    assert result["imported"] == 3
    assert result["catalog_id"] == str(catalog.id)
    assert result["mixed_currency_count"] == 1

    items = (await session.execute(select(CostItem).where(CostItem.catalog_id == catalog.id))).scalars().all()
    by_code = {item.code: item for item in items}
    assert by_code["R-001"].currency == "EUR"  # inherited
    assert by_code["R-002"].currency == "USD"  # kept as-is, never rewritten
    assert by_code["R-003"].currency == "EUR"


async def test_import_rejects_both_catalog_id_and_name(session: AsyncSession) -> None:
    csv_bytes = b"code,description,unit,rate\nP-001,Concrete,m3,100\n"
    with pytest.raises(HTTPException) as exc_info:
        await _import_file(
            session,
            csv_bytes,
            catalog_id=str(uuid.uuid4()),
            catalog_name="Also a name",
            catalog_currency="EUR",
        )
    assert exc_info.value.status_code == 422


# ── Required-mapping enforcement ───────────────────────────────────────────


async def test_import_rejects_unmapped_required_columns(session: AsyncSession) -> None:
    # Headers that map to neither description nor rate.
    csv_bytes = b"colA,colB\nfoo,bar\n"
    with pytest.raises(HTTPException) as exc_info:
        await _import_file(session, csv_bytes)
    assert exc_info.value.status_code == 422
    detail = str(exc_info.value.detail)
    assert "description" in detail
    assert "rate" in detail


async def test_import_rejects_missing_rate_even_with_description(session: AsyncSession) -> None:
    csv_bytes = b"code,description,unit\nP-001,Concrete,m3\n"
    with pytest.raises(HTTPException) as exc_info:
        await _import_file(session, csv_bytes)
    assert exc_info.value.status_code == 422
    detail = str(exc_info.value.detail)
    assert "rate" in detail
    assert "description" not in detail.split("not mapped:")[1].split(".")[0]


async def test_import_accepts_user_column_map_for_required_fields(session: AsyncSession) -> None:
    # Headers in a language the alias map does not know; the user mapping
    # must satisfy the required-column gate.
    csv_text = "Towar,Opis pozycji,Jednostka,Stawka\nB-1,Beton C25,m3,310.00\n"
    result = await _import_file(
        session,
        csv_text.encode("utf-8"),
        column_map='{"code":"Towar","description":"Opis pozycji","unit":"Jednostka","rate":"Stawka"}',
        catalog_name="Cennik",
        catalog_currency="PLN",
    )
    assert result["imported"] == 1


# ── Export ─────────────────────────────────────────────────────────────────


async def test_export_catalog_excel_roundtrip(session: AsyncSession) -> None:
    catalog_service = CostCatalogService(session)
    catalog = await catalog_service.create_catalog(CostCatalogCreate(name="Export Book 2026!", currency="EUR"))

    item_service = CostItemService(session)
    await item_service.create_cost_item(
        CostItemCreate(
            code="E-1",
            description="Excavation",
            unit="m3",
            rate="18.50",
            catalog_id=catalog.id,
            classification={"din276": "320"},
        )
    )
    # Row with no own currency (inherits at create) and a formula-looking text.
    await item_service.create_cost_item(
        CostItemCreate(
            code="E-2",
            description="=HYPERLINK evil",
            unit="pcs",
            rate="1.00",
            catalog_id=catalog.id,
        )
    )
    # An item OUTSIDE the catalog must not leak into the export.
    await item_service.create_cost_item(CostItemCreate(code="OUT-1", description="Other", unit="m", rate=9))

    response = await export_cost_catalog_excel(
        catalog.id,
        session,
        _USER_PAYLOAD,
        service=catalog_service,
    )
    disposition = response.headers["content-disposition"]
    assert 'filename="export-book-2026.xlsx"' in disposition

    body = await _read_streaming_body(response)
    wb = load_workbook(io.BytesIO(body), read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = [list(row) for row in ws.iter_rows(values_only=True)]
    wb.close()

    assert rows[0] == ["Code", "Description", "Unit", "Rate", "Currency", "Classification"]
    assert len(rows) == 3  # header + 2 catalog items, the outside item excluded
    by_code = {row[0]: row for row in rows[1:]}
    assert by_code["E-1"][1] == "Excavation"
    assert by_code["E-1"][3] == 18.5
    assert by_code["E-1"][4] == "EUR"
    assert "din276" in by_code["E-1"][5]
    # Formula-injection neutralised by _excel_safe.
    assert by_code["E-2"][1] == "'=HYPERLINK evil"


async def test_export_unknown_catalog_404(session: AsyncSession) -> None:
    with pytest.raises(HTTPException) as exc_info:
        await export_cost_catalog_excel(
            uuid.uuid4(),
            session,
            _USER_PAYLOAD,
            service=CostCatalogService(session),
        )
    assert exc_info.value.status_code == 404


# ── Delete modes ───────────────────────────────────────────────────────────


async def test_delete_catalog_keep_items_detaches(session: AsyncSession) -> None:
    catalog_service = CostCatalogService(session)
    catalog = await catalog_service.create_catalog(CostCatalogCreate(name="To delete", currency="EUR"))
    item_service = CostItemService(session)
    item = await item_service.create_cost_item(
        CostItemCreate(code="K-1", description="Kept", unit="m", rate=5, catalog_id=catalog.id)
    )

    result = await delete_cost_catalog(catalog.id, _USER_PAYLOAD, mode="keep_items", service=catalog_service)
    assert result["items_affected"] == 1

    assert await session.get(CostCatalog, catalog.id) is None
    survivor = await session.get(CostItem, item.id)
    assert survivor is not None
    assert survivor.catalog_id is None
    assert survivor.is_active is True


async def test_delete_catalog_delete_items_soft_deletes(session: AsyncSession) -> None:
    catalog_service = CostCatalogService(session)
    catalog = await catalog_service.create_catalog(CostCatalogCreate(name="Purge", currency="EUR"))
    item_service = CostItemService(session)
    item = await item_service.create_cost_item(
        CostItemCreate(code="D-1", description="Doomed", unit="m", rate=5, catalog_id=catalog.id)
    )

    result = await delete_cost_catalog(catalog.id, _USER_PAYLOAD, mode="delete_items", service=catalog_service)
    assert result["items_affected"] == 1

    assert await session.get(CostCatalog, catalog.id) is None
    soft_deleted = await session.get(CostItem, item.id)
    assert soft_deleted is not None
    assert soft_deleted.is_active is False


async def test_delete_unknown_catalog_404(session: AsyncSession) -> None:
    with pytest.raises(HTTPException) as exc_info:
        await delete_cost_catalog(uuid.uuid4(), _USER_PAYLOAD, mode="keep_items", service=CostCatalogService(session))
    assert exc_info.value.status_code == 404
