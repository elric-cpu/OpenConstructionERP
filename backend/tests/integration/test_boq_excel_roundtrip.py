# DDC-CWICR-OE: DataDrivenConstruction · OpenConstructionERP
# Copyright (c) 2026 Artem Boiko / DataDrivenConstruction
"""End-to-end BOQ Excel/CSV round-trip fidelity tests (GitHub #360).

Exercises the full export -> edit -> re-import cycle against the live API:

* the Excel export stamps a stable ``Position ID`` column;
* re-importing a sheet whose rows carry a known id UPDATES those positions
  in place instead of duplicating them;
* a blank id CREATES a new position, an unknown / foreign id is imported as
  a NEW position (never an update) and flagged;
* positions absent from the sheet are only deleted on the opt-in
  ``?delete_missing=true`` flag, and the count is always reported;
* an id belonging to ANOTHER BOQ can never mutate a position across the
  boundary.

Run::

    cd backend
    python -m pytest tests/integration/test_boq_excel_roundtrip.py -v --tb=short
"""

from __future__ import annotations

import asyncio
import io
import uuid

import pytest
import pytest_asyncio
from httpx import ASGITransport, AsyncClient

from app.main import create_app

# ── Module-scoped fixtures (shared client + admin auth) ─────────────────────


@pytest_asyncio.fixture(scope="module")
async def shared_client():
    app = create_app()

    from contextlib import asynccontextmanager

    @asynccontextmanager
    async def lifespan_ctx():
        async with app.router.lifespan_context(app):
            yield

    async with lifespan_ctx():
        transport = ASGITransport(app=app)
        async with AsyncClient(transport=transport, base_url="http://test") as ac:
            yield ac


@pytest_asyncio.fixture(scope="module")
async def shared_auth(shared_client: AsyncClient) -> dict[str, str]:
    """Register a fresh user, promote to admin via DB, return Bearer header."""
    unique = uuid.uuid4().hex[:8]
    email = f"boqroundtrip-{unique}@test.io"
    password = f"BoqRound{unique}9"

    reg = await shared_client.post(
        "/api/v1/users/auth/register",
        json={
            "email": email,
            "password": password,
            "full_name": "BOQ Round-Trip Tester",
            "role": "admin",
        },
    )
    assert reg.status_code == 201, f"Registration failed: {reg.text}"

    from ._auth_helpers import promote_to_admin

    await promote_to_admin(email)

    token = ""
    data: dict = {}
    for attempt in range(3):
        resp = await shared_client.post(
            "/api/v1/users/auth/login",
            json={"email": email, "password": password},
        )
        data = resp.json()
        token = data.get("access_token", "")
        if token:
            break
        if "Too many login attempts" in data.get("detail", ""):
            await asyncio.sleep(5 * (attempt + 1))
            continue
        break
    assert token, f"Login failed: {data}"
    return {"Authorization": f"Bearer {token}"}


# ── Scaffolding helpers ──────────────────────────────────────────────────────


async def _create_boq(client: AsyncClient, auth: dict[str, str]) -> str:
    resp = await client.post(
        "/api/v1/projects/",
        json={
            "name": f"Round-Trip Project {uuid.uuid4().hex[:6]}",
            "description": "Project for BOQ round-trip tests",
            "region": "DACH",
            "classification_standard": "din276",
            "currency": "EUR",
        },
        headers=auth,
    )
    assert resp.status_code == 201, resp.text
    project_id = resp.json()["id"]

    resp = await client.post(
        "/api/v1/boq/boqs/",
        json={
            "project_id": project_id,
            "name": "Round-Trip BOQ",
            "description": "BOQ used to assert export/import round-trip fidelity",
        },
        headers=auth,
    )
    assert resp.status_code == 201, resp.text
    return resp.json()["id"]


async def _add_position(
    client: AsyncClient,
    auth: dict[str, str],
    boq_id: str,
    *,
    ordinal: str,
    description: str,
    unit: str = "m3",
    quantity: float = 10.0,
    unit_rate: float = 120.0,
) -> str:
    """Add one position; return its id."""
    resp = await client.post(
        f"/api/v1/boq/boqs/{boq_id}/positions/",
        json={
            "boq_id": boq_id,
            "ordinal": ordinal,
            "description": description,
            "unit": unit,
            "quantity": quantity,
            "unit_rate": unit_rate,
        },
        headers=auth,
    )
    assert resp.status_code == 201, resp.text
    return resp.json()["id"]


async def _positions(client: AsyncClient, auth: dict[str, str], boq_id: str) -> list[dict]:
    resp = await client.get(f"/api/v1/boq/boqs/{boq_id}", headers=auth)
    assert resp.status_code == 200, resp.text
    return resp.json()["positions"]


async def _import_csv(
    client: AsyncClient,
    auth: dict[str, str],
    boq_id: str,
    csv_body: str,
    *,
    delete_missing: bool = False,
) -> dict:
    url = f"/api/v1/boq/boqs/{boq_id}/import/excel/"
    if delete_missing:
        url += "?delete_missing=true"
    resp = await client.post(
        url,
        files={"file": ("roundtrip.csv", csv_body.encode("utf-8"), "text/csv")},
        headers=auth,
    )
    assert resp.status_code == 200, f"import failed: {resp.status_code} {resp.text[:400]}"
    return resp.json()


# ── Export stamps the Position ID column ─────────────────────────────────────


@pytest.mark.asyncio
async def test_export_excel_contains_position_id_column(
    shared_client: AsyncClient,
    shared_auth: dict[str, str],
) -> None:
    """The Excel export must carry a ``Position ID`` header and stamp each
    position's UUID in that column so a re-import can match on it."""
    from openpyxl import load_workbook

    boq_id = await _create_boq(shared_client, shared_auth)
    pid = await _add_position(shared_client, shared_auth, boq_id, ordinal="01.001", description="Concrete slab")

    resp = await shared_client.get(
        f"/api/v1/boq/boqs/{boq_id}/export/excel", headers=shared_auth, follow_redirects=True
    )
    assert resp.status_code == 200, resp.text

    wb = load_workbook(io.BytesIO(resp.content), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header = [str(c) if c is not None else "" for c in rows[0]]
    assert "Position ID" in header, f"header missing Position ID: {header}"
    id_col = header.index("Position ID")

    # The exported UUID must appear in the id column of some data row.
    stamped = {str(r[id_col]) for r in rows[1:] if id_col < len(r) and r[id_col]}
    assert pid in stamped, f"exported id column {stamped} missing position {pid}"
    wb.close()


# ── Update-in-place + create + foreign-id safety ─────────────────────────────


@pytest.mark.asyncio
async def test_reimport_updates_in_place_and_creates_and_flags_foreign(
    shared_client: AsyncClient,
    shared_auth: dict[str, str],
) -> None:
    boq_id = await _create_boq(shared_client, shared_auth)
    p1 = await _add_position(
        shared_client, shared_auth, boq_id, ordinal="01.001", description="Concrete", quantity=10, unit_rate=120
    )
    p2 = await _add_position(
        shared_client, shared_auth, boq_id, ordinal="01.002", description="Rebar", unit="kg", quantity=500, unit_rate=2
    )
    p3 = await _add_position(
        shared_client,
        shared_auth,
        boq_id,
        ordinal="01.003",
        description="Formwork",
        unit="m2",
        quantity=80,
        unit_rate=45,
    )
    foreign = str(uuid.uuid4())

    # p1 edited (qty + description), p2 identical (-> unchanged), p3 dropped,
    # one blank-id new row, one foreign-id row.
    csv_body = (
        "Position ID,Pos,Description,Unit,Quantity,Unit Rate\n"
        f"{p1},01.001,Concrete C30/37 raised,m3,25,120\n"
        f"{p2},01.002,Rebar,kg,500,2\n"
        ",01.010,Waterproofing membrane,m2,40,25\n"
        f"{foreign},09.999,Injected foreign row,m,5,10\n"
    )
    body = await _import_csv(shared_client, shared_auth, boq_id, csv_body)

    assert body["round_trip"] is True
    assert body["updated"] == 1, body
    assert body["unchanged"] == 1, body
    assert body["created"] == 2, body  # blank id + foreign id both become new
    assert body["deleted"] == 0
    assert body["would_delete"] == 1, body  # p3 dropped from the sheet
    # The foreign id is surfaced as a round-trip problem, not silently updated.
    assert any(p.get("issue") == "unknown_id" for p in body["problems"]), body["problems"]

    positions = await _positions(shared_client, shared_auth, boq_id)
    by_id = {p["id"]: p for p in positions}

    # p1 updated in place (same id), p2 untouched, p3 still present.
    assert by_id[p1]["description"] == "Concrete C30/37 raised"
    assert float(by_id[p1]["quantity"]) == 25.0
    assert by_id[p2]["description"] == "Rebar"
    assert p3 in by_id, "dropped position must survive without delete_missing"
    assert by_id[p3]["description"] == "Formwork"

    # Two brand-new positions exist; neither reused an existing id.
    new_descs = {p["description"] for p in positions if p["id"] not in {p1, p2, p3}}
    assert "Waterproofing membrane" in new_descs
    assert "Injected foreign row" in new_descs
    # Exactly one position matches the original three ids each (no duplication).
    assert sum(1 for p in positions if p["id"] == p1) == 1


# ── Opt-in delete of positions dropped from the sheet ────────────────────────


@pytest.mark.asyncio
async def test_delete_missing_opt_in_removes_dropped_rows(
    shared_client: AsyncClient,
    shared_auth: dict[str, str],
) -> None:
    boq_id = await _create_boq(shared_client, shared_auth)
    keep = await _add_position(
        shared_client, shared_auth, boq_id, ordinal="01.001", description="Keep me", quantity=10, unit_rate=100
    )
    drop = await _add_position(
        shared_client, shared_auth, boq_id, ordinal="01.002", description="Drop me", quantity=5, unit_rate=50
    )

    # Sheet carries only ``keep``; ``drop`` is absent.
    csv_body = f"Position ID,Pos,Description,Unit,Quantity,Unit Rate\n{keep},01.001,Keep me,m3,10,100\n"

    # Without the flag: reported but not removed.
    body = await _import_csv(shared_client, shared_auth, boq_id, csv_body)
    assert body["would_delete"] == 1
    assert body["deleted"] == 0
    assert drop in {p["id"] for p in await _positions(shared_client, shared_auth, boq_id)}

    # With the flag: the dropped position is removed.
    body = await _import_csv(shared_client, shared_auth, boq_id, csv_body, delete_missing=True)
    assert body["deleted"] == 1, body
    remaining = {p["id"] for p in await _positions(shared_client, shared_auth, boq_id)}
    assert drop not in remaining
    assert keep in remaining


# ── Cross-BOQ id can never mutate a position in another BOQ ───────────────────


@pytest.mark.asyncio
async def test_foreign_boq_id_cannot_update_across_boundary(
    shared_client: AsyncClient,
    shared_auth: dict[str, str],
) -> None:
    boq_a = await _create_boq(shared_client, shared_auth)
    boq_b = await _create_boq(shared_client, shared_auth)
    pa = await _add_position(
        shared_client, shared_auth, boq_a, ordinal="01.001", description="A original", quantity=7, unit_rate=70
    )
    await _add_position(
        shared_client, shared_auth, boq_b, ordinal="01.001", description="B original", quantity=3, unit_rate=30
    )

    # Re-import into BOQ B a row carrying BOQ A's position id, trying to
    # overwrite A's description via the round trip.
    csv_body = f"Position ID,Pos,Description,Unit,Quantity,Unit Rate\n{pa},02.500,HIJACK ATTEMPT,m3,999,999\n"
    body = await _import_csv(shared_client, shared_auth, boq_b, csv_body)

    # The foreign id landed as a NEW row in B, flagged - never an update.
    assert body["updated"] == 0, body
    assert body["created"] == 1, body
    assert any(p.get("issue") == "unknown_id" for p in body["problems"]), body["problems"]

    # BOQ A's position is completely untouched.
    a_positions = {p["id"]: p for p in await _positions(shared_client, shared_auth, boq_a)}
    assert a_positions[pa]["description"] == "A original"
    assert float(a_positions[pa]["quantity"]) == 7.0
