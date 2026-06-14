"""Downloadable export formats for generated reports (PDF / XLSX / CSV).

Background
==========
Until this module existed the reporting feature could only hand back an
HTML body (``GET /reports/{id}/content`` returns ``text/html``). The
generated-report row carried a ``format`` column of ``pdf`` / ``excel`` /
``html`` but nothing ever produced a real binary in those formats - a user
who picked "Excel" still only had HTML to look at. This file closes that
gap by turning a report's ``data_snapshot`` (the same per-section dict the
:class:`~app.modules.reporting.renderer.ReportRenderer` consumes) into a
real downloadable file.

Design notes
============
- **Reuse, do not reinvent.** The PDF path uses the *same* reportlab
  platypus stack and the bundled Unicode fonts that the rest of the
  platform already ships (see ``app.modules.boq.pdf_export`` and
  ``app.core.pdf_fonts``). No second PDF library is introduced. The XLSX
  path uses ``openpyxl`` (already a dependency, used by the BOQ export at
  ``backend/app/modules/boq/router.py``). CSV uses the stdlib ``csv``
  module. Nothing heavy is added.
- **CSV / spreadsheet injection.** Every user-controlled string written
  into a CSV or XLSX cell is routed through
  :func:`app.core.csv_safety.neutralise_formula`, exactly as the BOQ
  exporters do.
- **Money stays Decimal-correct.** The snapshot already carries money as
  strings (e.g. ``"517103508.65 EUR"``) assembled by the service layer's
  ``_build_default_snapshot`` / retainage roll-ups. We never coerce those
  through ``float`` for display - they are emitted verbatim. Where a value
  is a bare numeric string we hand openpyxl a :class:`decimal.Decimal` so
  Excel stores it as a sortable number without precision loss.
- **One currency per report.** The resolved ISO 4217 code is stamped on the
  report row (``GeneratedReport.currency``) and into the snapshot; the
  exporters surface it in the header so every figure reads in one currency.

Public API
==========

>>> from app.modules.reporting.exporters import export_report
>>> filename, media_type, blob = export_report(
...     fmt="xlsx",
...     report_type="cost_report",
...     title="Q1 Cost Report",
...     project_name="Skyline Tower",
...     currency="EUR",
...     generated_at="2026-06-14T10:00:00",
...     template_data={"sections": [...]},
...     data_snapshot={"summary": {...}, "breakdown": [...]},
... )

The function is sync and pure - no DB, no network, no clock. The service
layer assembles the snapshot and resolves the currency before calling it.
"""

from __future__ import annotations

import csv
import html
import io
from decimal import Decimal, InvalidOperation
from typing import Any

from app.core.csv_safety import neutralise_formula

__all__ = [
    "ExportFormatError",
    "SUPPORTED_FORMATS",
    "export_report",
]

# Formats this module can produce. ``html`` is included so the download
# endpoint can serve the existing HTML body through the same code path and
# nothing regresses, but the heavy lifting here is pdf / xlsx / csv.
SUPPORTED_FORMATS: tuple[str, ...] = ("pdf", "xlsx", "csv", "html")

# MIME types keyed by format. Matches the BOQ export endpoints
# (``backend/app/modules/boq/router.py``) so the browser downloads a real
# file rather than rendering it inline.
_MEDIA_TYPES: dict[str, str] = {
    "pdf": "application/pdf",
    "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "csv": "text/csv; charset=utf-8",
    "html": "text/html; charset=utf-8",
}


class ExportFormatError(ValueError):
    """Raised when an unsupported export format is requested."""


# ── Section resolution ────────────────────────────────────────────────────
#
# Mirror the renderer's precedence so a downloaded file shows exactly the
# sections the HTML view shows: template sections first, else the built-in
# default list for the report type, else a single generic Summary block.


def _resolve_sections(
    report_type: str,
    template_data: dict[str, Any] | None,
) -> list[dict[str, Any]]:
    """Return the ordered section list to render (template > default > generic).

    Imports the renderer's default-section map so the two stay in lockstep -
    we deliberately do not maintain a second copy of the section ordering.
    """
    if isinstance(template_data, dict):
        sections = template_data.get("sections")
        if isinstance(sections, list) and sections:
            resolved = [s for s in sections if isinstance(s, dict) and s.get("id")]
            if resolved:
                return resolved

    from app.modules.reporting.renderer import _DEFAULT_SECTIONS

    return _DEFAULT_SECTIONS.get(report_type, [{"id": "summary", "title": "Summary"}])


def _section_title(section: dict[str, Any]) -> str:
    """Human-readable section title, falling back to a Title-Cased id."""
    sid = str(section.get("id", "")).strip()
    return str(section.get("title", sid.replace("_", " ").title()))


def _stringify(value: Any) -> str:
    """Stringify a scalar snapshot value for a flat text cell.

    Booleans become Yes/No (matching the renderer); ``None`` becomes an
    empty string. Dicts / lists are summarised compactly because flat CSV /
    PDF table cells cannot nest a sub-table the way the HTML renderer can.
    """
    if value is None:
        return ""
    if isinstance(value, bool):
        return "Yes" if value else "No"
    if isinstance(value, dict):
        # Compact "k: v" join so a nested object still conveys its content
        # in a single cell rather than printing a Python repr.
        return "; ".join(f"{_humanize_key(k)}: {_stringify(v)}" for k, v in value.items())
    if isinstance(value, list):
        return "; ".join(_stringify(v) for v in value)
    return str(value)


def _humanize_key(key: Any) -> str:
    """Title-case a snake_case snapshot key for a label cell."""
    return str(key).replace("_", " ").title()


def _looks_numeric(text: str) -> bool:
    """True when *text* is a bare number openpyxl can store as a Decimal.

    A money string like ``"1234.56 EUR"`` is intentionally NOT numeric here
    (it carries a currency suffix and should stay text); only a clean
    ``"1234.56"`` / ``"-3"`` qualifies.
    """
    if not text:
        return False
    try:
        d = Decimal(text)
    except (InvalidOperation, ValueError):
        return False
    return d.is_finite()


# ── Section shape detection ─────────────────────────────────────────────────
#
# A section payload renders as one of two table shapes (mirroring the HTML
# renderer): a key/value definition table (dict, or list of scalars) or a
# columnar table (list of dicts, e.g. a cost breakdown / incident log). The
# helpers below classify the payload and materialise the matching shape so
# every export format presents the same structure.


def _is_record_list(payload: Any) -> bool:
    """True when *payload* is a non-empty list whose items are all dicts.

    This is the "columnar table" shape (e.g. ``breakdown``: a list of
    ``{"trade": ..., "amount": ...}`` records).
    """
    return isinstance(payload, list) and len(payload) > 0 and all(isinstance(it, dict) for it in payload)


def _record_columns(records: list[dict[str, Any]]) -> list[str]:
    """Ordered union of keys across record dicts (first-seen order)."""
    columns: list[str] = []
    seen: set[str] = set()
    for item in records:
        for key in item:
            if key not in seen:
                seen.add(key)
                columns.append(key)
    return columns


def _flatten_keyvalue(payload: Any) -> list[tuple[str, str]]:
    """Flatten a dict / scalar / scalar-list into ``(label, value)`` rows.

    Money / Decimal values pass through ``_stringify`` unchanged so no float
    rounding is ever introduced. Used for the key/value table shape; record
    lists are handled separately via :func:`_record_columns`.
    """
    rows: list[tuple[str, str]] = []
    if isinstance(payload, dict):
        for key, value in payload.items():
            rows.append((_humanize_key(key), _stringify(value)))
    elif isinstance(payload, list):
        for item in payload:
            rows.append(("", _stringify(item)))
    else:
        rows.append(("", _stringify(payload)))
    return rows


def _section_is_empty(payload: Any) -> bool:
    """Mirror the renderer's skip rule: None / empty dict / empty list."""
    if payload is None:
        return True
    return bool(isinstance(payload, dict | list) and not payload)


# ── CSV export ──────────────────────────────────────────────────────────────


def _export_csv(
    *,
    report_type: str,
    title: str,
    project_name: str,
    currency: str,
    generated_at: str,
    template_data: dict[str, Any] | None,
    data_snapshot: dict[str, Any] | None,
) -> bytes:
    """Render the report snapshot as a CSV file.

    Layout: a small metadata preamble, then for each populated section a
    blank separator, a section-title row, and ``Field,Value`` rows. Every
    user-controlled string is neutralised against spreadsheet formula
    injection.
    """
    snapshot = data_snapshot or {}
    buffer = io.StringIO()
    writer = csv.writer(buffer)

    # Metadata preamble - first-party labels, dynamic values neutralised.
    writer.writerow(["Report", neutralise_formula(title)])
    writer.writerow(["Project", neutralise_formula(project_name)])
    writer.writerow(["Type", neutralise_formula(report_type)])
    if currency:
        writer.writerow(["Currency", neutralise_formula(currency)])
    writer.writerow(["Generated", neutralise_formula(generated_at)])

    rendered_any = False
    for section in _resolve_sections(report_type, template_data):
        sid = str(section.get("id", "")).strip()
        payload = snapshot.get(sid)
        if _section_is_empty(payload):
            continue
        rendered_any = True
        writer.writerow([])
        writer.writerow([neutralise_formula(_section_title(section))])
        if _is_record_list(payload):
            # Columnar table (e.g. cost breakdown): one header row of the
            # union of keys, then a row per record. Every cell neutralised.
            columns = _record_columns(payload)
            writer.writerow([neutralise_formula(_humanize_key(c)) for c in columns])
            for item in payload:
                writer.writerow([neutralise_formula(_stringify(item.get(c))) for c in columns])
        else:
            writer.writerow(["Field", "Value"])
            for label, value in _flatten_keyvalue(payload):
                writer.writerow([neutralise_formula(label), neutralise_formula(value)])

    if not rendered_any:
        writer.writerow([])
        writer.writerow(["No data available"])

    return buffer.getvalue().encode("utf-8-sig")


# ── XLSX export ─────────────────────────────────────────────────────────────


def _export_xlsx(
    *,
    report_type: str,
    title: str,
    project_name: str,
    currency: str,
    generated_at: str,
    template_data: dict[str, Any] | None,
    data_snapshot: dict[str, Any] | None,
) -> bytes:
    """Render the report snapshot as a formatted .xlsx workbook.

    One worksheet: a metadata header block, then a two-column
    ``Field | Value`` table per populated section with a shaded section
    heading. Bare numeric values are written as :class:`Decimal` so Excel
    stores a sortable number; everything else is neutralised text.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    snapshot = data_snapshot or {}

    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    bold = Font(bold=True)
    title_font = Font(bold=True, size=14)
    section_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    section_fill = PatternFill(start_color="E8E8EE", end_color="E8E8EE", fill_type="solid")
    wrap = Alignment(wrap_text=True, vertical="top")

    row = 1
    # ── Title + metadata block ──
    cell = ws.cell(row=row, column=1, value=neutralise_formula(title))
    cell.font = title_font
    row += 1

    meta_rows = [
        ("Project", project_name),
        ("Type", report_type),
    ]
    if currency:
        meta_rows.append(("Currency", currency))
    meta_rows.append(("Generated", generated_at))
    for label, value in meta_rows:
        lbl = ws.cell(row=row, column=1, value=label)
        lbl.font = bold
        ws.cell(row=row, column=2, value=neutralise_formula(value))
        row += 1
    row += 1  # blank spacer

    rendered_any = False
    for section in _resolve_sections(report_type, template_data):
        sid = str(section.get("id", "")).strip()
        payload = snapshot.get(sid)
        if _section_is_empty(payload):
            continue
        rendered_any = True

        if _is_record_list(payload):
            columns = _record_columns(payload)
            n_cols = max(len(columns), 1)

            # Section heading shaded across the table width.
            sec_cell = ws.cell(row=row, column=1, value=neutralise_formula(_section_title(section)))
            sec_cell.font = section_font
            sec_cell.fill = section_fill
            for c in range(2, n_cols + 1):
                ws.cell(row=row, column=c).fill = section_fill
            row += 1

            # Header row (one cell per column).
            for c_idx, col in enumerate(columns, start=1):
                hc = ws.cell(row=row, column=c_idx, value=neutralise_formula(_humanize_key(col)))
                hc.font = header_font
                hc.fill = header_fill
            row += 1

            for item in payload:
                for c_idx, col in enumerate(columns, start=1):
                    cell_val = _stringify(item.get(col))
                    if _looks_numeric(cell_val):
                        ws.cell(row=row, column=c_idx, value=Decimal(cell_val))
                    else:
                        vc = ws.cell(row=row, column=c_idx, value=neutralise_formula(cell_val))
                        vc.alignment = wrap
                row += 1
            row += 1  # blank spacer between sections
            continue

        # Key/value table shape.
        sec_cell = ws.cell(row=row, column=1, value=neutralise_formula(_section_title(section)))
        sec_cell.font = section_font
        sec_cell.fill = section_fill
        ws.cell(row=row, column=2).fill = section_fill
        row += 1

        h1 = ws.cell(row=row, column=1, value="Field")
        h1.font = header_font
        h1.fill = header_fill
        h2 = ws.cell(row=row, column=2, value="Value")
        h2.font = header_font
        h2.fill = header_fill
        row += 1

        for label, value in _flatten_keyvalue(payload):
            ws.cell(row=row, column=1, value=neutralise_formula(label))
            # Store clean numbers as Decimal so Excel treats them as numeric
            # (sortable, summable) without the lossy float roundtrip; money
            # strings with a currency suffix stay text.
            if _looks_numeric(value):
                ws.cell(row=row, column=2, value=Decimal(value))
            else:
                vcell = ws.cell(row=row, column=2, value=neutralise_formula(value))
                vcell.alignment = wrap
            row += 1
        row += 1  # blank spacer between sections

    if not rendered_any:
        ws.cell(row=row, column=1, value="No data available").font = bold

    # Reasonable column widths so the file opens readable.
    ws.column_dimensions[get_column_letter(1)].width = 32
    ws.column_dimensions[get_column_letter(2)].width = 60
    ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# ── PDF export ────────────────────────────────────────────────────────────────


def _export_pdf(
    *,
    report_type: str,
    title: str,
    project_name: str,
    currency: str,
    generated_at: str,
    template_data: dict[str, Any] | None,
    data_snapshot: dict[str, Any] | None,
) -> bytes:
    """Render an executive-summary PDF using the platform reportlab stack.

    Reuses ``app.core.pdf_fonts`` (the same bundled DejaVu Unicode faces the
    BOQ PDF export uses) and reportlab platypus. The layout is a titled
    cover header followed by one ``Field | Value`` table per populated
    section - a clean, correct executive summary rather than a faked binary.
    """
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_LEFT
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    from app.core.pdf_fonts import BODY_FONT, BOLD_FONT, register_pdf_fonts

    register_pdf_fonts()

    snapshot = data_snapshot or {}
    base = getSampleStyleSheet()
    styles = {
        "title": ParagraphStyle(
            "RepTitle",
            parent=base["Normal"],
            fontName=BOLD_FONT,
            fontSize=18,
            textColor=colors.HexColor("#16213e"),
            spaceAfter=4 * mm,
        ),
        "meta": ParagraphStyle(
            "RepMeta",
            parent=base["Normal"],
            fontName=BODY_FONT,
            fontSize=9,
            textColor=colors.HexColor("#555555"),
            spaceAfter=1 * mm,
        ),
        "section": ParagraphStyle(
            "RepSection",
            parent=base["Normal"],
            fontName=BOLD_FONT,
            fontSize=13,
            textColor=colors.HexColor("#1d4ed8"),
            spaceBefore=6 * mm,
            spaceAfter=2 * mm,
        ),
        "label": ParagraphStyle(
            "RepLabel",
            parent=base["Normal"],
            fontName=BOLD_FONT,
            fontSize=9,
            textColor=colors.HexColor("#333333"),
            leading=12,
        ),
        "value": ParagraphStyle(
            "RepValue",
            parent=base["Normal"],
            fontName=BODY_FONT,
            fontSize=9,
            textColor=colors.HexColor("#111111"),
            alignment=TA_LEFT,
            leading=12,
        ),
    }

    def _p(text: Any, style_key: str) -> Paragraph:
        """Escape arbitrary text and wrap it in a paragraph (XSS/markup safe).

        Mirrors ``boq.pdf_export._safe_para``: reportlab's Paragraph parses
        a subset of HTML, so user-controlled strings MUST be escaped or a
        payload like ``<font color="white">`` would render / a malformed tag
        would crash paraparser.
        """
        rendered = "" if text is None else str(text)
        return Paragraph(html.escape(rendered, quote=True), styles[style_key])

    flowables: list[Any] = []
    flowables.append(_p(title, "title"))
    flowables.append(_p(f"Project: {project_name}", "meta"))
    meta_line = f"Type: {report_type}"
    if currency:
        meta_line += f"  -  Currency: {currency}"
    flowables.append(_p(meta_line, "meta"))
    flowables.append(_p(f"Generated: {generated_at}", "meta"))
    flowables.append(Spacer(1, 4 * mm))

    usable_width = A4[0] - 40 * mm
    col_widths = [usable_width * 0.35, usable_width * 0.65]

    rendered_any = False
    for section in _resolve_sections(report_type, template_data):
        sid = str(section.get("id", "")).strip()
        payload = snapshot.get(sid)
        if _section_is_empty(payload):
            continue
        rendered_any = True
        flowables.append(_p(_section_title(section), "section"))

        if _is_record_list(payload):
            columns = _record_columns(payload)
            header_cells = [_p(_humanize_key(c), "label") for c in columns]
            table_rows = [header_cells]
            for item in payload:
                table_rows.append([_p(_stringify(item.get(c)), "value") for c in columns])
            n_cols = max(len(columns), 1)
            rec_col_widths = [usable_width / n_cols] * n_cols
            table = Table(table_rows, colWidths=rec_col_widths, repeatRows=1)
            table.setStyle(
                TableStyle(
                    [
                        ("VALIGN", (0, 0), (-1, -1), "TOP"),
                        ("TOPPADDING", (0, 0), (-1, -1), 1.5 * mm),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 1.5 * mm),
                        ("LEFTPADDING", (0, 0), (-1, -1), 2 * mm),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 2 * mm),
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a1a2e")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                        ("LINEBELOW", (0, 0), (-1, -1), 0.4, colors.HexColor("#e5e7eb")),
                        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8f8fb")]),
                    ]
                )
            )
            flowables.append(table)
            continue

        table_rows = []
        for label, value in _flatten_keyvalue(payload):
            table_rows.append([_p(label, "label"), _p(value, "value")])

        if not table_rows:
            continue
        table = Table(table_rows, colWidths=col_widths)
        table.setStyle(
            TableStyle(
                [
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("TOPPADDING", (0, 0), (-1, -1), 1.5 * mm),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 1.5 * mm),
                    ("LEFTPADDING", (0, 0), (-1, -1), 2 * mm),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 2 * mm),
                    ("LINEBELOW", (0, 0), (-1, -1), 0.4, colors.HexColor("#e5e7eb")),
                    ("ROWBACKGROUNDS", (0, 0), (-1, -1), [colors.white, colors.HexColor("#f8f8fb")]),
                ]
            )
        )
        flowables.append(table)

    if not rendered_any:
        flowables.append(_p("No data available", "section"))
        flowables.append(
            _p(
                "This report was generated with an empty data snapshot. Verify that "
                "the source modules returned data for the selected project.",
                "value",
            )
        )

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=20 * mm,
        rightMargin=20 * mm,
        topMargin=20 * mm,
        bottomMargin=18 * mm,
        title=title,
        author="OpenConstructionERP",
        subject="Project Report",
        creator="OpenConstructionERP - DataDrivenConstruction",
    )
    doc.build(flowables)
    return buffer.getvalue()


# ── Public entry point ──────────────────────────────────────────────────────


def export_report(
    *,
    fmt: str,
    report_type: str,
    title: str,
    project_name: str,
    currency: str,
    generated_at: str,
    template_data: dict[str, Any] | None,
    data_snapshot: dict[str, Any] | None,
    html_body: str | None = None,
) -> tuple[str, str, bytes]:
    """Render a generated report into a downloadable file.

    Args:
        fmt: One of :data:`SUPPORTED_FORMATS` (``pdf`` / ``xlsx`` / ``csv`` /
            ``html``).
        report_type: Report type token (drives the default section list).
        title: Report title (already schema-sanitised; re-escaped here).
        project_name: Owning project's display name.
        currency: Resolved ISO 4217 code stamped on the report row.
        generated_at: ISO timestamp shown in the file header.
        template_data: Bound template's ``template_data`` (optional).
        data_snapshot: Per-section payload dict.
        html_body: Pre-rendered HTML body, only used for ``fmt="html"`` so
            the existing HTML output is served unchanged. When ``None`` and
            ``fmt="html"`` the HTML is rendered fresh from the snapshot.

    Returns:
        ``(suggested_filename, media_type, file_bytes)``.

    Raises:
        ExportFormatError: when *fmt* is not supported.
    """
    fmt = (fmt or "").strip().lower()
    if fmt not in SUPPORTED_FORMATS:
        raise ExportFormatError(f"Unsupported export format '{fmt}'. Expected one of: {', '.join(SUPPORTED_FORMATS)}.")

    if fmt == "csv":
        blob = _export_csv(
            report_type=report_type,
            title=title,
            project_name=project_name,
            currency=currency,
            generated_at=generated_at,
            template_data=template_data,
            data_snapshot=data_snapshot,
        )
    elif fmt == "xlsx":
        blob = _export_xlsx(
            report_type=report_type,
            title=title,
            project_name=project_name,
            currency=currency,
            generated_at=generated_at,
            template_data=template_data,
            data_snapshot=data_snapshot,
        )
    elif fmt == "pdf":
        blob = _export_pdf(
            report_type=report_type,
            title=title,
            project_name=project_name,
            currency=currency,
            generated_at=generated_at,
            template_data=template_data,
            data_snapshot=data_snapshot,
        )
    else:  # html
        if html_body is not None:
            blob = html_body.encode("utf-8")
        else:
            from app.modules.reporting.renderer import ReportRenderer

            blob = (
                ReportRenderer()
                .render_html(
                    report_type=report_type,
                    title=title,
                    project_name=project_name,
                    template_data=template_data,
                    data_snapshot=data_snapshot,
                    generated_at=generated_at,
                )
                .encode("utf-8")
            )

    filename = f"{_safe_filename(title)}.{fmt}"
    return filename, _MEDIA_TYPES[fmt], blob


def _safe_filename(title: str) -> str:
    """ASCII-safe, quote-free base filename derived from the report title.

    Mirrors the BOQ export filename handling: non-ASCII is replaced and
    double quotes are swapped for single quotes so the value is safe inside
    a ``Content-Disposition: attachment; filename="..."`` header. Falls back
    to ``report`` when the title reduces to nothing.
    """
    base = (title or "").encode("ascii", errors="replace").decode("ascii").replace('"', "'").strip()
    # Collapse path separators that would confuse some download clients.
    base = base.replace("/", "-").replace("\\", "-")
    return base or "report"
