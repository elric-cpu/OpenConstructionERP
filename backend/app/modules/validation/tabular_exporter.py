# DDC-CWICR-OE: DataDrivenConstruction · OpenConstructionERP
# Copyright (c) 2026 Artem Boiko / DataDrivenConstruction
"""CSV / XLSX exporters for OpenConstructionERP validation reports.

Sibling to :mod:`app.modules.validation.sarif_exporter`. Where SARIF targets
code-scanning pipelines, these two formats target the people who live in
spreadsheets - QA leads, cost managers and clients who want the validation
findings as a flat table they can sort, filter and attach to an email.

Design (reuse, do not reinvent):
    * **One row mapping.** Both formats - and the SARIF export - share the
      same per-result projection produced by
      :func:`app.modules.validation.sarif_exporter._normalize_report`, so a
      finding reads identically across CSV, XLSX and SARIF. The flat row
      shape lives in :func:`report_to_rows`.
    * **Formula-injection safe.** Every user-controlled cell (rule name,
      message, suggestion, element ref, ...) is routed through
      :func:`app.core.csv_safety.neutralise_formula`, exactly as the BOQ and
      reporting exporters do. A description like ``=cmd|'/c calc'!A0`` is
      written as literal text, never an executable formula.
    * **openpyxl + stdlib csv only.** ``openpyxl`` is already a base
      dependency (used by the BOQ export); CSV uses the stdlib. Nothing new
      is added.

Public API:
    * :func:`report_to_rows`   - report -> (headers, list[row]) projection.
    * :func:`report_to_csv`    - report -> CSV bytes (UTF-8 BOM).
    * :func:`report_to_xlsx`   - report -> .xlsx workbook bytes.

All three are sync and pure - no DB, no network, no clock - so they are
directly unit-testable. The router layer does the IDOR-guarded report load
and wraps the bytes in an HTTP response.
"""

from __future__ import annotations

import csv
import io
from datetime import datetime
from typing import Any

from app.core.csv_safety import neutralise_formula
from app.modules.validation.sarif_exporter import _normalize_report

__all__ = [
    "ROW_HEADERS",
    "report_to_csv",
    "report_to_rows",
    "report_to_xlsx",
]

# Column order for the flat findings table. Kept as a module constant so the
# CSV header, the XLSX header and the unit tests all agree on one layout.
ROW_HEADERS: tuple[str, ...] = (
    "rule_id",
    "rule_name",
    "severity",
    "status",
    "category",
    "message",
    "element_ref",
    "suggestion",
)


def _status_for(item: dict[str, Any]) -> str:
    """Derive the human ``status`` cell for a normalised result.

    A passing check reads ``passed``; a failing one reads its severity
    (``error`` / ``warning`` / ``info``) so a reader scanning the column sees
    the actionable state directly, matching the on-screen badge text.
    """
    if item.get("passed", False):
        return "passed"
    severity = str(item.get("severity", "") or "").lower()
    return severity or "warning"


def report_to_rows(report: Any) -> tuple[list[str], list[list[str]]]:
    """Project a validation report into a flat ``(headers, rows)`` table.

    Accepts either an in-memory :class:`EngineReport` or a persisted ORM
    :class:`app.modules.validation.models.ValidationReport` row - both are
    normalised through the SARIF exporter's ``_normalize_report`` so the row
    mapping is shared by every export format.

    Returns:
        ``(headers, rows)`` where ``headers`` is :data:`ROW_HEADERS` as a
        list and each row is a list of plain strings in that column order.
        Cells are NOT yet formula-neutralised - that happens at the CSV /
        XLSX write boundary so this projection stays a faithful view of the
        data for tests and other consumers.
    """
    norm = _normalize_report(report)
    rows: list[list[str]] = []
    for item in norm["results"]:
        rows.append(
            [
                str(item.get("rule_id", "") or ""),
                str(item.get("rule_name", "") or item.get("rule_id", "") or ""),
                str(item.get("severity", "") or ""),
                _status_for(item),
                str(item.get("category", "") or ""),
                str(item.get("message", "") or ""),
                str(item.get("element_ref", "") or "") if item.get("element_ref") else "",
                str(item.get("suggestion", "") or "") if item.get("suggestion") else "",
            ]
        )
    return list(ROW_HEADERS), rows


def _meta_pairs(report: Any) -> list[tuple[str, str]]:
    """Build the small ``(label, value)`` metadata preamble for a report."""
    norm = _normalize_report(report)
    timestamp = norm.get("timestamp")
    if isinstance(timestamp, datetime):
        ts = timestamp.isoformat()
    else:
        ts = str(timestamp or "")
    rule_sets = norm.get("rule_sets") or []
    return [
        ("Report ID", str(norm.get("id", "") or "")),
        ("Target type", str(norm.get("target_type", "") or "")),
        ("Target ID", str(norm.get("target_id", "") or "")),
        ("Rule sets", "+".join(str(s) for s in rule_sets)),
        ("Generated", ts),
        ("Total checks", str(len(norm.get("results") or []))),
    ]


# ── CSV ──────────────────────────────────────────────────────────────────────


def report_to_csv(report: Any) -> bytes:
    """Render a validation report as CSV bytes (UTF-8 with BOM).

    Layout: a short metadata preamble, a blank separator, the findings
    header, then one row per result. Every cell - preamble values and result
    cells alike - is passed through :func:`neutralise_formula` so a value
    beginning with ``= + - @`` (or a leading tab / CR) can never execute as a
    spreadsheet formula when the file is opened.

    The UTF-8 BOM (``utf-8-sig``) makes Excel detect the encoding so non-ASCII
    rule messages render correctly on a double-click open.
    """
    headers, rows = report_to_rows(report)
    buffer = io.StringIO(newline="")
    writer = csv.writer(buffer)

    for label, value in _meta_pairs(report):
        # Static first-party label; dynamic value neutralised.
        writer.writerow([label, neutralise_formula(value)])
    writer.writerow([])

    writer.writerow([neutralise_formula(h) for h in headers])
    for row in rows:
        writer.writerow([neutralise_formula(cell) for cell in row])

    return buffer.getvalue().encode("utf-8-sig")


# ── XLSX ─────────────────────────────────────────────────────────────────────


def report_to_xlsx(report: Any) -> bytes:
    """Render a validation report as a formatted .xlsx workbook (bytes).

    One worksheet: a metadata header block, a blank spacer, a shaded header
    row, then one row per finding. Every user-controlled cell is neutralised
    against formula injection (identically to the CSV path). Severity cells on
    failing rows are lightly tinted so errors/warnings stand out, mirroring the
    on-screen badges.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    headers, rows = report_to_rows(report)

    wb = Workbook()
    ws = wb.active
    ws.title = "Validation"

    bold = Font(bold=True)
    title_font = Font(bold=True, size=14)
    header_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    wrap = Alignment(wrap_text=True, vertical="top")
    # Light severity tints for failing rows (error / warning / info).
    severity_fills = {
        "error": PatternFill(start_color="FDE2E1", end_color="FDE2E1", fill_type="solid"),
        "warning": PatternFill(start_color="FCEFC7", end_color="FCEFC7", fill_type="solid"),
        "info": PatternFill(start_color="DCEAFB", end_color="DCEAFB", fill_type="solid"),
    }
    severity_col = headers.index("severity") + 1 if "severity" in headers else None
    status_col = headers.index("status") + 1 if "status" in headers else None

    row_idx = 1
    title_cell = ws.cell(row=row_idx, column=1, value="Validation findings")
    title_cell.font = title_font
    row_idx += 1

    for label, value in _meta_pairs(report):
        lbl = ws.cell(row=row_idx, column=1, value=label)
        lbl.font = bold
        ws.cell(row=row_idx, column=2, value=neutralise_formula(value))
        row_idx += 1
    row_idx += 1  # blank spacer

    header_row = row_idx
    for c_idx, head in enumerate(headers, start=1):
        hc = ws.cell(row=row_idx, column=c_idx, value=neutralise_formula(head))
        hc.font = header_font
        hc.fill = header_fill
    row_idx += 1

    for row in rows:
        for c_idx, cell in enumerate(row, start=1):
            wc = ws.cell(row=row_idx, column=c_idx, value=neutralise_formula(cell))
            wc.alignment = wrap
        # Tint the row's severity + status cells when this is a failing finding
        # so errors/warnings stand out at a glance.
        status_val = row[headers.index("status")] if "status" in headers else ""
        if status_val and status_val != "passed":
            fill = severity_fills.get(status_val)
            if fill is not None:
                if severity_col:
                    ws.cell(row=row_idx, column=severity_col).fill = fill
                if status_col:
                    ws.cell(row=row_idx, column=status_col).fill = fill
        row_idx += 1

    if not rows:
        ws.cell(row=row_idx, column=1, value="No findings").font = bold

    # Column widths tuned so the file opens readable; message/suggestion wide.
    widths = {
        "rule_id": 28,
        "rule_name": 28,
        "severity": 12,
        "status": 12,
        "category": 16,
        "message": 60,
        "element_ref": 22,
        "suggestion": 50,
    }
    for c_idx, head in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(c_idx)].width = widths.get(head, 18)
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1).coordinate

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
