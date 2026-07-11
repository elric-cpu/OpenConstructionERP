# DDC-CWICR-OE: DataDrivenConstruction · OpenConstructionERP
# Copyright (c) 2026 Artem Boiko / DataDrivenConstruction
"""Bill-of-Quantities (BOQ) workbook builder.

The single most-requested BIM-to-estimating handoff: take the measured
quantities an IFC / RVT model already carries on every element (areas,
volumes, lengths, weights) straight into one Excel file a quantity
surveyor can price. This is the spreadsheet a site quantity-takeoff
person opens, not a CAFM handover (that is COBie, see ``cobie.py``).

The workbook has two sheets:

    BOQ      - one row per *group* (by element type, storey, discipline,
               or type x storey) with the element count and the summed
               quantities, plus a TOTAL row.
    Elements - one row per element with its individual quantities, so the
               summary numbers are auditable back to the source elements.

Source of truth
---------------
Everything is a PROJECTION of data already held on ``BIMElement``:
``element_type`` / ``storey`` / ``discipline`` for grouping, and the
``quantities`` JSON blob for the numbers. No SQL is issued here - the
service loads the (optionally filtered) elements and hands them in, so
this module stays a pure, snapshot-testable function.

Quantity keys
-------------
Converters are not perfectly consistent about quantity naming - some
write the canonical ``area_m2`` / ``volume_m3`` and some write raw IFC
BaseQuantities (``NetVolume``, ``GrossArea``). Each output column scans a
small alias list and takes the first numeric value present, so both
shapes land in the right column.

Determinism
-----------
Group rows are sorted by their group key and detail rows by a stable
composite key, so the VALUES are reproducible for a given input (the
bytes are not, because openpyxl embeds build metadata - the tests assert
on values, not bytes, exactly like the COBie tests).
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import UTC, datetime
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

# ── Quantity columns ─────────────────────────────────────────────────────
# (canonical key, human label, unit, alias keys scanned in order).
# The first alias with a numeric value wins, so both our canonical
# ``area_m2`` and a raw IFC ``NetArea`` map to the Area column.
QUANTITY_FIELDS: list[tuple[str, str, str, tuple[str, ...]]] = [
    ("area_m2", "Area", "m2", ("area_m2", "area", "NetArea", "GrossArea", "net_area", "gross_area")),
    (
        "volume_m3",
        "Volume",
        "m3",
        ("volume_m3", "volume", "NetVolume", "GrossVolume", "net_volume", "gross_volume"),
    ),
    ("length_m", "Length", "m", ("length_m", "length", "Length")),
    ("weight_kg", "Weight", "kg", ("weight_kg", "weight", "Weight", "mass", "Mass")),
]

# group_by tokens accepted by the builder; anything else falls back to
# element_type so a bad query param can never raise.
GROUP_BY_CHOICES = frozenset({"element_type", "storey", "discipline", "element_type_storey"})

UNCLASSIFIED = "Unclassified"
UNASSIGNED = "Unassigned"

_QTY_NUMBER_FORMAT = "#,##0.000"
_COUNT_NUMBER_FORMAT = "#,##0"
_HEADER_FILL = PatternFill(start_color="FFE0E0E0", end_color="FFE0E0E0", fill_type="solid")
_TOTAL_FILL = PatternFill(start_color="FFF3F4F6", end_color="FFF3F4F6", fill_type="solid")

# Fixed row layout of the BOQ summary sheet so tests have stable anchors.
TITLE_ROW = 1
META_ROW = 2
HEADER_ROW = 4
DATA_START_ROW = 5


@dataclass
class BoqExportOptions:
    """Builder configuration. ``frozen_now`` pins the generated-on stamp so
    snapshot tests stay deterministic; production passes ``None``."""

    title: str | None = None
    group_by: str = "element_type"
    include_detail: bool = True
    frozen_now: datetime | None = None


# ── Public API ───────────────────────────────────────────────────────────


def build_boq_workbook(
    model: Any,
    elements: list[Any],
    options: BoqExportOptions | None = None,
) -> bytes:
    """Build a BOQ workbook for the supplied model + (pre-filtered) elements.

    Args:
        model: BIMModel-like object exposing ``name`` (attribute access only,
            so tests can pass a ``SimpleNamespace``).
        elements: BIMElement-like objects exposing ``element_type``,
            ``storey``, ``discipline``, ``name``, ``stable_id`` and a
            ``quantities`` dict.
        options: builder configuration; defaults used when ``None``.

    Returns:
        XLSX file bytes.
    """
    opts = options or BoqExportOptions()
    group_by = opts.group_by if opts.group_by in GROUP_BY_CHOICES else "element_type"

    wb = Workbook()
    summary = wb.active
    if summary is not None:
        summary.title = "BOQ"
    _write_summary_sheet(wb["BOQ"], model, elements, group_by, opts)

    if opts.include_detail:
        _write_detail_sheet(wb.create_sheet("Elements"), elements)

    # Stamp authorship into docProps so a downloaded BOQ carries its origin
    # even when the UI strings are localised away (mirrors cobie.py).
    try:
        wb.properties.creator = "OpenConstructionERP · DataDrivenConstruction"
        wb.properties.lastModifiedBy = "OpenConstructionERP"
        wb.properties.title = f"Bill of Quantities - {getattr(model, 'name', 'Model')}"
        wb.properties.description = (
            "Generated by OpenConstructionERP (https://openconstructionerp.com) · DDC-CWICR-OE-2026"
        )
    except Exception:  # noqa: BLE001 - best-effort metadata stamp
        pass

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# ── Summary sheet ────────────────────────────────────────────────────────


def _group_columns(group_by: str) -> list[str]:
    if group_by == "storey":
        return ["Storey"]
    if group_by == "discipline":
        return ["Discipline"]
    if group_by == "element_type_storey":
        return ["Element Type", "Storey"]
    return ["Element Type"]


def _group_key(element: Any, group_by: str) -> tuple[str, ...]:
    etype = _clean(getattr(element, "element_type", None), UNCLASSIFIED)
    storey = _clean(getattr(element, "storey", None), UNASSIGNED)
    discipline = _clean(getattr(element, "discipline", None), UNASSIGNED)
    if group_by == "storey":
        return (storey,)
    if group_by == "discipline":
        return (discipline,)
    if group_by == "element_type_storey":
        return (etype, storey)
    return (etype,)


def _write_summary_sheet(
    ws: Worksheet,
    model: Any,
    elements: list[Any],
    group_by: str,
    opts: BoqExportOptions,
) -> None:
    group_cols = _group_columns(group_by)
    qty_headers = [f"{label} ({unit})" for _key, label, unit, _aliases in QUANTITY_FIELDS]
    columns = ["#", *group_cols, "Count", *qty_headers]

    # Title + meta block.
    title = opts.title or f"Bill of Quantities - {getattr(model, 'name', 'Model')}"
    ws.cell(row=TITLE_ROW, column=1, value=title).font = Font(bold=True, size=14)
    generated_on = _format_datetime(opts.frozen_now or datetime.now(UTC))
    ws.cell(
        row=META_ROW,
        column=1,
        value=f"Model: {getattr(model, 'name', 'Model')}    Generated: {generated_on}    Elements: {len(elements)}",
    ).font = Font(italic=True, color="FF6B7280")

    _write_header_row(ws, columns, HEADER_ROW)

    # Aggregate: group key -> [count, sum_area, sum_volume, sum_length, sum_weight].
    buckets: dict[tuple[str, ...], list[float]] = {}
    for el in elements:
        key = _group_key(el, group_by)
        bucket = buckets.setdefault(key, [0.0] * (1 + len(QUANTITY_FIELDS)))
        bucket[0] += 1
        qty = getattr(el, "quantities", None)
        for i, (_canon, _label, _unit, aliases) in enumerate(QUANTITY_FIELDS, start=1):
            bucket[i] += _quantity(qty, aliases)

    totals = [0.0] * (1 + len(QUANTITY_FIELDS))
    row_idx = DATA_START_ROW
    for line_no, key in enumerate(sorted(buckets), start=1):
        bucket = buckets[key]
        ws.cell(row=row_idx, column=1, value=line_no)
        col = 2
        for part in key:
            ws.cell(row=row_idx, column=col, value=part)
            col += 1
        count_cell = ws.cell(row=row_idx, column=col, value=int(bucket[0]))
        count_cell.number_format = _COUNT_NUMBER_FORMAT
        col += 1
        totals[0] += bucket[0]
        for i in range(1, len(QUANTITY_FIELDS) + 1):
            val = round(bucket[i], 3)
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.number_format = _QTY_NUMBER_FORMAT
            totals[i] += bucket[i]
            col += 1
        row_idx += 1

    # TOTAL row.
    total_cell = ws.cell(row=row_idx, column=1, value="TOTAL")
    total_cell.font = Font(bold=True)
    total_cell.fill = _TOTAL_FILL
    # blank the group columns, but fill them so the band reads as one row
    col = 2
    for _ in group_cols:
        ws.cell(row=row_idx, column=col).fill = _TOTAL_FILL
        col += 1
    c = ws.cell(row=row_idx, column=col, value=int(totals[0]))
    c.font = Font(bold=True)
    c.fill = _TOTAL_FILL
    c.number_format = _COUNT_NUMBER_FORMAT
    col += 1
    for i in range(1, len(QUANTITY_FIELDS) + 1):
        c = ws.cell(row=row_idx, column=col, value=round(totals[i], 3))
        c.font = Font(bold=True)
        c.fill = _TOTAL_FILL
        c.number_format = _QTY_NUMBER_FORMAT
        col += 1

    _autosize(ws, columns, header_row=HEADER_ROW)


# ── Detail sheet ─────────────────────────────────────────────────────────

DETAIL_COLUMNS = [
    "Stable ID",
    "Name",
    "Element Type",
    "Storey",
    "Discipline",
    "Area (m2)",
    "Volume (m3)",
    "Length (m)",
    "Weight (kg)",
]


def _write_detail_sheet(ws: Worksheet, elements: list[Any]) -> None:
    _write_header_row(ws, DETAIL_COLUMNS, 1)

    def _sort_key(el: Any) -> tuple[str, str, str, str]:
        return (
            _clean(getattr(el, "element_type", None), UNCLASSIFIED),
            _clean(getattr(el, "storey", None), UNASSIGNED),
            str(getattr(el, "name", "") or ""),
            str(getattr(el, "stable_id", "") or ""),
        )

    for el in sorted(elements, key=_sort_key):
        qty = getattr(el, "quantities", None)
        ws.append(
            [
                str(getattr(el, "stable_id", "") or ""),
                str(getattr(el, "name", "") or ""),
                _clean(getattr(el, "element_type", None), UNCLASSIFIED),
                _clean(getattr(el, "storey", None), UNASSIGNED),
                _clean(getattr(el, "discipline", None), UNASSIGNED),
                *(round(_quantity(qty, aliases), 3) for _k, _l, _u, aliases in QUANTITY_FIELDS),
            ]
        )
    # Number-format the four trailing quantity columns.
    first_qty_col = len(DETAIL_COLUMNS) - len(QUANTITY_FIELDS) + 1
    for r in range(2, ws.max_row + 1):
        for c in range(first_qty_col, len(DETAIL_COLUMNS) + 1):
            ws.cell(row=r, column=c).number_format = _QTY_NUMBER_FORMAT
    _autosize(ws, DETAIL_COLUMNS, header_row=1)


# ── Helpers ──────────────────────────────────────────────────────────────


def _write_header_row(ws: Worksheet, columns: list[str], row: int) -> None:
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=row, column=col_idx, value=col_name)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left")
        cell.fill = _HEADER_FILL


def _quantity(quantities: Any, aliases: tuple[str, ...]) -> float:
    """Return the first numeric value among ``aliases`` in the quantities
    blob, or 0.0. Tolerates missing/None blobs and string numbers."""
    if not isinstance(quantities, dict):
        return 0.0
    for key in aliases:
        if key in quantities:
            val = _coerce_float(quantities[key])
            if val is not None:
                return val
    return 0.0


def _coerce_float(value: Any) -> float | None:
    if isinstance(value, bool):  # bool is an int subclass - reject it
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        try:
            return float(value.strip())
        except (TypeError, ValueError):
            return None
    return None


def _clean(value: Any, fallback: str) -> str:
    if value is None:
        return fallback
    text = str(value).strip()
    return text or fallback


def _format_datetime(dt: datetime) -> str:
    if dt.tzinfo is not None:
        dt = dt.astimezone(UTC).replace(tzinfo=None)
    return dt.strftime("%Y-%m-%dT%H:%M:%S")


def _autosize(ws: Worksheet, columns: list[str], *, header_row: int) -> None:
    """Best-effort column widths so the sheet is readable on open. Caps the
    width so a long element name cannot blow a column out to the full row."""
    for col_idx, _name in enumerate(columns, start=1):
        longest = 0
        for r in range(header_row, ws.max_row + 1):
            v = ws.cell(row=r, column=col_idx).value
            if v is not None:
                longest = max(longest, len(str(v)))
        ws.column_dimensions[ws.cell(row=header_row, column=col_idx).column_letter].width = min(
            max(longest + 2, 10), 42
        )
