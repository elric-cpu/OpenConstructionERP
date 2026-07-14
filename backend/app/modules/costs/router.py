# DDC-CWICR-OE: DataDrivenConstruction · OpenConstructionERP
# Copyright (c) 2026 Artem Boiko / DataDrivenConstruction
"""Cost database API routes.

Endpoints:
    GET  /autocomplete    -- Fast text autocomplete for cost items (public)
    POST /                -- Create a cost item (auth required)
    GET  /                -- Search cost items (public, query params)
    GET  /{item_id}       -- Get cost item by ID
    PATCH /{item_id}      -- Update cost item (auth required)
    DELETE /{item_id}     -- Delete cost item (auth required)
    POST /bulk            -- Bulk import cost items (auth required)
    POST /import/file     -- Import cost items from Excel/CSV file (auth required)
    POST /load-cwicr/{db_id} -- Load CWICR regional database (auth required)
    POST /suggest-for-element          -- Rank cost items for a BIM element body
    POST /suggest-for-element/{id}     -- Same, loading the element by its UUID
"""

from __future__ import annotations

import csv
import io
import json
import logging
import math as _math
import re as _re
import urllib.parse
import uuid
import zipfile
from decimal import Decimal
from pathlib import Path
from typing import Any

from fastapi import APIRouter, Depends, File, Form, Header, HTTPException, Query, UploadFile, status
from fastapi.responses import JSONResponse, StreamingResponse
from pydantic import BaseModel, Field
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.file_signature import (
    SIGNATURE_BYTES_REQUIRED,
    FileSignatureMismatch,
)
from app.core.file_signature import (
    detect as detect_signature,
)
from app.core.file_signature import (
    require as require_signature,
)
from app.dependencies import (
    CurrentUserId,
    CurrentUserPayload,
    OptionalUserPayload,
    RequirePermission,
    RequireRole,
    SessionDep,
    verify_project_access,
)
from app.modules.costs import base_registry
from app.modules.costs.cwicr_v3_catalogue import CWICR_V3_CATALOGUES
from app.modules.costs.intelligence import (
    CostCertaintyService,
    CostUsageRecorder,
    RegionalIndexService,
)
from app.modules.costs.matcher import (
    MatchResult,
    match_cwicr_for_position,
    match_cwicr_items,
)
from app.modules.costs.models import CostItem
from app.modules.costs.repository import synonym_text_predicate  # noqa: F401
from app.modules.costs.resource_pricing import ResourcePriceService
from app.modules.costs.schemas import (
    BenchmarkRequest,
    BenchmarkResponse,
    CategoryTreeNode,
    CertaintyBadge,
    CostAutocompleteItem,
    CostCatalogCreate,
    CostCatalogResponse,
    CostCatalogUpdate,
    CostItemCreate,
    CostItemResponse,
    CostItemUpdate,
    CostSearchQuery,
    CostSuggestion,
    CwicrMatchFromPositionRequest,
    CwicrMatchRequest,
    MassApplyPreviewResponse,
    RecordUsageRequest,
    RegionalAdjustResponse,
    RegionalIndexResponse,
    RepriceResponse,
    ResourcePriceBulkRequest,
    ResourcePriceListResponse,
    ResourcePriceRow,
    ResourcePriceSetRequest,
    ResourcePriceStats,
    ResourceSeedResponse,
    SuggestCostsForElementRequest,
)
from app.modules.costs.service import CostBenchmarkService, CostCatalogService, CostItemService
from app.modules.costs.translations import localize_cost_row

# Round-7 upload safety: cap incoming bulk imports at 100 MB so a renamed
# binary can't waste arbitrary memory on the parse path before the
# magic-byte gate has a chance to reject it. A real-world CWICR
# CSV/Excel of 55K rows is ~8 MB; 100 MB leaves comfortable headroom for
# annotated columns without exposing the parser to multi-GB blobs.
_MAX_COST_UPLOAD_BYTES = 100 * 1024 * 1024

# Zip-bomb guard for OOXML (.xlsx) uploads: openpyxl inflates archive
# entries in memory, so a tiny upload that decompresses to gigabytes
# would bypass the on-the-wire size cap above. Reject archives whose
# declared UNCOMPRESSED payload or entry count is absurd for a workbook.
_MAX_COST_ZIP_UNCOMPRESSED_BYTES = 500 * 1024 * 1024
_MAX_COST_ZIP_ENTRIES = 10_000

# Magic-byte allow-list for the /import/file/ endpoint. ZIP covers OOXML
# (xlsx); OLE covers legacy .xls; pure CSV has no magic so we accept the
# ``None`` signature only when the body decodes as text and contains a
# common delimiter (handled inline in the route).
_ALLOWED_COST_IMPORT_SIGNATURES: frozenset[str] = frozenset({"zip", "ole"})

router = APIRouter(tags=["costs"])
logger = logging.getLogger(__name__)


from app.core.sql_json import json_path_text


class CertaintyBatchRequest(BaseModel):
    """Request body for ``POST /v1/costs/certainty/batch``.

    Carries the cost-item ids visible on one list page so the certainty
    badges can be resolved in a single round-trip instead of one HTTP
    request per row (an N+1 the list view fired on every page). Bounded
    to 200 ids - comfortably above the 10-row default page size while
    still capping the ``IN()`` fan-out.
    """

    ids: list[uuid.UUID] = Field(
        ...,
        min_length=1,
        max_length=200,
        description="Cost-item ids to grade (deduplicated server-side; unknown ids dropped).",
    )


class UsageCountsRequest(BaseModel):
    """Request body for ``POST /v1/costs/usage-counts/``.

    Carries the cost-item ids visible on one list page so the "used in N
    estimates" indicator resolves in a single grouped query instead of one
    request per row. Bounded to 200 ids, mirroring the certainty batch.
    """

    ids: list[uuid.UUID] = Field(
        ...,
        min_length=1,
        max_length=200,
        description="Cost-item ids to count usage for (deduplicated server-side).",
    )


# ── Region → currency map ─────────────────────────────────────────────────
#
# CWICR catalogues are imported per-region, but the parquet files don't
# carry an explicit currency column - every rate is denominated in the
# region's local currency. We resolve the right ISO 4217 code at ingestion
# time (so rates persist with their true currency) AND lazily on read for
# legacy rows that landed with ``currency = ''`` before this map existed.
#
# Single source of truth: the v3 catalogue registry
# (:data:`CWICR_V3_CATALOGUES`) already declares the ISO currency of every
# region DDC ships. Deriving the map from it means new catalogue rows are
# covered automatically and the two can never drift - the old hand-kept
# literal omitted ~18 live regions (KES/GHS/KRW/THB/VND/…) and silently
# mislabeled their rates as EUR.
#
# Legacy / alias keys that are NOT in the v3 registry (older parquet
# ``db_id`` tags the importer still accepts) are merged on top so they keep
# resolving. Keys follow the parquet ``db_id`` / ``region`` convention
# (UPPERCASE, country prefix).
_REGION_CURRENCY_LEGACY: dict[str, str] = {
    "DE_HAMBURG": "EUR",
    "BE_BRUSSELS": "EUR",
    "IE_DUBLIN": "EUR",
    "USA_NEWYORK": "USD",
    "SA_RIYADH": "SAR",
    # China authentic base (Beijing 2012 + Bortala 2022, prefixed rate_codes).
    # Loaded from our own work-items parquet, not a DDC v3 snapshot, so it lives
    # in the legacy overlay rather than the v3 registry.
    "ZH_CHINA": "CNY",
    # Turkey authentic national base (CSB analyses), separate from the legacy
    # metro id used by DDC snapshots.
    "TR_NATIONAL": "TRY",
    # Authentic national / regional bases loaded from our own work-items
    # parquet (official government sources), not DDC v3 snapshots, so they sit
    # in the legacy overlay. The parquet also carries a per-row currency column
    # that _resolve_currency prefers; these entries are the read-path fallback.
    "BR_NATIONAL": "BRL",
    "ES_ANDALUCIA": "EUR",
    "IT_TOSCANA": "EUR",
    "VN_NATIONAL": "VND",
    "ID_NATIONAL": "IDR",
    "GR_NATIONAL": "EUR",
    # NOTE: ``PT_SAOPAULO`` is intentionally NOT registered - it was a
    # mislabeled tag (São Paulo is Brazil; canonical key is ``BR_SAOPAULO``,
    # supplied by the v3 registry). A stray ``PT_SAOPAULO`` row should hit
    # the unknown-region path, not silently resolve.
}


def _build_region_currency_map() -> dict[str, str]:
    """Derive ``{region: ISO currency}`` from the v3 catalogue + legacy aliases."""
    out: dict[str, str] = {cat.region: cat.currency for cat in CWICR_V3_CATALOGUES if cat.currency}
    # Legacy/alias keys only fill gaps - never override a canonical v3 entry.
    for region, currency in _REGION_CURRENCY_LEGACY.items():
        out.setdefault(region, currency)
    return out


_REGION_CURRENCY: dict[str, str] = _build_region_currency_map()


# CWICR region tags follow the convention ``<2-letter country>_<UPPERCASE city>``
# (a few legacy tags use a 3-letter prefix like ``USA_``). Anything that doesn't
# match this shape is almost certainly junk / a typo and should not be silently
# resolved to "EUR" - we log a warning so operations can spot the bad row.
_REGION_FORMAT_RE = _re.compile(r"^[A-Z]{2,3}_[A-Z0-9]+$")


def _is_valid_region_format(region: str) -> bool:
    """Return True when ``region`` looks like a canonical CWICR region tag."""
    return bool(_REGION_FORMAT_RE.match(region))


def _resolve_currency(
    currency: str | None,
    region: str | None,
    *,
    warnings: list[str] | None = None,
) -> str:
    """Return the catalogue currency, deriving it from region when empty.

    The CWICR import historically stored ``currency = ''`` because the source
    parquet doesn't carry the field - every rate is in the region's local
    currency. This helper plugs that hole without forcing a re-import.

    Resolution order:
        1. Non-empty incoming ``currency`` (caller-supplied wins).
        2. ``_REGION_CURRENCY[region]`` when the region matches a known key
           (derived from the v3 catalogue registry, so every shipped region
           resolves to its true ISO code).
        3. ``""`` (unset) when the region is unknown or malformed.

    A genuinely unknown region returns an EMPTY string rather than a wrong
    "EUR" - mislabeling a Kenyan/Thai/Korean rate as EUR silently corrupts
    every downstream cross-currency conversion, whereas an empty currency is
    honestly "unknown" and is rendered as such (and skipped by FX maths).
    When the region can't be resolved a structured warning is emitted via
    ``logger.warning`` and - if a ``warnings`` list is supplied by the caller
    - a short human-readable message is appended so the route handler can
    surface it to the API response (frontend renders as a non-blocking toast).
    """
    if isinstance(currency, str):
        cleaned = currency.strip().upper()
        if cleaned:
            return cleaned
    if isinstance(region, str):
        normalized = region.strip().upper()
        if normalized:
            if not _is_valid_region_format(normalized):
                msg = (
                    f"Cost row uses non-canonical region tag {normalized!r} "
                    f"(expected ``XX_CITY``); currency left unset."
                )
                logger.warning(msg)
                if warnings is not None and msg not in warnings:
                    warnings.append(msg)
            else:
                mapped = _REGION_CURRENCY.get(normalized)
                if mapped:
                    return mapped
                msg = (
                    f"Unknown region {normalized!r} - no entry in _REGION_CURRENCY "
                    f"(add it to the CWICR catalogue registry); currency left unset."
                )
                logger.warning(msg)
                if warnings is not None and msg not in warnings:
                    warnings.append(msg)
    return ""


def _get_service(session: SessionDep) -> CostItemService:
    return CostItemService(session)


def _get_catalog_service(session: SessionDep) -> CostCatalogService:
    return CostCatalogService(session)


def _get_resource_price_service(session: SessionDep) -> ResourcePriceService:
    return ResourcePriceService(session)


def _parse_user_uuid(user_id: str | None) -> uuid.UUID | None:
    """Best-effort UUID parse of the auth user id (None for non-UUID ids)."""
    if not user_id:
        return None
    try:
        return uuid.UUID(str(user_id))
    except (ValueError, AttributeError, TypeError):
        return None


# ── Autocomplete metadata helpers (Phase F v2.7.0) ────────────────────────


_BREAKDOWN_KEYS: tuple[str, ...] = ("labor_cost", "material_cost", "equipment_cost")


def _extract_cost_breakdown(metadata: dict[str, Any] | None) -> dict[str, float] | None:
    """Pull labor / material / equipment numbers out of CWICR metadata.

    The CWICR ingest stamps these as ``round(value, 2)`` only when the
    source row carries a non-zero figure - so an absent key really means
    "no data" (not "zero"). Returns ``None`` when none of the three keys
    are present so the tooltip can hide the breakdown section gracefully.
    """
    if not isinstance(metadata, dict) or not metadata:
        return None
    out: dict[str, float] = {}
    for key in _BREAKDOWN_KEYS:
        v = metadata.get(key)
        if isinstance(v, (int, float)) and v >= 0:
            out[key] = float(v)
    return out or None


def _slim_autocomplete_metadata(metadata: dict[str, Any] | None) -> dict[str, Any] | None:
    """Project metadata to a tooltip-sized payload.

    Keeps:
      * ``variant_stats`` - rendered as the "N variants" hint.
      * ``variant_count`` - derived count when ``variants`` is present.
      * ``labor_hours`` / ``workers_per_unit`` - small auxiliary numbers.
      * ``scope_of_work`` - ordered list of work steps (truncated to 8
        entries to keep the payload bounded). Surfaced in the BOQ grid
        as an inline (i) hint next to the description.

    Strips the heavy ``variants`` array - full variant data is fetched
    lazily via ``GET /v1/costs/{id}/`` when the user actually applies
    the suggestion. The slim payload is bounded to roughly < 200 B per
    item so the autocomplete response stays snappy on slow links.
    """
    if not isinstance(metadata, dict) or not metadata:
        return None
    out: dict[str, Any] = {}
    stats = metadata.get("variant_stats")
    if isinstance(stats, dict) and stats:
        out["variant_stats"] = stats
    variants = metadata.get("variants")
    if isinstance(variants, list) and variants:
        out["variant_count"] = len(variants)
    for k in ("labor_hours", "workers_per_unit"):
        v = metadata.get(k)
        if isinstance(v, (int, float)) and v > 0:
            out[k] = float(v)
    sow = metadata.get("scope_of_work")
    if isinstance(sow, list) and sow:
        # Cap at 8 steps to keep the autocomplete payload small. The
        # full list (often 10–20 steps for complex CWICR rates) is
        # available via ``GET /v1/costs/{id}/`` when needed.
        out["scope_of_work"] = [str(s)[:300] for s in sow[:8] if str(s).strip()]
    return out or None


# ── Locale resolution ─────────────────────────────────────────────────────


def _resolve_cost_locale(
    locale_param: str | None,
    accept_language: str | None,
) -> str:
    """Pick the best CWICR translation locale for an HTTP request.

    Priority:
      1. ``?locale=ro`` query parameter (explicit, wins over header).
      2. First language tag of ``Accept-Language`` (RFC 7231, region stripped).
      3. ``"en"`` fallback.

    The CWICR translations module uses its own SUPPORTED_LOCALES (16 entries)
    independently of ``app.core.i18n`` (20 entries) - they overlap but the
    CWICR set adds ``ro``, ``bg``, ``hr``, ``id``, ``th``, ``vi`` that the
    UI-strings i18n doesn't ship yet.  Pulling the locale here keeps the
    cost-data path decoupled from the broader request-locale middleware so
    a missing UI locale doesn't accidentally lose a CWICR translation.
    """
    from app.modules.costs.translations import SUPPORTED_LOCALES as COST_LOCALES

    # 1. Explicit query param wins. Strip region (de-DE → de).
    if locale_param:
        norm = locale_param.strip().lower().split("-")[0]
        if norm in COST_LOCALES:
            return norm

    # 2. First entry of Accept-Language. Quality-weighted parsing isn't
    #    necessary here - the costs UI only needs a single best-match,
    #    and the existing AcceptLanguageMiddleware already does the
    #    full RFC 7231 dance for the rest of the app.
    if accept_language:
        for raw in accept_language.split(","):
            tag = raw.split(";", 1)[0].strip().lower().split("-")[0]
            if tag in COST_LOCALES:
                return tag

    return "en"


def _localize_response_payload(
    item_response: CostItemResponse,
    locale: str,
) -> dict[str, Any]:
    """Convert a CostItemResponse to a dict with localized mirror fields.

    Pydantic responses are immutable for safety, so we serialize → mutate →
    return a dict.

    Note on the ``metadata_`` key: ``CostItemResponse`` defines its
    metadata field with ``alias="metadata_"`` (SQLAlchemy reserves
    ``metadata`` for its DeclarativeBase namespace).  ``model_dump(
    by_alias=True)`` therefore emits the alias, and frontend clients
    already key off ``metadata_`` (see ``api.ts``
    ``CostItemMetadata``) - keep that contract intact.
    """
    payload = item_response.model_dump(by_alias=True, mode="json")
    cls = payload.get("classification") or {}
    # Schema uses alias="metadata_" → that's the dumped key here.
    md = payload.get("metadata_") or {}
    comps = payload.get("components") or []
    localize_cost_row(
        classification=cls,
        metadata=md,
        components=comps,
        locale=locale,
    )
    payload["classification"] = cls
    payload["metadata_"] = md
    payload["components"] = comps
    return payload


# ── Autocomplete ──────────────────────────────────────────────────────────


@router.get("/autocomplete/", response_model=list[CostAutocompleteItem])
async def autocomplete_cost_items(
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    service: CostItemService = Depends(_get_service),
    q: str = Query(
        ...,
        min_length=2,
        max_length=200,
        description=(
            "Search text (min 2 chars). A single word is expanded with "
            "multilingual construction synonyms, so 'rebar' also finds "
            "'reinforcement' / 'Bewehrung' / 'armatura' and 'beton' finds "
            "'concrete'; the word you type still matches as a substring."
        ),
    ),
    region: str | None = Query(default=None, description="Filter by region (e.g. DE_BERLIN)"),
    limit: int = Query(default=20, ge=1, le=200, description="Max results to return"),
    semantic: bool = Query(default=False, description="Use vector semantic search if available"),
    locale: str | None = Query(
        default=None,
        max_length=10,
        description="Localize CWICR-frozen-German fields (see search endpoint).",
    ),
    accept_language: str | None = Header(default=None, alias="accept-language"),
) -> list[CostAutocompleteItem]:
    """Fast autocomplete for cost items. Uses vector semantic search when available.

    When ``semantic=true`` and a vector index exists, uses AI embeddings
    to find semantically similar items (e.g. "concrete wall" finds
    "reinforced partition C30/37"). Falls back to a multilingual synonym
    text search otherwise, so "rebar" also finds "reinforcement" /
    "Bewehrung" / "armatura" without any vector index installed.

    The response carries a slim ``cost_breakdown`` (labor / material /
    equipment) and a thinned ``metadata_`` block so the BOQ description
    cell can render a rich hover tooltip (Phase F, v2.7.0) without a
    second round-trip. The variant array itself is intentionally omitted
    to keep the per-item delta well under 200 B - callers that need the
    full variant catalog should hit ``GET /v1/costs/{id}/`` on hover.
    """
    resolved_locale = _resolve_cost_locale(locale, accept_language)
    # Try vector search first if requested
    if semantic:
        try:
            from app.core.vector import encode_texts, vector_search, vector_status

            status = vector_status()
            if status.get("connected") and status.get("cost_collection"):
                query_vec = encode_texts([q])[0]
                results = vector_search(query_vec, region=region, limit=limit)
                if results:
                    # Vector results may not have components - look them up from DB
                    codes = [r.get("code", "") for r in results]
                    components_map: dict[str, list[dict[str, Any]]] = {}
                    metadata_map: dict[str, dict[str, Any]] = {}
                    try:
                        items_from_db = await service.get_by_codes(codes)
                        for db_item in items_from_db:
                            components_map[db_item.code] = db_item.components or []
                            metadata_map[db_item.code] = db_item.metadata_ or {}
                    except Exception:
                        logger.debug("Cost search: component lookup failed", exc_info=True)

                    out: list[CostAutocompleteItem] = []
                    for r in results:
                        cls = dict(r.get("classification") or {})
                        comps = list(components_map.get(r.get("code", ""), []))
                        md_full = metadata_map.get(r.get("code", ""), {})
                        # Mutates cls/comps in place to add *_localized keys.
                        localize_cost_row(
                            classification=cls,
                            metadata=None,
                            components=comps,
                            locale=resolved_locale,
                        )
                        breakdown = _extract_cost_breakdown(md_full)
                        slim_md = _slim_autocomplete_metadata(md_full)
                        out.append(
                            CostAutocompleteItem(
                                code=r.get("code", ""),
                                description=r.get("description", ""),
                                unit=r.get("unit", ""),
                                rate=float(r.get("rate", 0)),
                                currency=_resolve_currency(r.get("currency"), r.get("region")),
                                region=r.get("region"),
                                classification=cls,
                                components=comps,
                                cost_breakdown=breakdown,
                                metadata_=slim_md,
                            )
                        )
                    return out
        except Exception:
            logger.debug("Cost search: vector search failed, falling back to text", exc_info=True)

    # Standard text search - the "items WITH components first" priority
    # is pushed into SQL so we fetch exactly ``limit`` rows here (was
    # ``limit*3`` + Python sort). On a 110k-row catalogue this cut the
    # per-keystroke cost from ~80 ms (24-row fetch + 24-row Python sort)
    # to ~8 ms (single SQL roundtrip, no post-processing sort).
    items = await service.search_for_autocomplete(q=q, region=region, limit=limit)

    import json as _json

    def _parse_components(raw: object) -> list[dict[str, Any]]:
        if isinstance(raw, str):
            try:
                parsed = _json.loads(raw)
                return parsed if isinstance(parsed, list) else []
            except Exception:
                return []
        return raw if isinstance(raw, list) else []

    out: list[CostAutocompleteItem] = []
    for item in items:
        cls = dict(item.classification or {})
        comps = _parse_components(item.components)
        localize_cost_row(
            classification=cls,
            metadata=None,
            components=comps,
            locale=resolved_locale,
        )
        md_full = item.metadata_ or {}
        breakdown = _extract_cost_breakdown(md_full)
        slim_md = _slim_autocomplete_metadata(md_full)
        out.append(
            CostAutocompleteItem(
                code=item.code,
                description=item.description,
                unit=item.unit,
                rate=float(item.rate),
                currency=_resolve_currency(
                    getattr(item, "currency", None),
                    getattr(item, "region", None),
                ),
                region=getattr(item, "region", None),
                classification=cls,
                components=comps,
                cost_breakdown=breakdown,
                metadata_=slim_md,
            )
        )
    return out


# ── Create ────────────────────────────────────────────────────────────────


@router.post(
    "/",
    response_model=CostItemResponse,
    status_code=201,
    dependencies=[Depends(RequirePermission("costs.create"))],
)
async def create_cost_item(
    data: CostItemCreate,
    user: CurrentUserPayload,
    service: CostItemService = Depends(_get_service),
    catalog_service: CostCatalogService = Depends(_get_catalog_service),
) -> CostItemResponse:
    """Create a new cost item.

    When ``catalog_id`` is set, the caller must own that catalog (or be an
    admin) - this is the write-side half of the ownership gate already
    enforced on update/delete (``_enforce_item_catalog_ownership``), closing
    the gap where any caller holding ``costs.create`` could otherwise add a
    position into another user's private catalog by guessing its UUID. A
    non-owner gets a 404 (existence not leaked), matching
    ``get_owned_catalog``. Items with no ``catalog_id`` keep going into the
    shared global catalogue as before.
    """
    if data.catalog_id is not None:
        owner_id = _parse_user_uuid((user or {}).get("sub"))
        is_admin = (user or {}).get("role") == "admin"
        await catalog_service.get_owned_catalog(data.catalog_id, owner_id=owner_id, is_admin=is_admin)
    item = await service.create_cost_item(data)
    return CostItemResponse.model_validate(item)


# ── Search / List ─────────────────────────────────────────────────────────


@router.get("/")
async def search_cost_items(
    user: CurrentUserPayload = None,  # type: ignore[assignment]
    service: CostItemService = Depends(_get_service),
    catalog_service: CostCatalogService = Depends(_get_catalog_service),
    q: str | None = Query(
        default=None,
        description=(
            "Free-text search against code OR description. Routed through the "
            "multilingual construction-vocabulary matcher, so a single-word "
            "query like 'rebar' also finds 'reinforcement' / 'Bewehrung' / "
            "'armatura' and 'beton' finds 'concrete'; the word you type still "
            "matches as an (accent- and case-insensitive) substring. Canonical "
            "param: ``search`` and ``query`` are silently aliased to ``q`` at "
            "this boundary. The vector layer is a best-effort re-rank on top, "
            "never the only source of recall."
        ),
    ),
    search: str | None = Query(
        default=None,
        description="Alias of ``q``. Silently merged when both are passed.",
    ),
    query_param: str | None = Query(
        default=None,
        alias="query",
        description="Alias of ``q``. Silently merged when both are passed.",
    ),
    name: str | None = Query(
        default=None,
        description=(
            "Substring (ILIKE) filter against code only - CostItem rows "
            "have no separate name column, so the catalog code IS the "
            "name. AND-combined with ``q``."
        ),
    ),
    description: str | None = Query(
        default=None,
        description="Substring (ILIKE) filter against description only. AND-combined with ``q``.",
    ),
    unit: str | None = Query(default=None, description="Filter by unit"),
    source: str | None = Query(default=None, description="Filter by source"),
    region: str | None = Query(default=None, description="Filter by region (e.g. DE_BERLIN)"),
    category: str | None = Query(
        default=None, description="Filter by classification.collection (construction category)"
    ),
    classification_path: str | None = Query(
        default=None,
        description=(
            "Slash-delimited classification prefix path "
            "(collection/department/section/subsection). Prefix-matches "
            "at any depth; empty middle segments act as wildcards. "
            "AND-combined with the other filters."
        ),
    ),
    catalog_id: uuid.UUID | None = Query(
        default=None,
        description="Filter to items in one user-owned cost catalog.",
    ),
    min_rate: Decimal | None = Query(default=None, ge=0, description="Minimum rate"),
    max_rate: Decimal | None = Query(default=None, ge=0, description="Maximum rate"),
    limit: int = Query(default=50, ge=1, le=100),
    offset: int = Query(default=0, ge=0),
    cursor: str | None = Query(
        default=None,
        description=(
            "Opaque keyset cursor returned in the previous page's "
            "``next_cursor``. When set, ``offset`` is ignored and "
            "``total`` is omitted."
        ),
    ),
    fuzzy: bool = Query(
        default=True,
        description=(
            "Typo- and word-order-tolerant fuzzy ranking for ``q`` via "
            "PostgreSQL trigram similarity. Exact and prefix hits rank first, "
            "then trigram similarity. Falls back automatically to plain "
            "substring matching when the pg_trgm extension is unavailable, so "
            "results never regress. Set false to force the legacy substring path."
        ),
    ),
    lite: bool = Query(
        default=False,
        description=(
            "Return a slim payload - strip the per-row ``components`` "
            "array (cwicr items can carry 16+ resource entries averaging "
            "~31 KB/row) and reduce ``metadata_`` to ``variant_stats`` only. "
            "Adds a ``components_count`` integer so list UIs can still show "
            "the breakdown badge without the full array. When the user "
            "drills into a row, callers fetch the full document via "
            "``GET /v1/costs/{id}`` for components and metadata."
        ),
    ),
    locale: str | None = Query(
        default=None,
        max_length=10,
        description=(
            "Localize CWICR-frozen-German fields for this locale "
            "(e.g. 'ro', 'bg', 'sv'). Falls back to Accept-Language. "
            "Mirrors the source values into *_localized keys; the "
            "originals stay untouched for backwards compatibility."
        ),
    ),
    accept_language: str | None = Header(default=None, alias="accept-language"),
) -> dict[str, Any]:
    """Search cost items with optional filters. Public endpoint.

    Returns a keyset-paginated response with items, optional total count,
    next_cursor, and has_more. Each item carries
    ``classification.category_localized``,
    ``metadata.variant_stats.unit_localized`` / ``_group_localized``, and
    per-component ``unit_localized`` mirror fields when the locale has a
    translation table.  Originals are preserved so older clients
    continue to read the German source.

    Backwards compatibility: clients that don't send ``cursor`` continue
    to receive a non-null ``total``. The new fields ``next_cursor`` and
    ``has_more`` are additions to the response shape.

    When ``catalog_id`` is supplied, the caller must own that catalog (or
    be an admin) - listing is otherwise the one place a private catalog's
    items were readable by anyone who could guess its UUID, even though the
    single-item GET/PATCH/DELETE endpoints already scope by ownership. A
    non-owner gets a 404 (existence not leaked), matching
    ``get_owned_catalog``.
    """
    if catalog_id is not None:
        owner_id = _parse_user_uuid((user or {}).get("sub"))
        is_admin = (user or {}).get("role") == "admin"
        await catalog_service.get_owned_catalog(catalog_id, owner_id=owner_id, is_admin=is_admin)

    # Merge canonical ``q`` with the silent aliases ``search`` / ``query``.
    # First non-empty wins; explicit ``q`` always takes precedence so a
    # caller that mistakenly sends both ``q=foo&search=bar`` gets ``foo``
    # rather than a surprise. Empty strings ("") count as absent.
    canonical_q = q or search or query_param or None

    query = CostSearchQuery(
        q=canonical_q,
        name=name,
        description=description,
        unit=unit,
        source=source,
        region=region,
        category=category,
        classification_path=classification_path,
        catalog_id=catalog_id,
        min_rate=min_rate,
        max_rate=max_rate,
        limit=limit,
        offset=offset,
        cursor=cursor,
        fuzzy=fuzzy,
    )
    # Fast-path: when no text/category filters are present and we already
    # know the per-region totals from the prewarmed stats cache, skip the
    # COUNT(*) over the filtered subquery on the first page. The user's
    # bug report - "Add from Database" modal hangs - was traced to this
    # cold-cache count: 18 s on a 277 k-row catalog. Subsequent pages use
    # cursors which already skip the count, so this only affects page 1.
    skip_count_via_cache = False
    cached_total: int | None = None
    if (
        cursor is None
        and not canonical_q
        and not name
        and not description
        and not category
        and not classification_path
        and catalog_id is None
        and not unit
        and not source
        and min_rate is None
        and max_rate is None
    ):
        stats = _region_cache.get("stats")
        if isinstance(stats, list):
            if region:
                for entry in stats:
                    if entry.get("region") == region:
                        cached_total = int(entry.get("count", 0))
                        skip_count_via_cache = True
                        break
            else:
                # No region filter: the per-region sum omits region-less rows,
                # so use the cached grand total (which includes them). Only skip
                # the COUNT when that accurate total is available.
                total_active = _region_cache.get("total_active")
                if isinstance(total_active, int):
                    cached_total = total_active
                    skip_count_via_cache = True

    if skip_count_via_cache:
        items, _, has_more, next_cursor = await service.search_costs_paginated(
            query,
            skip_count=True,
        )
        total = cached_total
    else:
        items, total, has_more, next_cursor = await service.search_costs_paginated(query)
    resolved_locale = _resolve_cost_locale(locale, accept_language)

    # Currency-fallback warnings - accumulate per-row issues so the FE can
    # surface one non-blocking toast per request instead of one per row.
    # _resolve_currency() (called from the schema validator + the manual
    # payload paths below) appends de-duplicated messages here.
    currency_warnings: list[str] = []

    # Lite payload trim - drops the heavy ``components`` array and trims
    # ``metadata_`` to a small whitelist. CWICR rows average ~38 KB each
    # (31 KB components + 6.6 KB metadata); a 10-row page is 380 KB on
    # the wire, which dominates the perceived load time of /costs even
    # though the SQL is fast. With ``lite=true`` a 10-row page drops to
    # ~3 KB. ``components_count`` preserves the "has breakdown" hint.
    def _serialize(i: Any) -> dict[str, Any]:
        # Funnel each row's empty-currency rows through _resolve_currency()
        # explicitly so the warning list captures bad rows. The schema
        # validator runs internally too, but it can't reach the route-scoped
        # warnings list, so do the resolve here BEFORE model_validate.
        try:
            row_currency = getattr(i, "currency", None) or ""
            row_region = getattr(i, "region", None)
            if not (isinstance(row_currency, str) and row_currency.strip()):
                resolved = _resolve_currency(
                    row_currency,
                    row_region,
                    warnings=currency_warnings,
                )
                # Stamp the resolved value onto the ORM instance so the
                # schema validator (which can't see ``warnings``) sees a
                # populated currency and short-circuits.
                try:
                    i.currency = resolved
                except Exception:
                    pass
        except Exception:
            logger.debug("Currency warning capture skipped", exc_info=True)
        payload = _localize_response_payload(
            CostItemResponse.model_validate(i),
            resolved_locale,
        )
        if lite:
            comps = payload.get("components") or []
            payload["components_count"] = len(comps)
            payload["components"] = []
            md = payload.get("metadata_") or {}
            if isinstance(md, dict):
                # Whitelist tiny keys the list view + BOQ-add synth path
                # actually consume. Drops ``variants`` (~6 KB / row of
                # alternate price entries) and any other large arrays.
                payload["metadata_"] = {
                    k: md[k]
                    for k in (
                        "variant_stats",
                        "labor_cost",
                        "material_cost",
                        "equipment_cost",
                        "labor_hours",
                        "workers_per_unit",
                        "scope_of_work",
                    )
                    if k in md
                }
        return payload

    serialized = [_serialize(i) for i in items]
    response: dict[str, Any] = {
        "items": serialized,
        "total": total,
        "limit": limit,
        "offset": offset,
        "next_cursor": next_cursor,
        "has_more": has_more,
    }
    # ``warnings``: non-fatal data-quality issues the FE renders as a single
    # transient toast (one entry per distinct message - duplicates already
    # collapsed by _resolve_currency). Omitted when empty so clients that
    # don't know about the field see no change in the response shape.
    if currency_warnings:
        response["warnings"] = currency_warnings
    return response


# ── Regions ───────────────────────────────────────────────────────────────


# ── In-memory cache for slow aggregate queries ──────────────────────────────

import time as _time

_region_cache: dict[str, Any] = {"regions": None, "stats": None, "categories": None, "total_active": None, "ts": 0}
# 1-hour TTL on the regions/stats/categories aggregates. Originally 30 s,
# bumped to 5 min, then to 60 min after the BOQ "Add from Database" modal
# was reported as taking 18 seconds to open on a cold backend. With 100 k+
# active cost items the DISTINCT/COUNT scan can take 15-20 s on cold
# SQLite, and the cache is correctly invalidated on import/delete via
# ``_invalidate_cost_cache()``, so a long TTL never risks staleness.
_CACHE_TTL = 3600

# The category tree is much heavier to compute (single GROUP BY across the
# four classification depths against the full active catalog) and rarely
# changes between imports. Cache it longer, per-region, and key the entries
# off a separate timestamp so it doesn't piggy-back on the 30s general TTL.
# Bumped from 5 min to 60 min for the same reason as ``_CACHE_TTL`` above:
# cold tree GROUP BY can hit 80+ seconds on 100 k+ row catalogs, and the
# cache is wiped on import via ``_invalidate_cost_cache()`` so user-visible
# data is always fresh after a CWICR load.
_CATEGORY_TREE_CACHE_TTL = 3600  # 60 minutes
_category_tree_cache: dict[str, dict[str, Any]] = {}


def _invalidate_cost_cache() -> None:
    """Call after import/delete to force refresh on next request.

    Wipes every value slot explicitly, not just the shared ``ts`` timestamp,
    so future cache keys that don't piggy-back on ``ts`` are still cleared.
    """
    for key in list(_region_cache.keys()):
        if key == "ts":
            _region_cache[key] = 0
        else:
            _region_cache[key] = None
    _category_tree_cache.clear()


@router.get("/regions/", response_model=list[str])
async def list_loaded_regions(
    session: SessionDep,
) -> list[str]:
    """List distinct regions that have cost items loaded."""
    now = _time.monotonic()
    if _region_cache["regions"] is not None and now - _region_cache["ts"] < _CACHE_TTL:
        return _region_cache["regions"]

    from sqlalchemy import distinct, select

    from app.modules.costs.models import CostItem

    result = await session.execute(
        select(distinct(CostItem.region))
        .where(CostItem.is_active.is_(True))
        .where(CostItem.region.isnot(None))
        .where(CostItem.region != "")
    )
    regions = sorted(row[0] for row in result.all())
    _region_cache["regions"] = regions
    _region_cache["ts"] = now
    return regions


@router.get("/regions/stats/")
async def region_stats(
    session: SessionDep,
) -> list[dict]:
    """Return item count per loaded region. Cached for 30s."""
    now = _time.monotonic()
    if _region_cache["stats"] is not None and now - _region_cache["ts"] < _CACHE_TTL:
        return _region_cache["stats"]

    from sqlalchemy import func, select

    from app.modules.costs.models import CostItem

    result = await session.execute(
        select(CostItem.region, func.count(CostItem.id).label("cnt"))
        .where(CostItem.is_active.is_(True))
        .where(CostItem.region.isnot(None))
        .where(CostItem.region != "")
        .group_by(CostItem.region)
        .order_by(func.count(CostItem.id).desc())
    )
    stats = [{"region": row[0], "count": row[1]} for row in result.all()]
    _region_cache["stats"] = stats
    # The per-region stats above deliberately exclude region-less rows (region
    # IS NULL / ""). The unfiltered search counts ALL active items, so using the
    # sum of per-region counts as the no-region total undercounts by exactly the
    # region-less items. Cache the true grand total (regioned + region-less) so
    # the search fast-path stays accurate without losing the count optimisation.
    region_less = (
        await session.execute(
            select(func.count(CostItem.id))
            .where(CostItem.is_active.is_(True))
            .where((CostItem.region.is_(None)) | (CostItem.region == ""))
        )
    ).scalar_one()
    _region_cache["total_active"] = sum(s["count"] for s in stats) + int(region_less or 0)
    _region_cache["ts"] = now
    return stats


@router.delete(
    "/actions/clear-region/{region}",
    # Wholesale region wipe - admin only. ``costs.delete`` alone would let
    # any editor nuke a whole regional cost database. Keeps parity with
    # ``/actions/clear-database/`` which already requires admin.
    dependencies=[Depends(RequireRole("admin"))],
)
async def clear_region_database(
    region: str,
    session: SessionDep,
    _user_id: CurrentUserId,
) -> dict:
    """Delete all cost items for a specific region.

    E.g. ``DELETE /actions/clear-region/DE_BERLIN`` removes all DE_BERLIN items.
    """
    from sqlalchemy import delete as sql_delete

    from app.modules.costs.models import CostItem

    stmt = sql_delete(CostItem).where(CostItem.region == region)
    result = await session.execute(stmt)
    await session.commit()
    count = result.rowcount  # type: ignore[union-attr]

    logger.info("Cleared region %s: %d items deleted", region, count)
    _invalidate_cost_cache()
    return {"deleted": count, "region": region}


# ── Resource price sheet (make coefficient bases calculable) ─────────────────
#
# CWICR describes each work item through its resource lines (labour / material /
# machine) with a norm quantity each. Coefficient bases (Vietnam Dinh Muc,
# Indonesia AHSP) ship those quantities with NO prices - they are priced
# regionally - so their work items import with a zero rate. These endpoints hold
# one editable unit price per resource per region, seed it from whatever prices a
# base already carries, and re-price every work item from the sheet
# (rate = sum(component.quantity x price)). The same path re-prices a priced base
# after a local price edit, so it is uniform for coded and codeless bases.


def _resource_row(row: Any) -> ResourcePriceRow:
    return ResourcePriceRow.model_validate(row)


@router.get("/resource-prices/{region}/", response_model=ResourcePriceListResponse)
async def list_resource_prices(
    region: str,
    _user_id: CurrentUserId,
    service: ResourcePriceService = Depends(_get_resource_price_service),
    search: str | None = Query(default=None, description="Filter by resource name (substring)."),
    resource_type: str | None = Query(
        default=None,
        description="labor | material | equipment | operator | electricity | other.",
    ),
    only_unpriced: bool = Query(default=False, description="Only rows still at price 0."),
    limit: int = Query(default=100, ge=1, le=500),
    offset: int = Query(default=0, ge=0),
) -> ResourcePriceListResponse:
    """List the resource price sheet for a region (paginated) with coverage stats.

    The price sheet is what makes a coefficient base estimable: every resource is
    listed with its current local unit price (0 = still needs a price). Priced
    bases come pre-filled; edit any row and re-price.
    """
    rows, total = await service.list_prices(
        region,
        search=search,
        resource_type=resource_type,
        only_unpriced=only_unpriced,
        limit=limit,
        offset=offset,
    )
    stats = await service.region_stats(region)
    return ResourcePriceListResponse(
        region=region,
        total=total,
        limit=limit,
        offset=offset,
        stats=ResourcePriceStats(**stats),
        rows=[_resource_row(r) for r in rows],
    )


@router.get("/resource-prices/{region}/stats/", response_model=ResourcePriceStats)
async def resource_price_stats(
    region: str,
    _user_id: CurrentUserId,
    service: ResourcePriceService = Depends(_get_resource_price_service),
) -> ResourcePriceStats:
    """Coverage of a region's price sheet (how many resources still need a price)."""
    return ResourcePriceStats(**await service.region_stats(region))


@router.post(
    "/resource-prices/{region}/seed/",
    response_model=ResourceSeedResponse,
    dependencies=[Depends(RequirePermission("costs.update"))],
)
async def seed_resource_prices(
    region: str,
    _user_id: CurrentUserId,
    service: ResourcePriceService = Depends(_get_resource_price_service),
) -> ResourceSeedResponse:
    """(Re)build the price sheet for a region from its work items.

    Collects every distinct resource and seeds its observed unit price (0 for a
    coefficient base). Idempotent and safe to re-run: rows a user has edited are
    preserved. Normally runs automatically on region load; this is the manual
    rebuild.
    """
    result = await service.seed_region(region)
    if result.resources == 0:
        raise HTTPException(
            status_code=404,
            detail=(f"No cost items found for region '{region}'. Load the base first, then seed its price sheet."),
        )
    _invalidate_cost_cache()
    return ResourceSeedResponse(**result.as_dict())


@router.put(
    "/resource-prices/{region}/{resource_key:path}",
    response_model=ResourcePriceRow,
    dependencies=[Depends(RequirePermission("costs.update"))],
)
async def set_resource_price(
    region: str,
    resource_key: str,
    body: ResourcePriceSetRequest,
    user_id: CurrentUserId,
    service: ResourcePriceService = Depends(_get_resource_price_service),
) -> ResourcePriceRow:
    """Set one resource's unit price for a region.

    ``resource_key`` is the value from the price-sheet row (a resource code, or a
    ``name:...`` key for codeless bases). Marks the row user-edited so a later
    re-seed leaves it alone. Call the re-price endpoint afterwards to fold the new
    price into the region's work-item rates.
    """
    try:
        row = await service.set_price(
            region,
            resource_key,
            body.unit_price,
            currency=body.currency,
            unit=body.unit,
            resource_name=body.resource_name,
            resource_type=body.resource_type,
            updated_by=_parse_user_uuid(user_id),
        )
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    return _resource_row(row)


@router.post(
    "/resource-prices/{region}/bulk/",
    dependencies=[Depends(RequirePermission("costs.update"))],
)
async def set_resource_prices_bulk(
    region: str,
    body: ResourcePriceBulkRequest,
    user_id: CurrentUserId,
    service: ResourcePriceService = Depends(_get_resource_price_service),
) -> dict:
    """Apply many resource-price edits to a region in one transaction."""
    written = await service.set_prices_bulk(
        region,
        [item.model_dump() for item in body.items],
        updated_by=_parse_user_uuid(user_id),
    )
    return {"region": region, "written": written}


@router.post(
    "/resource-prices/{region}/reprice/",
    response_model=RepriceResponse,
    dependencies=[Depends(RequirePermission("costs.update"))],
)
async def reprice_region_endpoint(
    region: str,
    _user_id: CurrentUserId,
    service: ResourcePriceService = Depends(_get_resource_price_service),
    dry_run: bool = Query(default=False, description="Preview the effect without writing."),
) -> RepriceResponse:
    """Recompute every work item's rate in a region from the price sheet.

    ``rate = sum(component.quantity x sheet_price)``. Each component's unit price
    and cost, and the labour/material/equipment breakdown, are refreshed too so
    the stored rate stays explainable. Use ``dry_run`` to preview coverage before
    committing.
    """
    result = await service.reprice_region(region, dry_run=dry_run)
    if not dry_run:
        _invalidate_cost_cache()
    return RepriceResponse(**result.as_dict())


# ── Vector database (LanceDB embedded / Qdrant server) ──────────────────────


@router.get("/vector/status/")
async def get_vector_status() -> dict:
    """Check vector DB status (LanceDB embedded or Qdrant server)."""
    from app.core.vector import vector_status as vs

    return vs()


@router.get("/vector/download-status/")
async def vector_download_status() -> dict:
    """Embedder load state - used by /modules to poll while a model is being
    pulled from HuggingFace on first vector install.

    Returns the active model name (whichever of ``embedding_model_name`` or
    ``embedding_model_fallback`` successfully loaded), a coarse ``status``
    flag (``ready`` once the singleton is materialised, ``unavailable``
    otherwise - typically while the model is still downloading or after
    both candidates failed to load), and the configured embedding
    dimension.

    Idempotent: if the singleton is already loaded, returns immediately.
    Otherwise touches ``get_embedder()`` which is a no-op after the first
    attempt fails (``_embedder_tried`` short-circuit).
    """
    from app.config import get_settings
    from app.core.vector import active_model_name, get_embedder

    embedder = get_embedder()
    settings = get_settings()
    return {
        "model": active_model_name(),
        "status": "ready" if embedder is not None else "unavailable",
        "dimension": getattr(settings, "embedding_model_dim", 384),
    }


@router.get("/vector/regions/")
async def vector_region_stats() -> list[dict]:
    """Return per-region vector counts from the vector DB.

    Response: ``[{"region": "DE_BERLIN", "count": 55719}, ...]``
    """
    from app.core.vector import vector_status as vs

    status = vs()
    if not status.get("connected"):
        return []

    # LanceDB: query the table directly for per-region counts
    try:
        from app.core.vector import COST_TABLE, _backend, _get_lancedb

        if _backend() != "qdrant":
            db = _get_lancedb()
            if db is None:
                return []
            try:
                tbl = db.open_table(COST_TABLE)
            except Exception:
                logger.debug("LanceDB table %s not found", COST_TABLE)
                return []
            df = tbl.to_pandas()
            if "region" not in df.columns:
                return []
            counts = df.groupby("region").size().reset_index(name="count")
            return [
                {"region": r, "count": int(c)} for r, c in zip(counts["region"], counts["count"], strict=False) if r
            ]
        else:
            # For Qdrant, return total count only (per-region requires scroll)
            col = status.get("cost_collection")
            if col and col.get("vectors_count", 0) > 0:
                return [{"region": "all", "count": col["vectors_count"]}]
            return []
    except Exception:
        logger.debug("Vector stats query failed", exc_info=True)
        return []


@router.get("/vector/v3-status/")
async def vector_v3_status(
    db: SessionDep,
    user: OptionalUserPayload,
    country: str = Query(
        "",
        description=(
            "Region or country code (e.g. DE, DE_BERLIN, USA_USD). "
            "Resolves to the per-language v3 collection - "
            "DE_BERLIN → cwicr_de_v3, USA_USD → cwicr_en_v3, etc. "
            "Empty string returns the engine state without a collection probe."
        ),
    ),
    project_id: uuid.UUID | None = Query(
        None,
        description=(
            "Optional project id. When provided, the response includes "
            "``language_mismatch`` describing whether the project's bound "
            "cost catalogue speaks a different language than the project "
            "region - used to surface a 'wrong catalogue' warning on /match-elements."
        ),
    ),
) -> dict[str, Any]:
    """Per-language CWICR v3 collection readiness for /match-elements.

    Used by the match-elements page to surface a "vector DB ready / missing"
    banner in the same style as the BIM converter status panel. Single
    Qdrant probe - does NOT trigger reindexing; that lives on /costs.

    When ``project_id`` is supplied, also returns ``language_mismatch``
    diagnostics so the UI can warn about a cross-language catalogue
    binding (e.g. a US project bound to RU_MOSCOW would surface Russian
    descriptions). The mismatch is detected by resolving both the project
    region and the bound catalogue id through ``language_for``.
    """
    from app.core.match_service.region_language import language_for
    from app.core.vector import vector_status as vs
    from app.modules.costs.qdrant_adapter import country_to_collection

    base = vs()
    payload: dict[str, Any] = {
        "engine": base.get("engine", "unknown"),
        "connected": bool(base.get("connected")),
        "country": country or "",
        "language": language_for(country) if country else "",
        "collection": "",
        "exists": False,
        "points_count": 0,
        "status_band": "disconnected",
    }

    # Cross-language binding diagnostics - independent of Qdrant probe so
    # the warning fires even when the engine is offline.
    #
    # IDOR gate (Round-7): the diagnostics read MatchProjectSettings for
    # the supplied project. Verify the authenticated caller owns the
    # project (or is admin) before exposing it. Anonymous callers skip
    # the diagnostics entirely - surfacing them anonymously would leak
    # whether arbitrary project UUIDs exist + their bound catalogue.
    if project_id is not None:
        sub = (user or {}).get("sub") if user else None
        if sub:
            try:
                await verify_project_access(project_id, str(sub), db)
            except HTTPException:
                # Caller does not own the project - return the engine
                # status payload WITHOUT the language_mismatch diagnostic
                # so we neither 404 the unrelated probe nor leak the row.
                payload["language_mismatch"] = {
                    "status": "unknown",
                    "project_region": "",
                    "project_language": "",
                    "bound_catalogue": "",
                    "bound_language": "",
                }
            else:
                payload["language_mismatch"] = await _detect_language_mismatch(db, project_id)

    if not payload["connected"]:
        payload["error"] = base.get("error", "")
        return payload

    if not country:
        # Engine reachable but the caller didn't ask about a specific collection.
        payload["status_band"] = "no_country"
        return payload

    payload["collection"] = country_to_collection(country)

    if base.get("engine") != "qdrant":
        # LanceDB or other backend - v3 collection naming doesn't apply.
        payload["status_band"] = "non_qdrant"
        return payload

    try:
        from app.core.vector import _get_qdrant

        client = _get_qdrant()
        if client is None:
            payload["status_band"] = "disconnected"
            return payload
        names = {c.name for c in client.get_collections().collections}
        if payload["collection"] in names:
            payload["exists"] = True
            try:
                col = client.get_collection(payload["collection"])
                # Version-tolerant: ``points_count`` → ``vectors_count``
                # (older qdrant-client) → live count().
                pc_raw = getattr(col, "points_count", None)
                if pc_raw is None:
                    pc_raw = getattr(col, "vectors_count", None)
                if pc_raw is None:
                    pc_raw = client.count(payload["collection"]).count
                pc = int(pc_raw or 0)
                payload["points_count"] = pc
                payload["status_band"] = "ready" if pc > 0 else "empty"
            except Exception:
                payload["status_band"] = "ready"
        else:
            payload["status_band"] = "missing"
    except Exception as exc:
        logger.debug("Qdrant v3 status probe failed", exc_info=True)
        payload["error"] = str(exc)
        payload["status_band"] = "disconnected"

    return payload


async def _detect_language_mismatch(
    db: AsyncSession,
    project_id: uuid.UUID,
) -> dict[str, Any]:
    """Compare the project region's language with the bound catalogue's language.

    Returns a structured payload that the UI can render as either a
    ``ok`` / ``mismatch`` banner. The "mismatch" status fires when the
    bound ``cost_database_id`` resolves to a different ISO-639-1 code
    than the project's region - almost always a sign that
    ``auto_bind_dominant_catalogue`` picked by row count before the
    language-aware fix landed (#236).

    Status values:
        - ``unknown``      - project not found, or no region set
        - ``unbound``      - project has no cost_database_id yet
        - ``ok``           - languages match (or both fall back to default)
        - ``mismatch``     - project language ≠ catalogue language
    """
    from sqlalchemy import select  # noqa: PLC0415

    from app.core.match_service.region_language import language_for
    from app.modules.projects.models import MatchProjectSettings, Project

    out: dict[str, Any] = {
        "status": "unknown",
        "project_region": "",
        "project_language": "",
        "bound_catalogue": "",
        "bound_language": "",
    }
    try:
        project = await db.get(Project, project_id)
        if not project or not project.region:
            return out
        out["project_region"] = project.region
        out["project_language"] = language_for(project.region)

        # MatchProjectSettings uses an ``id`` PK with a unique FK on
        # ``project_id``; ``db.get`` cannot be used here.
        result = await db.execute(select(MatchProjectSettings).where(MatchProjectSettings.project_id == project_id))
        settings = result.scalar_one_or_none()
        if not settings or not settings.cost_database_id:
            out["status"] = "unbound"
            return out
        out["bound_catalogue"] = settings.cost_database_id
        out["bound_language"] = language_for(settings.cost_database_id)

        if out["project_language"] and out["bound_language"]:
            out["status"] = "ok" if out["project_language"] == out["bound_language"] else "mismatch"
    except Exception:  # pragma: no cover - defensive
        logger.debug("language mismatch probe failed", exc_info=True)
    return out


@router.get("/embedder/status/")
async def embedder_status() -> dict[str, Any]:
    """Free / open-source language model readiness for /match-elements.

    Surfaces enough information for the UI to render a "language model
    required" panel when the optional ``[semantic]`` extra is missing,
    plus reassurance ("MIT, multilingual, 100+ languages, runs locally")
    so users understand it's a one-time install of a free model rather
    than a paid API.

    Returns
    -------
    {
        "installed": bool,            # FlagEmbedding importable?
        "model_loaded": bool,         # encoder initialised in this process?
        "model_name": str,            # configured HF id, e.g. "BAAI/bge-m3"
        "model_id_runtime": str,      # actual HF id used at runtime
                                      # (may differ when int8_mode is True)
        "license": "MIT",
        "open_source": True,
        "homepage": str,
        "languages_supported": int,   # 100+ via BGE-M3
        "size_mb_int8": int,          # ONNX INT8 footprint
        "size_mb_fp32": int,          # full-precision footprint
        "int8_mode": bool,            # current setting
        "pip_command": str,           # one-liner the UI shows in a copy box
        "missing_packages": list[str],
        "extra_name": "semantic",     # hint for advanced users
    }

    Always returns 200 - the UI distinguishes states from the payload,
    not from HTTP status, so a missing-extra install can render a clean
    install card instead of an error toast.
    """
    from app.config import get_settings  # noqa: PLC0415

    s = get_settings()
    model_name = getattr(s, "cwicr_embedding_model", "BAAI/bge-m3")
    int8_mode = bool(getattr(s, "cwicr_embedding_int8", True))
    runtime_id = "gpahal/bge-m3-onnx-int8" if int8_mode else model_name

    missing: list[str] = []
    installed = False
    try:
        import FlagEmbedding  # type: ignore[import-not-found]  # noqa: F401, PLC0415

        installed = True
    except ImportError:
        missing.append("FlagEmbedding")
    try:
        import qdrant_client  # type: ignore[import-not-found]  # noqa: F401, PLC0415
    except ImportError:
        missing.append("qdrant-client")

    # Probe whether the encoder has been initialised in this worker
    # without forcing a load - qdrant_adapter._encoder is a module-level
    # singleton that is None until the first /qdrant-search hits.
    model_loaded = False
    try:
        from app.modules.costs import qdrant_adapter  # noqa: PLC0415

        model_loaded = qdrant_adapter._encoder is not None  # type: ignore[attr-defined]
    except Exception:
        model_loaded = False

    return {
        "installed": installed and not missing,
        "model_loaded": model_loaded,
        "model_name": model_name,
        "model_id_runtime": runtime_id,
        "license": "MIT",
        "open_source": True,
        "homepage": f"https://huggingface.co/{model_name}",
        "languages_supported": 100,
        "size_mb_int8": 700,
        "size_mb_fp32": 2300,
        "int8_mode": int8_mode,
        "pip_command": "pip install --upgrade openconstructionerp[semantic]",
        "missing_packages": missing,
        "extra_name": "semantic",
    }


@router.get("/qdrant-search/")
async def qdrant_smoke_search(
    q: str = Query(..., min_length=1, description="Query text - passed verbatim as the CORE query"),
    country: str = Query("DE", description="Region or country code, e.g. DE, DE_BERLIN, USA_USD"),
    limit: int = Query(10, ge=1, le=500),
    is_abstract: bool | None = Query(False, description="Drop aggregator headers (None to leave open)"),
    department_code: str | None = Query(None, description="DIN-276-derived trade bucket (optional)"),
    unit_dim: str | None = Query(None, description="volume / area / length / count (optional)"),
    diag: bool = Query(False, description="Return diagnostics (resolved collection + parquet path)"),
) -> dict[str, Any]:
    """Smoke endpoint for the new BGE-M3 + Qdrant CWICR pipeline.

    One-shot hybrid search: dense + sparse fused via Qdrant native RRF,
    then parquet lookup attaches the 84-column rate data. Use this to
    verify the new pipeline before wiring it into ``/match-elements``.

    Example:
        GET /api/v1/costs/qdrant-search/?q=Stahlbetonwand%20C30/37&country=DE
    """

    from app.modules.costs.parquet_lookup import parquet_path_for_country, parquet_root
    from app.modules.costs.qdrant_adapter import (
        country_to_collection,
        lookup_full_rows,
        search,
    )

    filters: dict[str, Any] = {}
    if is_abstract is not None:
        filters["is_abstract"] = is_abstract
    if department_code:
        filters["department_code"] = department_code
    if unit_dim:
        filters["unit_dim"] = unit_dim

    try:
        hits = await search(
            country=country,
            core_query=q,
            filters=filters,
            limit=limit,
        )
    except (ImportError, ModuleNotFoundError) as exc:
        # The optional [semantic] extra (qdrant_client / FlagEmbedding) is
        # not installed. A lazy ``from qdrant_client...`` deep inside
        # search() raised a bare ModuleNotFoundError - never echo the raw
        # "No module named 'qdrant_client'" text to the client (NEW-B-105).
        logger.info("CWICR Qdrant search unavailable (optional extra missing): %s", exc)
        raise HTTPException(
            status_code=503,
            detail=(
                "Semantic search is not available on this deployment. "
                "Install the optional extra: pip install openconstructionerp[semantic]"
            ),
        ) from exc
    except RuntimeError as exc:
        # Optional [semantic] extra missing or no Qdrant reachable.
        raise HTTPException(status_code=503, detail=str(exc)) from exc
    except Exception as exc:  # noqa: BLE001
        logger.exception("CWICR Qdrant smoke search failed")
        raise HTTPException(status_code=500, detail="qdrant search failed (see server logs)") from exc

    rate_codes = [h.rate_code for h in hits]
    full_rows = await lookup_full_rows(country=country, rate_codes=rate_codes)
    full_by_code = {str(r.get("rate_code")): r for r in full_rows}

    response_hits = [
        {
            "rate_code": h.rate_code,
            "country": h.country,
            "score": h.score,
            "payload": h.payload,
            "full": full_by_code.get(h.rate_code),
        }
        for h in hits
    ]

    body: dict[str, Any] = {"hits": response_hits, "count": len(response_hits)}
    if diag:
        body["diagnostics"] = {
            "collection": country_to_collection(country),
            "parquet_root": str(parquet_root()),
            "parquet_file": str(parquet_path_for_country(country) or ""),
            "parquet_rows_attached": len(full_rows),
        }
    return body


@router.post(
    "/vector/index/",
    dependencies=[Depends(RequirePermission("costs.create"))],
    # Either ``JSONResponse`` (when the vector backend is unavailable → 503)
    # or a plain ``dict`` (happy path). FastAPI can't auto-derive a single
    # response model from that union, and we want the bare-dict happy-path
    # serialisation untouched - so we opt out of response-model generation.
    response_model=None,
)
async def vectorize_cost_items(
    session: SessionDep,
    _user_id: CurrentUserId,
    region: str | None = Query(default=None, description="Only index items from this region"),
    batch_size: int = Query(default=256, ge=32, le=1024),
) -> JSONResponse | dict:
    """Generate embeddings and index cost items into vector DB.

    Thin HTTP wrapper around :func:`vectorize_region`; the work lives in that
    module-level helper so the partner-pack one-click installer can build the
    vector DB through the same path without going through HTTP.
    """
    return await vectorize_region(session, region=region, batch_size=batch_size)


async def vectorize_region(
    session: AsyncSession,
    *,
    region: str | None = None,
    batch_size: int = 256,
) -> JSONResponse | dict:
    """Embed and index the cost items of one region into the vector DB.

    Uses FastEmbed/ONNX (all-MiniLM-L6-v2, 384d) locally - no API key needed.
    Default backend: LanceDB (embedded, no Docker required).

    Returns ``503 Service Unavailable`` (as a ``JSONResponse``) when the vector
    backend (Qdrant / LanceDB / embedding model) is not reachable or not
    installed. The body keeps the legacy ``{"indexed": 0, "message": ...,
    "error": ...}`` shape so existing clients still parse it. The happy path
    returns a plain ``dict``. Reusable building block shared by the
    ``POST /vector/index/`` route and the partner-pack ``full-install``
    orchestrator (which treats a 503 as graceful degradation, not a failure).
    """
    import asyncio
    import time

    from sqlalchemy import select

    # Quick check: can we even import the vector module?
    try:
        from app.core.vector import encode_texts, get_embedder, vector_index
    except Exception as exc:
        logger.warning("Vector module import failed: %s", exc)
        return JSONResponse(
            content={
                "indexed": 0,
                "message": "Vector indexing is not available: vector module failed to load.",
                "error": str(exc),
            },
            status_code=503,
        )

    # Verify embedding model is loadable (run in thread with short timeout
    # so a slow model download doesn't hang the request indefinitely).
    try:
        embedder = await asyncio.wait_for(asyncio.to_thread(get_embedder), timeout=30)
        if embedder is None:
            return JSONResponse(
                content={
                    "indexed": 0,
                    "message": "Vector indexing is not available: no embedding model "
                    "found. Install sentence-transformers (pip install "
                    "sentence-transformers).",
                },
                status_code=503,
            )
    except TimeoutError:
        return JSONResponse(
            content={
                "indexed": 0,
                "message": "Vector indexing is not available: embedding model loading "
                "timed out. The model may need to be downloaded first - try again later.",
            },
            status_code=503,
        )
    except Exception as exc:
        logger.warning("Embedding model check failed: %s", exc)
        return JSONResponse(
            content={
                "indexed": 0,
                "message": f"Vector indexing is not available: {exc}",
            },
            status_code=503,
        )

    from app.modules.costs.models import CostItem

    start = time.monotonic()

    # Fetch cost items
    stmt = select(CostItem).where(CostItem.is_active.is_(True))
    if region:
        stmt = stmt.where(CostItem.region == region)

    result = await session.execute(stmt)
    items = result.scalars().all()

    if not items:
        return {"indexed": 0, "message": "No cost items found to index"}

    logger.info("Vectorizing %d cost items (region=%s)...", len(items), region or "all")

    # Pre-extract all data from ORM objects before they expire
    items_data = []
    for item in items:
        cls = item.classification or {}
        items_data.append(
            {
                "id": str(item.id),
                "code": item.code,
                "description": (item.description or "")[:200],
                "unit": item.unit or "",
                "rate": float(item.rate) if item.rate else 0.0,
                "region": item.region or "",
                "text": " ".join(
                    p
                    for p in [
                        item.description or "",
                        item.unit or "",
                        cls.get("collection", ""),
                        cls.get("department", ""),
                        cls.get("section", ""),
                    ]
                    if p
                ),
            }
        )

    # Run CPU-heavy embedding in a thread to not block event loop.
    # NOTE: Uses ThreadPoolExecutor (not Process) to avoid pickling issues
    # with global model singletons and LanceDB connections.
    def _vectorize_batch(data: list[dict], bs: int) -> int:
        total = 0
        for i in range(0, len(data), bs):
            batch = data[i : i + bs]
            texts = [d["text"] for d in batch]
            vectors = encode_texts(texts)
            records = [
                {**{k: d[k] for k in ("id", "code", "description", "unit", "rate", "region")}, "vector": vectors[j]}
                for j, d in enumerate(batch)
            ]
            total += vector_index(records)
        return total

    try:
        indexed = await asyncio.to_thread(_vectorize_batch, items_data, batch_size)
    except Exception as exc:
        # Graceful error when vector backend is unavailable:
        # - RuntimeError: no embedding model or LanceDB not installed
        # - ImportError: sentence-transformers / lancedb not installed
        # - Any other error during vectorization
        logger.warning("Vector indexing failed: %s", exc)
        return {
            "indexed": 0,
            "message": f"Vector indexing failed: {exc}",
        }

    duration = round(time.monotonic() - start, 1)
    logger.info("Indexed %d cost items in %.1fs", indexed, duration)

    return {
        "indexed": indexed,
        "region": region or "all",
        "duration_seconds": duration,
    }


@router.post(
    "/vector/load-github/{db_id}",
    dependencies=[Depends(RequirePermission("costs.create"))],
)
async def load_vector_from_github(
    db_id: str,
    session: SessionDep,
    _user_id: CurrentUserId,
) -> dict:
    """Download pre-built vector embeddings from GitHub and index into LanceDB.

    Downloads a parquet file with pre-computed 384d embeddings (all-MiniLM-L6-v2)
    for the given region, so users don't need to run the embedding model locally.
    """
    import time

    from app.core.vector import vector_index

    start = time.monotonic()

    github_path = _GITHUB_CWICR_FILES.get(db_id)
    if not github_path:
        raise HTTPException(404, f"Unknown database ID: {db_id}")

    # Vector parquet is stored alongside the regular parquet in the same repo
    vector_filename = f"{db_id}_vectors.parquet"
    vector_github_path = f"{db_id}/{vector_filename}"
    url = f"{_GITHUB_CWICR_BASE_URL}/{vector_github_path}"

    # Cache locally
    cache_dir = _CWICR_CACHE_DIR / "vectors"
    cache_dir.mkdir(parents=True, exist_ok=True)
    local_path = cache_dir / vector_filename

    # Download if not cached
    github_available = False
    if not local_path.exists() or local_path.stat().st_size < 1000:
        logger.info("Downloading vector data for %s from GitHub: %s", db_id, url)
        try:
            _download_to_file(url, local_path)
            if local_path.exists() and local_path.stat().st_size > 1000:
                github_available = True
        except Exception as exc:
            local_path.unlink(missing_ok=True)
            logger.info("GitHub vectors not available for %s, will generate locally: %s", db_id, exc)
    else:
        github_available = True

    # Fallback: generate vectors locally from cost items in DB
    if not github_available:
        logger.info("Generating vectors locally for %s from cost database", db_id)
        try:
            import asyncio
            from concurrent.futures import ThreadPoolExecutor

            from app.core.vector import encode_texts
            from app.core.vector import vector_index as vi
            from app.modules.costs.repository import CostItemRepository

            repo = CostItemRepository(session)
            items_list, total = await repo.search(region=db_id, limit=5000)
            if not items_list:
                items_list, total = await repo.search(limit=5000)

            if not items_list:
                raise HTTPException(400, f"No cost items found for '{db_id}'.")

            # Run embedding generation in a thread to not block event loop
            def _generate_vectors(items_data):
                batch_size = 128
                indexed = 0
                for i in range(0, len(items_data), batch_size):
                    batch = items_data[i : i + batch_size]
                    texts = [f"{it['code']} {it['desc']}" for it in batch]
                    vectors = encode_texts(texts)
                    records = [
                        {
                            "id": it["id"],
                            "vector": vec,
                            "code": it["code"],
                            "description": it["desc"],
                            "unit": it["unit"],
                            "rate": it["rate"],
                            "region": it["region"],
                        }
                        for it, vec in zip(batch, vectors, strict=False)
                    ]
                    if records:
                        indexed += vi(records)
                return indexed

            # Prepare data outside the thread (ORM objects can't cross threads)
            items_data = [
                {
                    "id": str(ci.id),
                    "code": ci.code or "",
                    "desc": (ci.description or "")[:200],
                    "unit": ci.unit or "",
                    "rate": float(ci.rate) if ci.rate else 0.0,
                    "region": ci.region or db_id,
                }
                for ci in items_list
            ]

            loop = asyncio.get_event_loop()
            with ThreadPoolExecutor(max_workers=1) as pool:
                indexed = await loop.run_in_executor(pool, _generate_vectors, items_data)

            duration = round(time.monotonic() - start, 1)
            logger.info("Generated %d vectors locally for %s in %.1fs", indexed, db_id, duration)
            return {
                "indexed": indexed,
                "database": db_id,
                "source": "local",
                "duration_seconds": duration,
            }
        except HTTPException:
            raise
        except Exception as gen_err:
            logger.exception("Failed to generate vectors for %s", db_id)
            raise HTTPException(
                500,
                f"Vector generation failed for '{db_id}': {gen_err}. "
                f"Ensure sentence-transformers and lancedb are installed.",
            ) from gen_err

    logger.info("Loading vector data from %s", local_path)

    # Read parquet: columns = id, vector, code, description, unit, rate, region
    import pandas as pd

    df = pd.read_parquet(local_path)
    total = len(df)

    if total == 0:
        return {"indexed": 0, "database": db_id, "message": "Empty vector file"}

    # Index in batches
    batch_size = 256
    indexed = 0
    for i in range(0, total, batch_size):
        batch = df.iloc[i : i + batch_size]
        records = []
        for _, row in batch.iterrows():
            vec = row.get("vector")
            if vec is None:
                continue
            # Convert numpy array to list if needed
            if hasattr(vec, "tolist"):
                vec = vec.tolist()
            elif isinstance(vec, str):
                import json

                vec = json.loads(vec)

            records.append(
                {
                    "id": str(row.get("id", "")),
                    "vector": vec,
                    "code": str(row.get("code", "")),
                    "description": str(row.get("description", ""))[:200],
                    "unit": str(row.get("unit", "")),
                    "rate": float(row.get("rate", 0)),
                    "region": str(row.get("region", db_id)),
                }
            )

        if records:
            indexed += vector_index(records)

    duration = round(time.monotonic() - start, 1)
    logger.info("Loaded %d vectors for %s from GitHub in %.1fs", indexed, db_id, duration)

    return {
        "indexed": indexed,
        "database": db_id,
        "source": "github",
        "duration_seconds": duration,
    }


# Mapping db_id to GitHub folder and snapshot filename (3072d embeddings)
_GITHUB_SNAPSHOT_FILES: dict[str, str] = {
    "USA_USD": "US___DDC_CWICR/USA_USD_workitems_costs_resources_EMBEDDINGS_3072_DDC_CWICR.snapshot",
    "UK_GBP": "UK___DDC_CWICR/UK_GBP_workitems_costs_resources_EMBEDDINGS_3072_DDC_CWICR.snapshot",
    "DE_BERLIN": "DE___DDC_CWICR/DE_BERLIN_workitems_costs_resources_EMBEDDINGS_3072_DDC_CWICR.snapshot",
    "ENG_TORONTO": "EN___DDC_CWICR/EN_TORONTO_workitems_costs_resources_EMBEDDINGS_3072_DDC_CWICR.snapshot",
    # CA_TORONTO alias - same canonical id as the cost-DB map above.
    "CA_TORONTO": "EN___DDC_CWICR/EN_TORONTO_workitems_costs_resources_EMBEDDINGS_3072_DDC_CWICR.snapshot",
    "FR_PARIS": "FR___DDC_CWICR/FR_PARIS_workitems_costs_resources_EMBEDDINGS_3072_DDC_CWICR.snapshot",
    "SP_BARCELONA": "ES___DDC_CWICR/SP_BARCELONA_workitems_costs_resources_EMBEDDINGS_3072_DDC_CWICR.snapshot",
    "PT_SAOPAULO": "PT___DDC_CWICR/PT_SAOPAULO_workitems_costs_resources_EMBEDDINGS_3072_DDC_CWICR.snapshot",
    "RU_STPETERSBURG": "RU___DDC_CWICR/RU_STPETERSBURG_workitems_costs_resources_EMBEDDINGS_3072_DDC_CWICR.snapshot",
    "AR_DUBAI": "AR___DDC_CWICR/AR_DUBAI_workitems_costs_resources_EMBEDDINGS_3072_DDC_CWICR.snapshot",
    "ZH_SHANGHAI": "ZH___DDC_CWICR/ZH_SHANGHAI_workitems_costs_resources_EMBEDDINGS_3072_DDC_CWICR.snapshot",
    "ZH_CHINA": "ZH___DDC_CWICR/ZH_SHANGHAI_workitems_costs_resources_EMBEDDINGS_3072_DDC_CWICR.snapshot",
    "HI_MUMBAI": "HI___DDC_CWICR/HI_MUMBAI_workitems_costs_resources_EMBEDDINGS_3072_DDC_CWICR.snapshot",
}


@router.post(
    "/vector/restore-snapshot/{db_id}",
    dependencies=[Depends(RequirePermission("costs.create"))],
)
async def restore_qdrant_snapshot(
    db_id: str,
    _user_id: CurrentUserId,
) -> dict:
    """Download a pre-built Qdrant snapshot from GitHub and restore it.

    Downloads 3072d embeddings snapshot (~1.1 GB) and restores into Qdrant.
    Requires Qdrant server running (Docker or binary).
    """
    import asyncio
    import time

    from app.core.vector import _get_qdrant

    start = time.monotonic()

    client = _get_qdrant()
    if client is None:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            "Qdrant not available. Start Qdrant: docker run -p 6333:6333 qdrant/qdrant",
        )

    snapshot_path = _GITHUB_SNAPSHOT_FILES.get(db_id)
    if not snapshot_path:
        available = ", ".join(sorted(_GITHUB_SNAPSHOT_FILES.keys()))
        raise HTTPException(
            status.HTTP_404_NOT_FOUND,
            f"Unknown database ID: {db_id}. Available: {available}",
        )

    url = f"{_GITHUB_CWICR_BASE_URL}/{snapshot_path}"

    # Cache locally to avoid re-downloading ~1.1 GB
    cache_dir = Path.home() / ".openestimator" / "cache" / "snapshots"
    cache_dir.mkdir(parents=True, exist_ok=True)
    local_path = cache_dir / f"{db_id}.snapshot"

    # Download if not already cached
    if not local_path.exists() or local_path.stat().st_size < 10_000:
        logger.info(
            "Downloading Qdrant snapshot for %s from GitHub (~1.1 GB): %s",
            db_id,
            url,
        )
        try:
            loop = asyncio.get_event_loop()
            await loop.run_in_executor(
                None,
                _download_to_file,
                url,
                local_path,
                600.0,
            )
        except Exception as exc:
            local_path.unlink(missing_ok=True)
            logger.error("Failed to download snapshot for %s: %s", db_id, exc)
            raise HTTPException(
                status.HTTP_502_BAD_GATEWAY,
                f"Failed to download snapshot from GitHub: {exc}",
            )

        if not local_path.exists() or local_path.stat().st_size < 10_000:
            local_path.unlink(missing_ok=True)
            raise HTTPException(
                status.HTTP_502_BAD_GATEWAY,
                "Downloaded snapshot file is too small or missing. The file may not exist on GitHub yet.",
            )
        logger.info(
            "Snapshot downloaded for %s: %.1f MB",
            db_id,
            local_path.stat().st_size / (1024 * 1024),
        )
    else:
        logger.info(
            "Using cached snapshot for %s: %s (%.1f MB)",
            db_id,
            local_path,
            local_path.stat().st_size / (1024 * 1024),
        )

    # Restore snapshot into Qdrant
    collection_name = f"cwicr_{db_id.lower()}"

    try:
        from qdrant_client.models import Distance, VectorParams

        # Create collection if it does not already exist
        existing = [c.name for c in client.get_collections().collections]
        if collection_name not in existing:
            client.create_collection(
                collection_name,
                vectors_config=VectorParams(size=3072, distance=Distance.COSINE),
            )
            logger.info("Created Qdrant collection: %s", collection_name)

        # recover_snapshot() can only pull from a URI the Qdrant SERVER itself
        # reaches (http://, s3://, or file:// on its own disk). The snapshot
        # we hold lives on the app container instead, so push the raw bytes
        # straight to /collections/{name}/snapshots/upload over multipart.
        from app.modules.costs import qdrant_snapshot_loader

        target_url = _v3_qdrant_url()
        if not target_url:  # nowhere to send the restore request
            detail = "Qdrant URL not configured - set QDRANT_URL or CWICR_QDRANT_URL"
            raise HTTPException(status.HTTP_500_INTERNAL_SERVER_ERROR, detail)
        restore_kwargs = {
            "qdrant_url": target_url,
            "collection_name": collection_name,
            "snapshot_path": local_path,
            "timeout_s": 1800,
        }
        restored = await asyncio.to_thread(qdrant_snapshot_loader.restore_snapshot_file, **restore_kwargs)
        if not restored:
            detail = f"Failed to restore Qdrant snapshot for {db_id}. Check Qdrant logs."
            raise HTTPException(status.HTTP_500_INTERNAL_SERVER_ERROR, detail)
        logger.info("Snapshot restored for collection %s", collection_name)
    except Exception as exc:
        logger.error("Failed to restore snapshot for %s: %s", db_id, exc)
        raise HTTPException(
            status.HTTP_500_INTERNAL_SERVER_ERROR,
            f"Failed to restore Qdrant snapshot: {exc}",
        )

    duration = round(time.monotonic() - start, 1)

    # Get collection info after restore. ``CollectionInfo.vectors_count``
    # was removed in newer qdrant-client - read ``points_count`` first and
    # fall back resiliently so a client version bump can't crash this.
    try:
        col_info = client.get_collection(collection_name)
        vectors_count = getattr(col_info, "points_count", None)
        if vectors_count is None:
            vectors_count = getattr(col_info, "vectors_count", None)
        if vectors_count is None:
            try:
                vectors_count = client.count(collection_name).count
            except Exception:
                vectors_count = None
    except Exception:
        vectors_count = None

    logger.info(
        "Qdrant snapshot restore complete for %s: collection=%s, vectors=%s, duration=%.1fs",
        db_id,
        collection_name,
        vectors_count,
        duration,
    )

    return {
        "restored": True,
        "collection": collection_name,
        "database": db_id,
        "vectors_count": vectors_count,
        "source": "github_snapshot",
        "duration_seconds": duration,
    }


# ── v3 BGE-M3 catalogues (UI-facing install flow) ─────────────────────────
#
# `/vector/restore-snapshot/{db_id}` above ships the legacy 3072-dim snapshots
# and writes into per-region collections (`cwicr_de_berlin`, …). The v3 path
# below ships BGE-M3 v3 snapshots and writes into per-language collections
# (`cwicr_de_v3`, …) - the schema the production /match-elements pipeline
# already searches against. Once Phase 5 cleanup retires the 3072 path, the
# legacy endpoint goes away; until then they coexist so existing installs
# don't break.


def _v3_snapshot_cache_path(region: str) -> Path:
    """Local cache path for a v3 snapshot file. One per CWICR region.

    Cache lives under ``~/.openestimator/cache/snapshots-v3/`` so it
    doesn't collide with the 3072-dim cache (``cache/snapshots/``) used
    by the legacy endpoint. Re-using the same file across restores
    avoids re-downloading 400+ MB on retry.
    """
    return Path.home() / ".openestimator" / "cache" / "snapshots-v3" / f"{region}.snapshot"


def _snapshot_error_hint(err: str) -> str | None:
    """Map Qdrant's recover-from-URL error to a user-actionable hint.

    Qdrant 1.18+ on native Windows reliably trips ``os error 5`` on
    ``newest_clocks.json`` (and occasionally other clock/WAL files)
    during snapshot recovery because Windows Defender real-time
    scanning holds a handle on the file Qdrant has just written and
    is trying to ``fsync``. The download succeeds - only the final
    sync fails - so the user sees "could not fetch or restore" with
    no obvious cause. Surface a concrete fix instead of leaving them
    to grep the backend log.
    """
    if not err:
        return None
    low = err.lower()
    if "os error 5" in low or "access is denied" in low:
        return (
            "Windows Defender is locking files in Qdrant's storage folder during "
            "fsync. The download succeeded - only the final disk write was blocked. "
            "Fix: open PowerShell AS ADMINISTRATOR and run:\n"
            '  Add-MpPreference -ExclusionPath "$env:USERPROFILE\\.openestimator"\n'
            "Then click Install again. (No restart needed - Qdrant picks it up "
            "on the next attempt.) GUI alternative: Settings → Update & Security → "
            "Windows Security → Virus & threat protection → Manage settings → "
            "Add or remove exclusions → Folder → pick %USERPROFILE%\\.openestimator."
        )
    if "no space left" in low or "out of space" in low or "disk full" in low:
        return (
            "Disk is full - the BGE-M3 snapshot needs ~1–2 GB free during restore. "
            "Free up space on the drive holding ~/.openestimator and retry."
        )
    if "status - 404" in low or "404 not found" in low:
        return (
            "The DDC catalogue file is missing on HuggingFace. The region may "
            "still be marked 'available' in the registry before DDC publishes. "
            "Check huggingface.co/datasets/DataDrivenConstruction/cwicr-vector-db-bgem3-v3."
        )
    if "timed out" in low or "timeout" in low or "connection refused" in low:
        return (
            "Qdrant timed out fetching the snapshot. Verify outbound HTTPS to "
            "huggingface.co works from the Qdrant host, then retry."
        )
    return None


def _v3_qdrant_url() -> str | None:
    """Resolve the server-mode Qdrant URL for the v3 catalogue path.

    Prefers ``settings.cwicr_qdrant_url`` (the dedicated v3 setting);
    falls back to ``settings.qdrant_url`` which the legacy adapter uses
    - in single-server dev they point at the same instance and the
    fallback removes one configuration step. Returns ``None`` when
    neither is set so the caller can surface a clear "no server" error.
    """
    from app.config import get_settings

    s = get_settings()
    return getattr(s, "cwicr_qdrant_url", None) or getattr(s, "qdrant_url", None)


def _demo_mode_enabled() -> bool:
    """True on the public hosted demo (env ``OE_DEMO_MODE``).

    Mirrors the flag exposed by ``GET /api/system/status``. On the demo we
    block catalogue downloads: each DDC v3 snapshot is 400-900 MB and the
    shared demo box has neither the disk nor the single-core CPU budget to
    host them. Self-hosted installs leave ``OE_DEMO_MODE`` unset and keep
    full access to every published catalogue.
    """
    import os

    return os.environ.get("OE_DEMO_MODE", "").lower() in ("1", "true", "yes")


@router.get("/catalogues-v3/")
async def list_v3_catalogues() -> dict:
    """List the 30 CWICR v3 catalogues with install status.

    Frontend powers the `/setup/databases` "Quick install from DDC" grid
    from this endpoint. Each row gets a flag, name, currency, size,
    and a status suitable for picking the right CTA:

    * ``loaded`` - the collection exists on the configured Qdrant server
      and has at least one point. Multiple regions sharing a language
      (USA_USD + GB_LONDON → cwicr_en_v3) all report ``loaded`` because
      the search-time collection is populated; per-region cache state
      is exposed separately as ``snapshot_cached``.
    * ``installing`` - reserved; current implementation runs the install
      synchronously, so the only way to see this state is via the
      transient cache file probe. Kept in the schema so the frontend
      can reuse it once we move to background jobs.
    * ``available`` - DDC has published the v3 snapshot but it's not
      installed on this server. The "Install" CTA is enabled.
    * ``coming_soon`` - registry knows the region but the snapshot
      hasn't shipped yet. CTA is disabled with a "coming soon" hint.
    """
    from app.modules.costs.cwicr_v3_catalogue import CWICR_V3_CATALOGUES

    qdrant_url = _v3_qdrant_url()

    # Probe Qdrant once for the whole list. A single REST call is much
    # cheaper than per-region collection lookups, especially for the 30
    # cards where most regions share a language collection anyway.
    server_collections: set[str] = set()
    server_reachable = False
    if qdrant_url:
        try:
            from app.modules.costs.qdrant_snapshot_loader import server_collections as _probe

            server_collections = set(_probe(qdrant_url=qdrant_url))
            server_reachable = True
        except Exception as exc:  # pragma: no cover - defensive
            logger.debug("v3 catalogues: server probe failed: %s", exc)

    catalogues: list[dict] = []
    for cat in CWICR_V3_CATALOGUES:
        cache_path = _v3_snapshot_cache_path(cat.region)
        snapshot_cached = cache_path.exists() and cache_path.stat().st_size > 1_000_000

        if cat.collection in server_collections:
            install_status = "loaded"
        elif not cat.available:
            install_status = "coming_soon"
        else:
            install_status = "available"

        catalogues.append(
            {
                "region": cat.region,
                "country_iso": cat.country_iso,
                "city": cat.city,
                "language": cat.language,
                "currency": cat.currency,
                "collection": cat.collection,
                "size_mb": cat.size_mb,
                "available": cat.available,
                "snapshot_cached": snapshot_cached,
                "install_status": install_status,
            }
        )

    # Sort: loaded → available → coming_soon, then alpha by country_iso
    # within each bucket. The UI can re-sort, but a sensible default
    # surfaces actionable rows at the top.
    _STATUS_ORDER = {"loaded": 0, "available": 1, "installing": 2, "coming_soon": 3}
    catalogues.sort(key=lambda c: (_STATUS_ORDER.get(c["install_status"], 9), c["country_iso"]))

    return {
        "catalogues": catalogues,
        # Public-demo flag so the frontend can disable installs and show the
        # "download to self-host" notice. False on every normal install.
        "demo_mode": _demo_mode_enabled(),
        "server": {
            "url": qdrant_url,
            "reachable": server_reachable,
            "total_collections": len(server_collections),
            "v3_collections": sorted(c for c in server_collections if c.endswith("_v3")),
        },
    }


@router.post(
    "/catalogues-v3/{region}/install",
    dependencies=[Depends(RequirePermission("costs.create"))],
)
async def install_v3_catalogue(
    region: str,
    _user_id: CurrentUserId,
) -> dict:
    """Install a DDC v3 BGE-M3 snapshot into Qdrant via recover-from-URL.

    Synchronous: the call returns when the restore completes, which for
    a 400 MB snapshot is typically 5–15 minutes on a decent connection
    (the Qdrant server downloads from HuggingFace and restores inline).
    Returns ``409`` when the catalogue isn't yet published by DDC,
    ``404`` for unknown region ids, ``503`` when no Qdrant server is
    reachable, ``502`` when Qdrant could not fetch or restore the file.

    Implementation note: this endpoint deliberately does NOT use
    Qdrant's multipart ``/snapshots/upload`` because its default
    ``service.max_request_size_mb`` is 32 MB - every v3 BGE-M3 snapshot
    is several times that. The recover-from-URL endpoint
    (``PUT /collections/{name}/snapshots/recover``) has Qdrant download
    the file itself with no body-size ceiling.
    """
    import asyncio
    import time

    from app.modules.costs.cwicr_v3_catalogue import get_catalogue
    from app.modules.costs.qdrant_snapshot_loader import (
        SnapshotRestoreError,
        restore_snapshot_from_url,
        server_collections,
    )

    cat = get_catalogue(region)
    if cat is None:
        raise HTTPException(
            status.HTTP_404_NOT_FOUND,
            f"Unknown CWICR region: {region}. See GET /catalogues-v3/ for the list.",
        )
    if not cat.available:
        raise HTTPException(
            status.HTTP_409_CONFLICT,
            f"DDC has not yet published the v3 snapshot for {region}. "
            "Track the catalogue in /setup/databases and try again later.",
        )

    # Public demo: block catalogue downloads. Every v3 snapshot is 400-900 MB
    # and the shared demo box (small disk, single core) cannot host them. The
    # demo ships its own small reference cost index; loading additional cost
    # databases is a self-host feature.
    if _demo_mode_enabled():
        raise HTTPException(
            status.HTTP_403_FORBIDDEN,
            "Installing cost databases is disabled on the public demo. Each "
            "catalogue is 400+ MB and the demo runs on a small shared box. To "
            "load this region or your own cost data, download and run "
            "OpenConstructionERP on your own machine: pip install openconstructionerp.",
        )

    qdrant_url = _v3_qdrant_url()
    if not qdrant_url:
        raise HTTPException(
            status.HTTP_503_SERVICE_UNAVAILABLE,
            "No Qdrant server configured. Set CWICR_QDRANT_URL or QDRANT_URL "
            "and ensure the server is reachable (docker compose up -d qdrant).",
        )

    start = time.monotonic()

    # Build the public DDC URL the Qdrant server itself will fetch. Two
    # host layouts coexist in the registry:
    #   • Legacy: `<XX>___DDC_CWICR/<region>_workitems_…` lives on
    #     GitHub LFS at _GITHUB_CWICR_BASE_URL.
    #   • Current: `<XX>/<stem>_workitems_…` lives on HuggingFace under
    #     HF_CWICR_BASE_URL. DDC stopped publishing to GitHub LFS in v3
    #     because the LFS bandwidth cap kept getting hit.
    # Detection is on the path shape, not on `cat.region`, so adding new
    # HF rows in cwicr_v3_catalogue.py needs no router changes.
    from app.modules.costs.cwicr_v3_catalogue import HF_CWICR_BASE_URL

    snapshot_url = (
        f"{_GITHUB_CWICR_BASE_URL}/{cat.ddc_path}"
        if "___DDC_CWICR" in cat.ddc_path
        else f"{HF_CWICR_BASE_URL}/{cat.ddc_path}"
    )

    # Restore via the recover-from-URL path. Qdrant's multipart
    # ``snapshots/upload`` endpoint is bounded by
    # ``service.max_request_size_mb`` (default 32 MB) - every BGE-M3 v3
    # snapshot is 400–800 MB and gets rejected with HTTP 500 "An error
    # occurred processing field: snapshot". The recover-from-URL
    # endpoint instead has Qdrant download the file itself with no body
    # size limit, then restore inline. The Qdrant call is synchronous -
    # it returns once the restore finishes (5–15 min on a typical link
    # for 400 MB). We run it in the executor so the event loop stays
    # responsive for other tenants' requests during the wait.
    logger.info(
        "Installing v3 snapshot %s via Qdrant recover-from-URL (~%d MB): %s",
        cat.region,
        cat.size_mb,
        snapshot_url,
    )
    loop = asyncio.get_event_loop()
    try:
        await loop.run_in_executor(
            None,
            lambda: restore_snapshot_from_url(
                qdrant_url=qdrant_url,
                collection_name=cat.collection,
                snapshot_url=snapshot_url,
                # ~25 min budget: HF can be slow to first-byte for cold
                # objects, and the restore step itself is CPU+disk
                # bound on the Qdrant side.
                timeout_s=1500,
            ),
        )
    except SnapshotRestoreError as exc:
        qdrant_err = str(exc)
        hint = _snapshot_error_hint(qdrant_err)
        raise HTTPException(
            status.HTTP_502_BAD_GATEWAY,
            (
                f"Qdrant could not restore the snapshot from {snapshot_url}.\n"
                f"Qdrant said: {qdrant_err}" + (f"\nHint: {hint}" if hint else "")
            ),
        ) from exc
    except RuntimeError as exc:
        raise HTTPException(
            status.HTTP_500_INTERNAL_SERVER_ERROR,
            f"Snapshot restore failed: {exc}",
        ) from exc

    # Verify the collection actually registered. Qdrant returns
    # ``result: true`` once the recover RPC finishes, but the
    # ``/collections`` listing is cached briefly so we poll for up to
    # 30s before declaring success. Without this we returned
    # status="ok" while the catalogue stayed listed as "available" in
    # the UI, exactly the "downloads to 100% then nothing happens"
    # symptom.
    appeared = False
    poll_deadline = time.monotonic() + 30.0
    poll_delay = 0.5
    while time.monotonic() < poll_deadline:
        collections_after = await loop.run_in_executor(None, lambda: server_collections(qdrant_url=qdrant_url))
        if cat.collection in collections_after:
            appeared = True
            break
        await asyncio.sleep(poll_delay)
        poll_delay = min(poll_delay * 1.5, 3.0)

    duration = round(time.monotonic() - start, 1)

    logger.info(
        "v3 catalogue install: region=%s collection=%s appeared=%s duration=%.1fs",
        cat.region,
        cat.collection,
        appeared,
        duration,
    )

    if not appeared:
        raise HTTPException(
            status.HTTP_500_INTERNAL_SERVER_ERROR,
            f"Snapshot recovered but collection {cat.collection!r} did not "
            f"appear on Qdrant after 30s. Check Qdrant server logs and "
            f"disk space, then retry.",
        )

    return {
        "status": "ok",
        "region": cat.region,
        "collection": cat.collection,
        "snapshot_size_mb": cat.size_mb,
        "duration_seconds": duration,
        "collection_appeared": appeared,
    }


@router.get("/vector/search/")
async def semantic_search(
    q: str = Query(..., min_length=2, max_length=500, description="Natural language query"),
    region: str | None = Query(default=None, description="Filter by region"),
    limit: int = Query(default=25, ge=1, le=500),
) -> list[dict]:
    """Semantic search using vector similarity.

    Finds cost items whose descriptions are semantically similar
    to the query, even if the exact words don't match.
    E.g. "concrete wall" finds "reinforced partition C30/37".

    Degrades gracefully (NEW-B-105): when the optional ``[semantic]``
    extra is not installed (no embedding model / no ``qdrant_client``)
    the endpoint returns an empty result list with HTTP 200 instead of
    leaking an ``ImportError`` / ``RuntimeError`` as a 500. The lexical
    SQL search (``/costs/?q=``) remains the always-available path.
    """
    try:
        from app.core.vector import encode_texts, vector_search

        query_vector = encode_texts([q])[0]
        return vector_search(query_vector, region=region, limit=limit)
    except (ImportError, ModuleNotFoundError, RuntimeError) as exc:
        # Optional semantic stack absent / no embedding model loaded.
        # Never surface the raw import text to the client.
        logger.info("Semantic search unavailable, returning empty result: %s", exc)
        return []


# ── Categories (distinct classification.collection values) ───────────────


@router.get("/categories/", response_model=list[str])
async def list_categories(
    session: SessionDep,
    region: str | None = Query(default=None, description="Filter by region"),
) -> list[str]:
    """Return distinct classification.collection values. Cached for 30s."""
    cache_key = f"categories_{region or 'all'}"
    now = _time.monotonic()
    if _region_cache.get(cache_key) is not None and now - _region_cache["ts"] < _CACHE_TTL:
        return _region_cache[cache_key]

    from sqlalchemy import distinct, select

    from app.database import engine as _engine
    from app.modules.costs.models import CostItem

    _url = str(_engine.url)
    if "sqlite" in _url:
        collection_expr = json_path_text(CostItem.classification, "$.collection")
    else:
        collection_expr = CostItem.classification["collection"].as_string()

    stmt = (
        select(distinct(collection_expr))
        .where(CostItem.is_active.is_(True))
        .where(collection_expr.isnot(None))
        .where(collection_expr != "")
    )

    if region:
        stmt = stmt.where(CostItem.region == region)

    stmt = stmt.order_by(collection_expr)

    result = await session.execute(stmt)
    cats = [row[0] for row in result.all() if row[0]]
    _region_cache[cache_key] = cats
    _region_cache["ts"] = now
    return cats


# ── Category tree (4-level classification hierarchy) ─────────────────────


@router.get("/category-tree/", response_model=list[CategoryTreeNode])
async def get_category_tree(
    service: CostItemService = Depends(_get_service),
    region: str | None = Query(
        default=None,
        description="Restrict the aggregation to a single region (e.g. DE_BERLIN).",
    ),
    depth: int = Query(
        default=4,
        ge=1,
        le=4,
        description=(
            "How many classification levels to return (1..4). The BOQ "
            "'From Database' modal opens with depth=2 to keep the first "
            "paint snappy on cold catalogs; deeper levels are reachable "
            "via the search endpoint's classification_path filter."
        ),
    ),
    parent_path: str | None = Query(
        default=None,
        description=(
            "Optional slash-delimited prefix to scope the aggregation to "
            "a sub-branch (e.g. 'Concrete/Walls'). Lets clients lazily "
            "drill into a node without refetching the whole tree."
        ),
    ),
) -> list[CategoryTreeNode]:
    """Return the classification tree for a region.

    The tree is nested as
    ``collection → department → section → subsection``. NULL / empty
    values at any depth coalesce into the sentinel ``"__unspecified__"``;
    the frontend is expected to localize this label.

    Cached for 5 minutes per (region, depth, parent_path). The cache is
    wiped on any import / delete via ``_invalidate_cost_cache()``, so
    post-import catalogues become visible immediately on the next request.
    """
    cache_key = f"tree::{region or '__all__'}::d={depth}::p={parent_path or ''}"
    now = _time.monotonic()
    cached = _category_tree_cache.get(cache_key)
    if cached is not None and now - cached.get("ts", 0) < _CATEGORY_TREE_CACHE_TTL:
        return cached["nodes"]

    raw = await service.category_tree(region=region, depth=depth, parent_path=parent_path)
    nodes = [CategoryTreeNode.model_validate(n) for n in raw]
    _category_tree_cache[cache_key] = {"nodes": nodes, "ts": now}
    return nodes


# ── Available CWICR databases ─────────────────────────────────────────────


@router.get("/available-databases/")
async def list_available_databases() -> list[dict]:
    """List all available CWICR regional databases with their IDs.

    Use these IDs with POST /load-cwicr/{db_id} to import cost data.
    """
    return [
        {"id": db_id, "folder": folder.split("/")[0].replace("___DDC_CWICR", "")}
        for db_id, folder in _GITHUB_CWICR_FILES.items()
    ]


# ── Loaded CWICR databases (with vectorisation status) ────────────────────


@router.get("/loaded-databases/")
async def list_loaded_databases(
    session: SessionDep,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
) -> list[dict]:
    """List CWICR catalogues that are *actually loaded* into the SQL table.

    Returns one entry per region with both the SQL row count and the
    LanceDB vector count, so the Match-panel selector can surface three
    distinct UI states without an extra round-trip:

    * ``count == 0``                              - never happens (the
                                                    catalogue wouldn't be
                                                    listed) but guarded.
    * ``count > 0`` and ``vectorized_count == 0`` - "Catalogue loaded but
                                                    not vectorised yet"
    * ``count > 0`` and ``vectorized_count > 0``  - ready for matching

    Cheap: the SQL count comes from a single ``GROUP BY region`` against
    the indexed ``region`` column; the vector count comes from
    ``vector_count_for_region`` which counts the same ``cost_items`` store
    the ``/vector/index/`` action writes, filtered by the region field, so
    the badge tracks vectorisation progress instead of staying stuck (#170).
    """
    _ = user_id  # auth required via dependency; unused beyond that
    from sqlalchemy import func, select  # noqa: PLC0415

    from app.core.match_service.region_language import language_for  # noqa: PLC0415
    from app.core.vector import vector_count_for_region  # noqa: PLC0415
    from app.modules.costs.models import CostItem  # noqa: PLC0415

    result = await session.execute(
        select(CostItem.region, func.count(CostItem.id).label("cnt"))
        .where(CostItem.is_active.is_(True))
        .where(CostItem.region.isnot(None))
        .where(CostItem.region != "")
        .group_by(CostItem.region)
        .order_by(func.count(CostItem.id).desc())
    )
    out: list[dict] = []
    for region, sql_count in result.all():
        if not region or sql_count <= 0:
            continue
        vectorized = vector_count_for_region(region)
        out.append(
            {
                "id": region,
                "count": int(sql_count),
                "vectorized_count": int(vectorized),
                "ready": vectorized > 0,
                # Lets the /match-elements smart advisor filter by project
                # language without a second round-trip or a frontend mirror
                # of REGION_LANGUAGE.
                "language": language_for(region),
            }
        )
    return out


# ── User cost catalogs ─────────────────────────────────────────────────────
#
# A catalog is the user's own named "справочник работ и расценок": a
# first-class container with a REQUIRED currency that imported / manually
# created items belong to. Routes are registered BEFORE the ``/{item_id}``
# wildcard below so ``/catalogs/...`` never gets swallowed by it.


def _catalog_response(catalog: Any, item_count: int) -> CostCatalogResponse:
    """Build a CostCatalogResponse with the live item count attached."""
    response = CostCatalogResponse.model_validate(catalog)
    response.item_count = item_count
    return response


@router.post(
    "/catalogs/",
    response_model=CostCatalogResponse,
    status_code=201,
    dependencies=[Depends(RequirePermission("costs.create"))],
)
async def create_cost_catalog(
    data: CostCatalogCreate,
    user_id: CurrentUserId,
    service: CostCatalogService = Depends(_get_catalog_service),
) -> CostCatalogResponse:
    """Create a user-owned cost catalog (name + currency required)."""
    catalog = await service.create_catalog(data, created_by=_parse_user_uuid(user_id), source="manual")
    return _catalog_response(catalog, 0)


@router.get(
    "/catalogs/",
    response_model=list[CostCatalogResponse],
    dependencies=[Depends(RequirePermission("costs.list"))],
)
async def list_cost_catalogs(
    user: CurrentUserPayload,
    service: CostCatalogService = Depends(_get_catalog_service),
) -> list[CostCatalogResponse]:
    """List the caller's own cost catalogs with their active item counts.

    Scoped by ownership (``CostCatalog.created_by``): a non-admin caller
    sees only the catalogs they created; admins see all. This prevents
    cross-tenant disclosure of other users' catalog names / item counts.
    """
    owner_id = _parse_user_uuid((user or {}).get("sub"))
    is_admin = (user or {}).get("role") == "admin"
    catalogs = await service.list_catalogs(owner_id=owner_id, is_admin=is_admin)
    return [_catalog_response(catalog, count) for catalog, count in catalogs]


@router.patch(
    "/catalogs/{catalog_id}",
    response_model=CostCatalogResponse,
    dependencies=[Depends(RequirePermission("costs.update"))],
)
async def update_cost_catalog(
    catalog_id: uuid.UUID,
    data: CostCatalogUpdate,
    user: CurrentUserPayload,
    service: CostCatalogService = Depends(_get_catalog_service),
) -> CostCatalogResponse:
    """Update a catalog's name / description / currency.

    Scoped by ownership (``CostCatalog.created_by``): a non-owner caller
    gets a 404 (existence is not leaked), matching the module convention.
    Admins bypass the ownership check.

    A currency change is rejected (409) while the catalog has items: the
    stored rates are denominated in the old currency and would be silently
    corrupted by a re-label.
    """
    owner_id = _parse_user_uuid((user or {}).get("sub"))
    is_admin = (user or {}).get("role") == "admin"
    await service.get_owned_catalog(catalog_id, owner_id=owner_id, is_admin=is_admin)
    catalog = await service.update_catalog(catalog_id, data)
    return _catalog_response(catalog, await service.count_items(catalog_id))


@router.delete(
    "/catalogs/{catalog_id}",
    dependencies=[Depends(RequirePermission("costs.delete"))],
)
async def delete_cost_catalog(
    catalog_id: uuid.UUID,
    user: CurrentUserPayload,
    mode: str = Query(
        default="keep_items",
        pattern="^(keep_items|delete_items)$",
        description=(
            "keep_items detaches the items (they stay in the global cost table); "
            "delete_items soft-deletes them together with the catalog."
        ),
    ),
    service: CostCatalogService = Depends(_get_catalog_service),
) -> dict[str, Any]:
    """Delete a catalog, either keeping or soft-deleting its items.

    Scoped by ownership (``CostCatalog.created_by``): a non-owner caller
    gets a 404 (existence is not leaked), matching the module convention.
    Admins bypass the ownership check.
    """
    owner_id = _parse_user_uuid((user or {}).get("sub"))
    is_admin = (user or {}).get("role") == "admin"
    await service.get_owned_catalog(catalog_id, owner_id=owner_id, is_admin=is_admin)
    affected = await service.delete_catalog(catalog_id, mode=mode)
    _invalidate_cost_cache()
    return {"deleted": str(catalog_id), "mode": mode, "items_affected": affected}


@router.get(
    "/catalogs/{catalog_id}/export-excel/",
    dependencies=[Depends(RequirePermission("costs.list"))],
)
async def export_cost_catalog_excel(
    catalog_id: uuid.UUID,
    session: SessionDep,
    user: CurrentUserPayload,
    service: CostCatalogService = Depends(_get_catalog_service),
) -> StreamingResponse:
    """Export ONE catalog's items as an Excel file.

    Scoped by ownership (``CostCatalog.created_by``): a non-owner caller
    gets a 404 (existence is not leaked) so this never streams another
    user's rate list. Admins bypass the ownership check.

    Columns: code, description, unit, rate, currency, classification.
    Text cells go through ``_excel_safe`` so user-supplied values can not
    smuggle spreadsheet formulas into the export. Batched fetch keeps
    memory flat for large catalogs.
    """
    from openpyxl import Workbook

    owner_id = _parse_user_uuid((user or {}).get("sub"))
    is_admin = (user or {}).get("role") == "admin"
    catalog = await service.get_owned_catalog(catalog_id, owner_id=owner_id, is_admin=is_admin)

    wb = Workbook(write_only=True)
    ws = wb.create_sheet(title="Catalog")
    ws.append(["Code", "Description", "Unit", "Rate", "Currency", "Classification"])

    batch_size = 1000
    offset = 0
    base_stmt = (
        select(CostItem)
        .where(CostItem.catalog_id == catalog_id, CostItem.is_active.is_(True))
        .order_by(CostItem.code, CostItem.id)
    )
    while True:
        result = await session.execute(base_stmt.offset(offset).limit(batch_size))
        items = result.scalars().all()
        if not items:
            break
        for item in items:
            try:
                rate_val: object = float(item.rate)
            except (ValueError, TypeError):
                rate_val = 0
            classification = json.dumps(item.classification, ensure_ascii=False) if item.classification else ""
            ws.append(
                [
                    _excel_safe(item.code),
                    _excel_safe(item.description),
                    _excel_safe(item.unit),
                    rate_val,
                    _excel_safe(item.currency or catalog.currency),
                    _excel_safe(classification),
                ]
            )
        if len(items) < batch_size:
            break
        offset += batch_size

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # ASCII slug as the legacy fallback plus RFC 5987 filename* so non-Latin
    # catalog names ("Моя смета 2026") keep their real name in the download.
    slug = _re.sub(r"[^A-Za-z0-9]+", "-", catalog.name).strip("-").lower() or "catalog"
    utf8_name = urllib.parse.quote(f"{catalog.name}.xlsx", safe="")
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": (f"attachment; filename=\"{slug}.xlsx\"; filename*=UTF-8''{utf8_name}")},
    )


# ── Get by ID ─────────────────────────────────────────────────────────────


async def _enforce_item_catalog_ownership(
    item: CostItem,
    user: dict[str, Any] | None,
    catalog_service: CostCatalogService,
) -> None:
    """Reject access to a CostItem bound to a catalog the caller does not own.

    Most ``CostItem`` rows are the intentionally-shared global CWICR
    catalogue (``catalog_id IS NULL``) and stay globally readable/editable
    by anyone holding the relevant ``costs.*`` permission. Rows that DO
    carry a ``catalog_id``, however, belong to a user-owned
    :class:`CostCatalog` (``CostCatalog.created_by``); the catalog-level
    endpoints already lock those to their owner via
    :meth:`CostCatalogService.get_owned_catalog`. This mirrors that gate at
    the item level so the per-item GET/PATCH/DELETE handlers cannot be used
    to read, tamper with, or soft-delete another user's private catalog
    item by guessing its UUID.

    Ownership is masked as a 404 (not 403), matching ``get_owned_catalog``,
    so the response is indistinguishable from a missing catalog and cannot
    be used as a UUID-existence oracle. Admins bypass the check.
    """
    if item.catalog_id is None:
        return
    owner_id = _parse_user_uuid((user or {}).get("sub"))
    is_admin = (user or {}).get("role") == "admin"
    # Raises 404 when the caller is neither the catalog owner nor an admin.
    await catalog_service.get_owned_catalog(item.catalog_id, owner_id=owner_id, is_admin=is_admin)


@router.get("/{item_id}")
async def get_cost_item(
    item_id: uuid.UUID,
    user: OptionalUserPayload,
    service: CostItemService = Depends(_get_service),
    catalog_service: CostCatalogService = Depends(_get_catalog_service),
    locale: str | None = Query(
        default=None,
        max_length=10,
        description="Localize CWICR-frozen-German fields (see search endpoint).",
    ),
    accept_language: str | None = Header(default=None, alias="accept-language"),
) -> dict[str, Any]:
    """Get a cost item by ID, with optional locale-specific translation
    of CWICR's frozen-German vocabulary columns (variant_stats,
    classification.category, component units).

    Items in the shared global catalogue (``catalog_id IS NULL``) are
    readable by anyone, matching the public search/autocomplete contract.
    Items bound to a user-owned catalog are scoped to their owner (404 to
    everyone else) so a private catalog's rates can't be read by UUID.
    """
    item = await service.get_cost_item(item_id)
    await _enforce_item_catalog_ownership(item, user, catalog_service)
    response = CostItemResponse.model_validate(item)
    resolved_locale = _resolve_cost_locale(locale, accept_language)
    return _localize_response_payload(response, resolved_locale)


# ── Mass apply preview ──────────────────────────────────────────────────────


class MassApplyPreviewRequest(BaseModel):
    """Request body for ``POST /v1/costs/{id}/apply-preview``.

    ``quantity`` is the length (in the item's own ``unit``, e.g. metres) the
    section is applied to. Accepts a JSON number or numeric string; the
    service coerces it to a non-negative Decimal.
    """

    quantity: Decimal = Field(default=Decimal("1"), ge=0, description="Length quantity in the item's unit.")


@router.post(
    "/{item_id}/apply-preview/",
    response_model=MassApplyPreviewResponse,
    dependencies=[Depends(RequirePermission("costs.read"))],
)
async def preview_mass_apply(
    item_id: uuid.UUID,
    data: MassApplyPreviewRequest,
    _user_id: CurrentUserId,
    service: CostItemService = Depends(_get_service),
) -> MassApplyPreviewResponse:
    """Preview the effective rate + line total of applying an item to a length.

    Mass-aware: a section priced per tonne / per kg (``mass_basis`` set with
    a positive ``mass_per_unit``) is converted to its effective per-length
    rate (``mass_per_unit * rate / 1000`` for tonnes) and the derived mass is
    returned. A plain per-unit item simply echoes the catalog rate, so the
    same endpoint is safe to call for any item. Read-only; never persists.
    """
    payload = await service.mass_apply_preview(item_id, data.quantity)
    return MassApplyPreviewResponse.model_validate(payload)


# ── Update ────────────────────────────────────────────────────────────────


@router.patch(
    "/{item_id}",
    response_model=CostItemResponse,
    dependencies=[Depends(RequirePermission("costs.update"))],
)
async def update_cost_item(
    item_id: uuid.UUID,
    data: CostItemUpdate,
    user: CurrentUserPayload,
    service: CostItemService = Depends(_get_service),
    catalog_service: CostCatalogService = Depends(_get_catalog_service),
) -> CostItemResponse:
    """Update a cost item.

    Items bound to a user-owned catalog (``catalog_id`` set) are scoped to
    their owner: a non-owner non-admin caller gets a 404 (no write, no
    existence leak) even though they hold ``costs.update``. Global-catalogue
    items (``catalog_id IS NULL``) stay editable by anyone with the
    permission, matching the existing shared-catalogue contract.
    """
    existing = await service.get_cost_item(item_id)
    await _enforce_item_catalog_ownership(existing, user, catalog_service)
    item = await service.update_cost_item(item_id, data)
    return CostItemResponse.model_validate(item)


# ── Delete ────────────────────────────────────────────────────────────────


@router.delete(
    "/{item_id}",
    status_code=204,
    dependencies=[Depends(RequirePermission("costs.delete"))],
)
async def delete_cost_item(
    item_id: uuid.UUID,
    user: CurrentUserPayload,
    service: CostItemService = Depends(_get_service),
    catalog_service: CostCatalogService = Depends(_get_catalog_service),
) -> None:
    """Soft-delete a cost item.

    Items bound to a user-owned catalog (``catalog_id`` set) are scoped to
    their owner: a non-owner non-admin caller gets a 404 (no delete, no
    existence leak) even though they hold ``costs.delete``. Global-catalogue
    items (``catalog_id IS NULL``) stay deletable by anyone with the
    permission, matching the existing shared-catalogue contract.
    """
    existing = await service.get_cost_item(item_id)
    await _enforce_item_catalog_ownership(existing, user, catalog_service)
    await service.delete_cost_item(item_id)


# ── Bulk import ───────────────────────────────────────────────────────────


@router.post(
    "/bulk/",
    response_model=list[CostItemResponse],
    status_code=201,
    dependencies=[Depends(RequirePermission("costs.create"))],
)
async def bulk_import_cost_items(
    data: list[CostItemCreate],
    _user_id: CurrentUserId,
    service: CostItemService = Depends(_get_service),
) -> list[CostItemResponse]:
    """Bulk import cost items. Skips duplicates by code."""
    items = await service.bulk_import(data)
    if items:
        _invalidate_cost_cache()
    return [CostItemResponse.model_validate(i) for i in items]


# ── File import (CSV / Excel) ────────────────────────────────────────────

# Column name aliases for flexible matching (all lowercased)
# Header aliases for auto-detecting columns in a user-uploaded catalogue.
# Covers EN, DE and RU so a Russian estimator's Excel (Наименование / Цена /
# Ед.изм / Шифр) maps without a manual step. When auto-detection still misses a
# column (any other language, an unusual header), the import endpoint accepts an
# explicit ``column_map`` from the mapping UI - so import never silently drops a
# column the user needs.
_COST_COLUMN_ALIASES: dict[str, list[str]] = {
    "code": [
        "code",
        "item code",
        "cost code",
        "artikelnummer",
        "art.nr.",
        "item",
        "nr",
        "nr.",
        "no",
        "no.",
        "#",
        "id",
        "position",
        # RU
        "код",
        "шифр",
        "артикул",
        "№",
        "номер",
        "позиция",
        "обоснование",
        # FR / ES / IT / PT / PL / TR
        "réf",
        "référence",
        "código",
        "codice",
        "kod",
        "lp",
        "lp.",
        "poz",
        "partida",
    ],
    "description": [
        "description",
        "beschreibung",
        "desc",
        "text",
        "bezeichnung",
        "item description",
        "name",
        "title",
        # RU
        "наименование",
        "наименование работ",
        "описание",
        "работа",
        "работы",
        "наименование работ и затрат",
        # FR / ES / IT / PT / PL / TR
        "désignation",
        "libellé",
        "descripción",
        "designación",
        "concepto",
        "descrizione",
        "voce",
        "descrição",
        "opis",
        "nazwa",
        "açıklama",
        "tanım",
    ],
    "unit": [
        "unit",
        "einheit",
        "me",
        "uom",
        "unit of measure",
        "measure",
        # RU
        "ед",
        "ед.",
        "ед.изм",
        "ед. изм.",
        "ед.изм.",
        "единица",
        "единица измерения",
        "изм",
        # FR / ES / IT / PT / PL / TR
        "unité",
        "unidad",
        "ud",
        "unità",
        "um",
        "u.m.",
        "unidade",
        "un",
        "jedn",
        "j.m.",
        "birim",
    ],
    "rate": [
        "rate",
        "price",
        "cost",
        "unit rate",
        "unit price",
        "unit cost",
        "ep",
        "einheitspreis",
        "preis",
        "amount",
        "value",
        # RU
        "цена",
        "расценка",
        "стоимость",
        "цена за единицу",
        "цена за ед.",
        "цена за ед",
        "цена ед.",
        "цена/ед",
        "стоимость единицы",
        "стоимость ед.",
        "сметная стоимость",
        "тариф",
        # FR / ES / IT / PT / PL / TR
        "prix",
        "prix unitaire",
        "pu",
        "tarif",
        "coût",
        "precio",
        "precio unitario",
        "importe",
        "coste",
        "prezzo",
        "prezzo unitario",
        "costo",
        "importo",
        "preço",
        "custo",
        "cena",
        "stawka",
        "koszt",
        "fiyat",
        "birim fiyat",
        "tutar",
    ],
    "currency": [
        "currency",
        "währung",
        "curr",
        "cur",
        # RU
        "валюта",
        # FR / ES / IT / PT / PL / TR
        "devise",
        "monnaie",
        "moneda",
        "divisa",
        "valuta",
        "moeda",
        "waluta",
        "para birimi",
    ],
    "classification": [
        "classification",
        "din 276",
        "din276",
        "kg",
        "cost group",
        "nrm",
        "masterformat",
        "class",
        "category",
        "group",
        # RU
        "классификация",
        "класс",
        "раздел",
        "гэсн",
        "фер",
        "категория",
        "группа",
    ],
}

# Canonical column keys the importer understands, in display order.
_COST_CANONICAL_COLUMNS: tuple[str, ...] = (
    "code",
    "description",
    "unit",
    "rate",
    "currency",
    "classification",
)


def _match_cost_column(header: str) -> str | None:
    """Match a header string to a canonical column name using the alias map.

    Args:
        header: Raw column header text from the uploaded file.

    Returns:
        Canonical column key or None if unrecognised.
    """
    normalised = header.strip().lower()
    for canonical, aliases in _COST_COLUMN_ALIASES.items():
        if normalised in aliases:
            return canonical
    return None


def _resolve_cost_column_map(
    raw_headers: list[Any],
    overrides: dict[str, str] | None = None,
) -> dict[int, str]:
    """Build a {column_index -> canonical_key} map for a header row.

    ``overrides`` (from the mapping UI) maps a canonical key to the exact raw
    header text the user chose; it wins over alias auto-detection so any-language
    columns import correctly. Without an override a column falls back to the
    EN/DE/RU alias map.
    """
    # Normalise overrides to {raw_header_lower -> canonical} for lookup.
    override_by_header: dict[str, str] = {}
    if overrides:
        for canonical, header in overrides.items():
            if canonical in _COST_CANONICAL_COLUMNS and header:
                override_by_header[str(header).strip().lower()] = canonical

    column_map: dict[int, str] = {}
    for idx, hdr in enumerate(raw_headers):
        if hdr is None:
            continue
        key = str(hdr).strip().lower()
        if not key:
            continue
        canonical = override_by_header.get(key) or _match_cost_column(str(hdr))
        if canonical:
            # First column wins for a given canonical key (avoid a later
            # duplicate header silently overwriting the chosen one).
            if canonical not in column_map.values():
                column_map[idx] = canonical
    return column_map


def _safe_float(value: Any, default: float = 0.0) -> float:
    """Parse a value to float, returning *default* on failure.

    Handles strings with comma decimal separators (e.g. "1.234,56" -> 1234.56).
    """
    if value is None:
        return default
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return default
    # Scientific notation ("1E+05") parses directly but would be mangled by
    # the separator stripping below ("1E+05" -> "+05" -> 5), so try the raw
    # string first. Reject non-finite results ("inf", "nan") - those are not
    # rates and must fall through to the default.
    try:
        direct = float(text)
    except (ValueError, TypeError):
        pass
    else:
        if _math.isfinite(direct):
            return direct
    # Strip everything that is not part of the number itself: whitespace
    # thousands separators ("8 500,00" incl. non-breaking/thin spaces),
    # apostrophe thousands ("1'250.00", Swiss), ANY currency symbol or
    # code in any language (₽ € $ £ ₺ ¥ kr zł Kč RUB ...), and percent.
    # Keeping only digits, separators and sign is locale-agnostic, so no
    # currency list to maintain.
    text = _re.sub(r"[^0-9.,+-]", "", text)
    if not text:
        return default
    # Handle European-style numbers: "1.234,56" -> "1234.56"
    if "," in text and "." in text:
        last_comma = text.rfind(",")
        last_dot = text.rfind(".")
        if last_comma > last_dot:
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        # Comma-only: treat commas as thousands separators ONLY when there
        # are MULTIPLE 3-digit groups ("1,234,567"). A SINGLE comma is
        # always the decimal separator, even with exactly 3 digits after
        # it ("0,500", "12,345") - European 3-decimal rates are far more
        # common in cost files than a lone English thousands group, and
        # inflating a rate 1000x is the worse failure mode. Mixed
        # comma+dot inputs ("1,234.56") are handled by the branch above.
        if text.count(",") > 1 and _re.fullmatch(r"[+-]?\d{1,3}(,\d{3})+", text):
            text = text.replace(",", "")
        else:
            text = text.replace(",", ".")
    elif text.count(".") > 1:
        # Multiple dots can only be thousands groups ("12.345.678").
        text = text.replace(".", "")
    try:
        return float(text)
    except (ValueError, TypeError):
        return default


def _parse_cost_rows_from_csv(
    content_bytes: bytes,
    overrides: dict[str, str] | None = None,
) -> tuple[list[dict[str, Any]], set[str]]:
    """Parse rows from a CSV file for cost import.

    Tries UTF-8 first, then Latin-1 as fallback (common for DACH region files).

    Returns:
        ``(rows, mapped_keys)`` where ``mapped_keys`` is the set of canonical
        column names the header row resolved to (auto-detection + overrides),
        so the caller can enforce required-column coverage with a clear error
        instead of silently dropping unmapped columns.
    """
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            text = content_bytes.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise ValueError("Unable to decode CSV file -- unsupported encoding")

    sniffer = csv.Sniffer()
    try:
        dialect = sniffer.sniff(text[:4096], delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel  # type: ignore[assignment]

    reader = csv.reader(io.StringIO(text), dialect)
    raw_headers = next(reader, None)
    if not raw_headers:
        raise ValueError("CSV file is empty or has no header row")

    column_map = _resolve_cost_column_map(list(raw_headers), overrides)

    rows: list[dict[str, Any]] = []
    for raw_row in reader:
        row: dict[str, Any] = {}
        for idx, val in enumerate(raw_row):
            canonical = column_map.get(idx)
            if canonical:
                row[canonical] = val.strip() if isinstance(val, str) else val
        if row:
            rows.append(row)

    return rows, set(column_map.values())


def _csv_headers_and_sample(content_bytes: bytes, sample: int = 5) -> tuple[list[str], list[list[str]]]:
    """Return (raw_headers, first ``sample`` data rows) for the mapping preview."""
    text = None
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            text = content_bytes.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise ValueError("Unable to decode CSV file -- unsupported encoding")
    try:
        dialect = csv.Sniffer().sniff(text[:4096], delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel  # type: ignore[assignment]
    reader = csv.reader(io.StringIO(text), dialect)
    raw_headers = next(reader, None)
    if not raw_headers:
        raise ValueError("CSV file is empty or has no header row")
    rows: list[list[str]] = []
    for raw_row in reader:
        rows.append([str(c) for c in raw_row])
        if len(rows) >= sample:
            break
    return [str(h) for h in raw_headers], rows


def _parse_cost_rows_from_excel(
    content_bytes: bytes,
    overrides: dict[str, str] | None = None,
) -> tuple[list[dict[str, Any]], set[str]]:
    """Parse rows from an Excel (.xlsx) file for cost import.

    Returns ``(rows, mapped_keys)`` - see :func:`_parse_cost_rows_from_csv`.
    """
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content_bytes), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        raise ValueError("Excel file has no worksheets")

    rows_iter = ws.iter_rows(values_only=True)
    raw_headers = next(rows_iter, None)
    if not raw_headers:
        raise ValueError("Excel file is empty or has no header row")

    column_map = _resolve_cost_column_map(list(raw_headers), overrides)

    rows: list[dict[str, Any]] = []
    for raw_row in rows_iter:
        row: dict[str, Any] = {}
        for idx, val in enumerate(raw_row):
            canonical = column_map.get(idx)
            if canonical and val is not None:
                row[canonical] = val
        if row:
            rows.append(row)

    wb.close()
    return rows, set(column_map.values())


def _excel_headers_and_sample(content_bytes: bytes, sample: int = 5) -> tuple[list[str], list[list[str]]]:
    """Return (raw_headers, first ``sample`` data rows) for the mapping preview."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content_bytes), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        raise ValueError("Excel file has no worksheets")
    rows_iter = ws.iter_rows(values_only=True)
    raw_headers = next(rows_iter, None)
    if not raw_headers:
        wb.close()
        raise ValueError("Excel file is empty or has no header row")
    rows: list[list[str]] = []
    for raw_row in rows_iter:
        rows.append(["" if c is None else str(c) for c in raw_row])
        if len(rows) >= sample:
            break
    wb.close()
    return ["" if h is None else str(h) for h in raw_headers], rows


def _validate_cost_upload(content: bytes, filename: str) -> bool:
    """Run the size + magic-byte gate on an upload; return True if CSV.

    Shared by preview and import so both reject renamed binaries identically.
    Raises HTTPException on any violation.
    """
    if not content:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Uploaded file is empty.")
    if len(content) > _MAX_COST_UPLOAD_BYTES:
        raise HTTPException(
            status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
            detail=(
                f"Uploaded file is too large ({len(content) / (1024 * 1024):.1f} MB > "
                f"{_MAX_COST_UPLOAD_BYTES / (1024 * 1024):.0f} MB limit). Split the catalogue into smaller batches."
            ),
        )
    head = content[:SIGNATURE_BYTES_REQUIRED]
    signature = detect_signature(head)
    is_csv = filename.endswith(".csv")
    if is_csv:
        # A genuine CSV is plain text and carries no magic-byte signature
        # (``detect`` returns None). ANY recognised signature - including
        # "xml" - means the .csv is really a binary / structured container
        # with a renamed extension, so reject it. XML payloads belong on the
        # .xlsx path, not here.
        if signature is not None:
            raise HTTPException(
                status_code=status.HTTP_415_UNSUPPORTED_MEDIA_TYPE,
                detail=f"File extension is .csv but the content is a {signature} container. Re-upload as a real CSV.",
            )
        if b"\x00" in head:
            raise HTTPException(
                status_code=status.HTTP_415_UNSUPPORTED_MEDIA_TYPE,
                detail="CSV upload contains NUL bytes - likely a binary file with a renamed extension.",
            )
    else:
        try:
            require_signature(head, _ALLOWED_COST_IMPORT_SIGNATURES, filename=filename)
        except FileSignatureMismatch as exc:
            raise HTTPException(status_code=status.HTTP_415_UNSUPPORTED_MEDIA_TYPE, detail=str(exc)) from exc
        # Zip-bomb gate for OOXML containers: sum the DECLARED uncompressed
        # sizes before openpyxl inflates anything. Legacy OLE .xls is not a
        # zip - BadZipFile falls through and the signature gate above stays
        # the authority for those.
        try:
            with zipfile.ZipFile(io.BytesIO(content)) as zf:
                infos = zf.infolist()
                if len(infos) > _MAX_COST_ZIP_ENTRIES:
                    raise HTTPException(
                        status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                        detail=(
                            f"Excel archive contains {len(infos)} entries "
                            f"(> {_MAX_COST_ZIP_ENTRIES} limit) - not a valid workbook."
                        ),
                    )
                total_uncompressed = sum(info.file_size for info in infos)
                if total_uncompressed > _MAX_COST_ZIP_UNCOMPRESSED_BYTES:
                    raise HTTPException(
                        status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
                        detail=(
                            f"Excel archive decompresses to "
                            f"{total_uncompressed / (1024 * 1024):.0f} MB "
                            f"(> {_MAX_COST_ZIP_UNCOMPRESSED_BYTES / (1024 * 1024):.0f} MB limit). "
                            f"Split the catalogue into smaller batches."
                        ),
                    )
        except zipfile.BadZipFile:
            pass
    return is_csv


@router.post(
    "/import/preview/",
    dependencies=[Depends(RequirePermission("costs.create"))],
)
async def preview_cost_file(
    user: CurrentUserPayload,
    file: UploadFile = File(..., description="Excel (.xlsx) or CSV (.csv) file"),
    catalog_id: str | None = Form(
        default=None,
        description="Optional target catalog UUID; echoed back with its name/currency so the import UI can confirm the destination.",
    ),
    catalog_service: CostCatalogService = Depends(_get_catalog_service),
) -> dict[str, Any]:
    """Inspect an uploaded catalogue and return its columns + a suggested mapping.

    The mapping UI calls this first: it shows the file's real headers and a few
    sample rows next to a suggested column->field mapping (auto-detected in
    EN/DE/RU). The user confirms or corrects the mapping, then calls
    ``/import/file/`` with the chosen ``column_map`` so columns in any language
    import correctly instead of being silently dropped.

    ``has_currency_column`` tells the UI whether the file carries its own
    currency values - when it does not, importing into a NEW catalog requires
    an explicit ``catalog_currency``.
    """
    filename = (file.filename or "").lower()
    if not filename.endswith((".xlsx", ".csv", ".xls")):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Unsupported file type. Please upload an Excel (.xlsx) or CSV (.csv) file.",
        )
    content = await file.read()
    is_csv = _validate_cost_upload(content, filename)
    try:
        headers, sample = _csv_headers_and_sample(content) if is_csv else _excel_headers_and_sample(content)
    except ValueError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Failed to read file: {exc}") from exc
    except Exception as exc:  # noqa: BLE001
        logger.warning("Cost preview parse failed: %s", exc)
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Could not read the file headers. Check that the first row contains column names.",
        ) from exc

    # Suggested mapping: canonical field -> the raw header we auto-detected.
    detected = _resolve_cost_column_map(headers, None)
    suggested: dict[str, str] = {}
    for idx, canonical in detected.items():
        if canonical not in suggested and idx < len(headers):
            suggested[canonical] = headers[idx]

    # Echo the target catalog (when given) so the UI can confirm the
    # destination + default currency before the user hits import.
    catalog_info: dict[str, str] | None = None
    if catalog_id:
        try:
            catalog_uuid = uuid.UUID(catalog_id)
        except ValueError as exc:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail=f"catalog_id is not a valid UUID: {catalog_id!r}",
            ) from exc
        # Ownership gate: echoing a catalog's name/currency must require the
        # caller to own it (or be admin), matching the import endpoint. The
        # unscoped get_catalog used here leaked any catalog's name + currency to
        # any holder of costs.create who guessed its UUID.
        owner_uuid = _parse_user_uuid((user or {}).get("sub"))
        is_admin = (user or {}).get("role") == "admin"
        catalog = await catalog_service.get_owned_catalog(catalog_uuid, owner_id=owner_uuid, is_admin=is_admin)
        catalog_info = {"id": str(catalog.id), "name": catalog.name, "currency": catalog.currency}

    return {
        "headers": headers,
        "sample_rows": sample,
        "suggested_map": suggested,
        "target_fields": list(_COST_CANONICAL_COLUMNS),
        # Import rejects files where the mapping does not cover description
        # AND rate - surface that contract to the mapping UI up front.
        "required_fields": ["description", "rate"],
        "has_currency_column": "currency" in detected.values(),
        "catalog": catalog_info,
    }


@router.post(
    "/import/file/",
    dependencies=[Depends(RequirePermission("costs.create"))],
)
async def import_cost_file(
    user: CurrentUserPayload,
    file: UploadFile = File(..., description="Excel (.xlsx) or CSV (.csv) file"),
    column_map: str | None = Form(
        default=None,
        description='Optional JSON mapping of canonical field to raw header, e.g. {"description":"Наименование","rate":"Цена"}.',
    ),
    catalog_id: str | None = Form(
        default=None,
        description="Import into this EXISTING catalog (UUID). Mutually exclusive with catalog_name.",
    ),
    catalog_name: str | None = Form(
        default=None,
        description=(
            "Create a NEW catalog with this name and import into it. Also stamped "
            "as the items' region tag so they are findable as one database."
        ),
    ),
    catalog_currency: str | None = Form(
        default=None,
        description=(
            "ISO 4217 currency of the new catalog. REQUIRED with catalog_name when "
            "the file has no mapped currency column."
        ),
    ),
    service: CostItemService = Depends(_get_service),
    catalog_service: CostCatalogService = Depends(_get_catalog_service),
) -> dict[str, Any]:
    """Import cost items from an Excel or CSV file upload.

    Accepts a multipart file upload. The file must be .xlsx or .csv.

    Expected columns (flexible auto-detection):
    - **Code / Item Code / Nr.** -- unique cost item code (auto-generated when missing)
    - **Description / Beschreibung / Text** -- description (required)
    - **Unit / Einheit / ME** -- unit of measurement
    - **Rate / Price / Cost / EP** -- unit rate or price (required)
    - **Currency / Wahrung** -- currency code (defaults to the catalog currency)
    - **Classification / DIN 276 / KG** -- classification code

    The mapping (auto-detection + user ``column_map``) MUST cover
    ``description`` and ``rate`` - otherwise the import is rejected with
    422 instead of silently importing rows with empty text or zero rates.

    Catalog targeting: pass ``catalog_id`` to import into an existing
    catalog, or ``catalog_name`` (+ ``catalog_currency`` when the file has
    no currency column) to create one inline. Rows without a currency
    inherit the catalog currency; rows carrying a DIFFERENT currency are
    imported as-is and counted in ``mixed_currency_count``.

    Returns:
        Summary with counts of imported, skipped, and error details per row.
    """
    # Validate file extension (advisory - magic-byte gate below is the
    # real check; extensions are attacker-controlled).
    filename = (file.filename or "").lower()
    if not filename.endswith((".xlsx", ".csv", ".xls")):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Unsupported file type. Please upload an Excel (.xlsx) or CSV (.csv) file.",
        )

    # Read file content
    content = await file.read()
    if not content:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Uploaded file is empty.",
        )

    # Size cap + magic-byte gate (shared with the preview endpoint): a renamed
    # binary can't exhaust memory on the openpyxl path or slip past the .csv/.exe
    # vector.
    is_csv_request = _validate_cost_upload(content, filename)

    # An explicit column_map from the mapping UI overrides alias auto-detection,
    # so headers in any language (Russian, French, ...) import instead of being
    # dropped. Tolerate a malformed map by ignoring it rather than 500-ing.
    overrides: dict[str, str] | None = None
    if column_map:
        try:
            parsed_map = json.loads(column_map)
            if isinstance(parsed_map, dict):
                overrides = {str(k): str(v) for k, v in parsed_map.items() if v}
        except (ValueError, TypeError):
            logger.warning("Ignoring malformed column_map on cost import")

    # Parse rows based on file type
    try:
        if filename.endswith(".xlsx") or filename.endswith(".xls"):
            rows, mapped_keys = _parse_cost_rows_from_excel(content, overrides)
        else:
            rows, mapped_keys = _parse_cost_rows_from_csv(content, overrides)
    except ValueError as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Failed to parse file: {exc}",
        )
    except Exception as exc:
        logger.exception("Unexpected error parsing cost import file: %s", exc)
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Failed to parse file. Please check the format and try again.",
        )

    # Required-column gate: silently importing a file whose mapping missed the
    # description or rate column produces rows of empty text / zero prices, so
    # reject with an actionable message instead. Other columns stay optional.
    missing_required = [key for key in ("description", "rate") if key not in mapped_keys]
    if missing_required:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=(
                f"Required column(s) not mapped: {', '.join(missing_required)}. "
                f"Neither auto-detection nor your column mapping covered them. "
                f"Use the import preview to map a file column to each of "
                f"'description' and 'rate', then retry."
            ),
        )

    if not rows:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No data rows found in file. Check that the first row contains column headers.",
        )

    # ── Resolve the target catalog ─────────────────────────────────────────
    #
    # Either an existing catalog (catalog_id) or a new one created inline
    # (catalog_name [+ catalog_currency]). The catalog currency is the
    # default for rows that carry none of their own.
    if catalog_id and catalog_name and catalog_name.strip():
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="Pass either catalog_id (existing catalog) or catalog_name (new catalog), not both.",
        )

    owner_uuid = _parse_user_uuid((user or {}).get("sub"))
    is_admin = (user or {}).get("role") == "admin"

    catalog: Any = None
    if catalog_id:
        try:
            catalog_uuid = uuid.UUID(catalog_id)
        except ValueError as exc:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail=f"catalog_id is not a valid UUID: {catalog_id!r}",
            ) from exc
        # Ownership gate: importing into an EXISTING catalog requires the
        # caller to own it (or be an admin). A non-owner gets a 404 so the
        # catalog's existence is not leaked, matching the module convention.
        catalog = await catalog_service.get_owned_catalog(
            catalog_uuid,
            owner_id=owner_uuid,
            is_admin=is_admin,
        )
    elif catalog_name and catalog_name.strip():
        file_has_currency = "currency" in mapped_keys
        resolved_currency = (catalog_currency or "").strip().upper()
        if not resolved_currency and file_has_currency:
            # Derive the catalog currency from the file: the most common
            # non-empty currency value across the parsed rows.
            counts: dict[str, int] = {}
            for row in rows:
                value = str(row.get("currency", "")).strip().upper()
                if value:
                    counts[value] = counts.get(value, 0) + 1
            if counts:
                resolved_currency = max(counts, key=lambda k: counts[k])
        if not resolved_currency:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail=(
                    "catalog_currency is required when creating a new catalog from a "
                    "file that has no mapped currency column. Pass a 3-letter ISO 4217 "
                    "code (e.g. EUR, USD)."
                ),
            )
        try:
            create_payload = CostCatalogCreate(
                name=catalog_name.strip(),
                currency=resolved_currency,
                description=None,
            )
        except ValueError as exc:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail=f"Invalid catalog_currency {resolved_currency!r}: expected a 3-letter ISO 4217 code.",
            ) from exc
        catalog = await catalog_service.create_catalog(
            create_payload,
            created_by=owner_uuid,
            source="import",
        )

    # A user-supplied catalogue name groups the imported rows under one findable
    # region so the user sees their own catalogue (instead of items vanishing
    # into an unnamed bucket). Sanitised to a short region tag. When importing
    # into an existing catalog, its name plays the same role.
    catalog_region: str | None = None
    if catalog is not None:
        catalog_region = str(catalog.name).strip()[:50] or None
    elif catalog_name and catalog_name.strip():
        catalog_region = catalog_name.strip()[:50]

    # Convert rows to CostItemCreate objects and import via service
    items_to_import: list[CostItemCreate] = []
    skipped = 0
    errors: list[dict[str, Any]] = []
    auto_code = 1
    # Auto-generated codes are salted with a request-unique prefix: a bare
    # counter restarts at 1 per request, so a second code-less file into the
    # same catalog would dedupe against the first and silently import 0 rows.
    auto_code_salt = uuid.uuid4().hex[:6]
    mixed_currency_count = 0
    rate_parse_failures = 0

    for row_idx, row in enumerate(rows, start=2):
        try:
            code = str(row.get("code", "")).strip()
            description = str(row.get("description", "")).strip()

            # Skip rows without both code and description
            if not code and not description:
                skipped += 1
                continue

            # Auto-generate code if missing
            if not code:
                code = f"IMPORT-{auto_code_salt}-{auto_code:04d}"
            auto_code += 1

            # Skip obvious summary rows
            desc_lower = description.lower()
            if desc_lower in (
                "total",
                "grand total",
                "summe",
                "gesamt",
                "gesamtsumme",
                "subtotal",
                "zwischensumme",
            ):
                skipped += 1
                continue

            # Parse unit (default: pcs)
            unit = str(row.get("unit", "pcs")).strip()
            if not unit:
                unit = "pcs"

            # Parse rate. NaN sentinel distinguishes "value present but
            # unparseable" from a genuine 0 - _safe_float itself never
            # returns NaN (non-finite direct parses fall to the default).
            raw_rate = row.get("rate")
            rate = _safe_float(raw_rate, default=_math.nan)
            if _math.isnan(rate):
                if raw_rate is not None and str(raw_rate).strip():
                    rate_parse_failures += 1
                rate = 0.0

            # Parse currency - empty if absent, never country-default. Inside
            # a catalog, an empty row currency inherits the CATALOG currency;
            # a different non-empty currency is kept as-is but counted so the
            # caller can surface a mixed-currency warning.
            currency = str(row.get("currency", "")).strip().upper()
            if catalog is not None:
                if not currency:
                    currency = catalog.currency
                elif currency != catalog.currency:
                    mixed_currency_count += 1

            # Build classification
            classification: dict[str, str] = {}
            class_value = str(row.get("classification", "")).strip()
            if class_value:
                classification["code"] = class_value

            items_to_import.append(
                CostItemCreate(
                    code=code,
                    description=description,
                    unit=unit,
                    rate=rate,
                    currency=currency,
                    source="file_import",
                    classification=classification,
                    region=catalog_region,
                    catalog_id=catalog.id if catalog is not None else None,
                )
            )

        except Exception as exc:
            errors.append(
                {
                    "row": row_idx,
                    "error": str(exc),
                    "data": {k: str(v)[:100] for k, v in row.items()},
                }
            )
            logger.warning("Cost import error at row %d: %s", row_idx, exc)

    # Bulk import via service (handles duplicate detection)
    imported_items = await service.bulk_import(items_to_import) if items_to_import else []
    imported_count = len(imported_items)
    skipped_by_duplicate = len(items_to_import) - imported_count

    logger.info(
        "Cost file import complete: imported=%d, skipped=%d (empty) + %d (duplicate), errors=%d",
        imported_count,
        skipped,
        skipped_by_duplicate,
        len(errors),
    )

    if imported_count:
        _invalidate_cost_cache()

    return {
        "imported": imported_count,
        "skipped": skipped + skipped_by_duplicate,
        "errors": errors,
        "total_rows": len(rows),
        "catalog": catalog_region,
        "catalog_id": str(catalog.id) if catalog is not None else None,
        "catalog_currency": catalog.currency if catalog is not None else None,
        # Rows whose own currency differs from the catalog currency. Imported
        # as-is (never silently rewritten) - this is a warning, not a block.
        "mixed_currency_count": mixed_currency_count,
        # Rows whose rate cell held a value that could not be parsed to a
        # number (imported with rate 0). Surfaced as a warning, not a block.
        "rate_parse_failures": rate_parse_failures,
    }


# ── Load CWICR database from local DDC Toolkit ──────────────────────────────

# GitHub repository info for downloading CWICR parquet files
_GITHUB_CWICR_BASE_URL = "https://github.com/datadrivenconstruction/OpenConstructionEstimate-DDC-CWICR/raw/main"

# Mapping from db_id to the work-item parquet path under the CWICR repo, built
# from the single-source base registry (app.modules.costs.base_registry) so this
# map, the resource-catalog folder map in app.modules.catalog.router and the
# /base-catalog API can never drift. The registry carries the current nested
# layout: the 30 global-CWICR markets live under CIS-Russia-GESN-FER-TER/ and
# each national base under its own folder. CA_TORONTO is the canonical internal
# id for Canada (demo packs, the batimatech-ca partner pack, region_language); it
# shares the EN/Toronto parquet so load-cwicr resolves it instead of 404ing.
_GITHUB_CWICR_FILES: dict[str, str] = base_registry.github_workitems_files()
_GITHUB_CWICR_FILES.setdefault("CA_TORONTO", _GITHUB_CWICR_FILES["ENG_TORONTO"])

CWICR_SEARCH_PATHS = [
    "../../DDC_Toolkit/pricing/data/excel",
    "../DDC_Toolkit/pricing/data/excel",
    "../../WORLD_COST_BASES",
    "../WORLD_COST_BASES",
    "WORLD_COST_BASES",
    str(Path.home() / "DDC_Toolkit" / "pricing" / "data" / "excel"),
    str(Path.home() / "Desktop" / "CodeProjects" / "DDC_Toolkit" / "pricing" / "data" / "excel"),
    str(Path.home() / "Desktop" / "CodeProjects" / "ERP_26030500" / "WORLD_COST_BASES"),
]

_LOCAL_CWICR_FILE_ALIASES: dict[str, tuple[str, ...]] = {
    # The canonical local export is still named `TR_*`, while the product id
    # must distinguish it from the legacy metro snapshot `TR_ISTANBUL`.
    "TR_NATIONAL": ("TR",),
    # Bare-country national bases: the product id carries a `_NATIONAL` suffix,
    # but the exported parquet keeps the short country prefix. ES_ANDALUCIA and
    # IT_TOSCANA already match their own file names, so they need no alias.
    "BR_NATIONAL": ("BR",),
    "VN_NATIONAL": ("VN",),
    "ID_NATIONAL": ("ID",),
    "GR_NATIONAL": ("GR",),
}

# Local cache directory for downloaded parquet files. The cache is
# persistent: a successfully downloaded regional parquet is kept and reused
# on the next load, so a region survives GitHub being unreachable later
# (fresh offline installs were left with no rates at all when the runtime
# download failed).
_CWICR_CACHE_DIR = Path.home() / ".openestimator" / "cache"

# Optional bundled CWICR parquet files inside the installed package. The
# regional workitems parquet files are 25-60 MB each (30 regions, ~1 GB
# total), far beyond any sane wheel budget, so this directory ships empty.
# It exists as a lookup hook: partner packs, installers or operators can
# drop ``{db_id}*.parquet`` files here (or in a source checkout) to make
# regional rates fully offline. The resource *catalog* CSVs are likewise not
# bundled - the catalog module downloads them on demand and caches them.
_BUNDLED_CWICR_DIR = Path(__file__).resolve().parents[2] / "data" / "cwicr"


def _download_to_file(url: str, dest: Path, timeout: float = 120.0) -> None:
    """Stream a URL to a local file with proper SSL trust on every platform.

    Switched away from ``urllib.request.urlretrieve`` because Python's stdlib
    ``ssl`` module on Windows does NOT read the OS certificate store. Every
    HTTPS download to ``raw.githubusercontent.com`` fails with
    ``CERTIFICATE_VERIFY_FAILED: unable to get local issuer certificate``
    (issue #104, reporter skolodi v2.6.37 trying to load SP_BARCELONA).

    ``httpx`` is already a project dependency and ships with ``certifi``'s
    Mozilla CA bundle baked in, so verification just works on Windows out of
    the box. Streams in 1 MB chunks so the 1.1 GB Qdrant snapshot doesn't
    blow up memory.
    """
    import httpx

    with httpx.stream(
        "GET",
        url,
        follow_redirects=True,
        timeout=timeout,
        headers={"User-Agent": "openconstructionerp"},
    ) as resp:
        resp.raise_for_status()
        with open(dest, "wb") as f:
            for chunk in resp.iter_bytes(chunk_size=1024 * 1024):
                f.write(chunk)


_LAST_DOWNLOAD_ERROR: dict[str, str] = {}


def _download_cwicr_from_github_sync(db_id: str) -> Path | None:
    """Download a CWICR parquet file from GitHub if available (sync version).

    Downloads to ~/.openestimator/cache/{db_id}.parquet.
    Returns the local path on success, None on failure. The most recent
    failure reason for ``db_id`` lands in ``_LAST_DOWNLOAD_ERROR`` so the
    HTTPException emitted upstream can surface it instead of a generic
    "not found" message.
    """
    github_path = _GITHUB_CWICR_FILES.get(db_id)
    if not github_path:
        _LAST_DOWNLOAD_ERROR[db_id] = (
            f"backend has no GitHub mapping for '{db_id}'. Upgrade with "
            f"`pip install --upgrade openconstructionerp` (≥ v2.6.23 added "
            f"the 19 new regions; older backends know only the original 11)."
        )
        return None

    url = f"{_GITHUB_CWICR_BASE_URL}/{github_path}"
    cache_dir = _CWICR_CACHE_DIR
    cache_dir.mkdir(parents=True, exist_ok=True)
    local_path = cache_dir / f"{db_id}.parquet"

    # ``_find_cwicr_file`` already consulted the cache before calling this
    # helper, so anything still on disk here is a stale or partial leftover
    # (e.g. a stuck 0-byte file from an interrupted fetch). Wipe it before
    # the download so we never serve stale bytes. On success the fresh file
    # STAYS in the cache and is reused by later loads (offline resilience).
    if local_path.exists():
        local_path.unlink(missing_ok=True)

    logger.info("Downloading CWICR %s from GitHub: %s", db_id, url)
    try:
        _download_to_file(url, local_path)
        if local_path.exists() and local_path.stat().st_size > 1000:
            logger.info("Downloaded CWICR %s: %d bytes", db_id, local_path.stat().st_size)
            _LAST_DOWNLOAD_ERROR.pop(db_id, None)
            return local_path
        size = local_path.stat().st_size if local_path.exists() else 0
        logger.warning(
            "Downloaded file too small or missing: %s (%d bytes)",
            local_path,
            size,
        )
        _LAST_DOWNLOAD_ERROR[db_id] = (
            f"GitHub download for '{db_id}' returned {size} bytes (expected ≥ 1 KB). "
            f"URL: {url}. Likely upstream 404 or proxy strip - try re-checking "
            f"https://github.com/datadrivenconstruction/OpenConstructionEstimate-DDC-CWICR "
            f"is reachable from this network."
        )
        local_path.unlink(missing_ok=True)
        return None
    except Exception as exc:
        logger.warning("Failed to download CWICR %s from GitHub: %s", db_id, exc)
        _LAST_DOWNLOAD_ERROR[db_id] = f"GitHub download failed: {exc.__class__.__name__}: {exc}. URL: {url}"
        local_path.unlink(missing_ok=True)
        return None


async def _download_cwicr_from_github(db_id: str) -> Path | None:
    """Async wrapper: runs the sync download in a thread pool to avoid blocking."""
    import asyncio

    return await asyncio.to_thread(_download_cwicr_from_github_sync, db_id)


async def _find_cwicr_file(db_id: str) -> Path | None:
    """Find a CWICR database file by database ID (e.g., DE_BERLIN).

    Lookup order, offline-first so a region only needs the network once:
      1. Local DDC_Toolkit checkout (Parquet > Excel SIMPLE > Excel any).
      2. Persistent local cache (``~/.openestimator/cache/{db_id}.parquet``,
         written by a previous successful GitHub download).
      3. Parquet bundled/dropped into the installed package at
         ``app/data/cwicr`` (empty by default - see ``_BUNDLED_CWICR_DIR``).
      4. GitHub download as the last resort.
    """
    candidate_prefixes = (db_id, *_LOCAL_CWICR_FILE_ALIASES.get(db_id, ()))

    # Priority 1: Parquet files in local DDC_Toolkit / generated base folders
    # (fastest and most reliable)
    for search_path in CWICR_SEARCH_PATHS:
        roots = [Path(search_path), Path(search_path).parent / "parquet"]
        for parquet_path in roots:
            if not parquet_path.exists():
                continue
            matches = [
                f
                for f in parquet_path.iterdir()
                if f.suffix == ".parquet" and any(f.name.startswith(prefix) for prefix in candidate_prefixes)
            ]
            if matches:
                # A region prefix can match more than one parquet in the same
                # folder (e.g. the canonical IT_TOSCANA_workitems_costs_resources
                # base alongside an auxiliary IT_TOSCANA_province_price_variants).
                # Always prefer the canonical work-items base; sort otherwise so
                # the choice is deterministic rather than filesystem-order.
                canonical = [f for f in matches if "workitems_costs_resources" in f.name]
                return canonical[0] if canonical else sorted(matches)[0]

    # Priority 2: Excel SIMPLE
    for search_path in CWICR_SEARCH_PATHS:
        p = Path(search_path)
        if not p.exists():
            continue
        for f in p.iterdir():
            if (
                any(f.name.startswith(prefix) for prefix in candidate_prefixes)
                and "_SIMPLE" in f.name
                and f.suffix == ".xlsx"
            ):
                return f

    # Priority 3: Any Excel
    for search_path in CWICR_SEARCH_PATHS:
        p = Path(search_path)
        if not p.exists():
            continue
        for f in p.iterdir():
            if any(f.name.startswith(prefix) for prefix in candidate_prefixes) and f.suffix == ".xlsx":
                return f

    # Priority 4: Persistent local cache from a previous download. The 1 KB
    # floor skips a stuck 0-byte/partial file (the download helper re-fetches
    # and replaces it in that case).
    cached = _CWICR_CACHE_DIR / f"{db_id}.parquet"
    try:
        if cached.is_file() and cached.stat().st_size > 1000:
            logger.info("Using cached CWICR parquet for %s: %s", db_id, cached)
            return cached
    except OSError:
        logger.warning("Unreadable cached CWICR parquet at %s, ignoring", cached)

    # Priority 5: Parquet shipped inside (or dropped into) the package data dir.
    try:
        if _BUNDLED_CWICR_DIR.is_dir():
            for f in sorted(_BUNDLED_CWICR_DIR.iterdir()):
                if f.name.startswith(db_id) and f.suffix == ".parquet":
                    logger.info("Using bundled CWICR parquet for %s: %s", db_id, f)
                    return f
    except OSError:
        logger.warning("Unreadable bundled CWICR dir at %s, ignoring", _BUNDLED_CWICR_DIR)

    # Priority 6: Download from GitHub (last resort - runs in thread to not
    # block the event loop; the result is cached for the next run)
    downloaded = await _download_cwicr_from_github(db_id)
    if downloaded:
        return downloaded

    return None


@router.get(
    "/base-catalog",
    dependencies=[Depends(RequirePermission("costs.read"))],
)
@router.get(
    # The app runs with redirect_slashes=False, so register both path forms.
    "/base-catalog/",
    dependencies=[Depends(RequirePermission("costs.read"))],
    include_in_schema=False,
)
async def get_base_catalog(session: SessionDep) -> dict:
    """Return the full catalog of loadable CWICR cost bases.

    Enumerates the nine base families and every loadable market variant with its
    work-item ("position") count, currency, language and flag, so the import
    page, database setup and onboarding can all render one consistent, honest
    picker before anything is downloaded. Merges the live loaded counts from
    ``oe_costs_item`` so a base that is already imported shows its real figure
    and a "loaded" badge.
    """
    from sqlalchemy import func

    stmt = select(CostItem.region, func.count()).where(CostItem.is_active.is_(True)).group_by(CostItem.region)
    rows = (await session.execute(stmt)).all()
    loaded_counts = {region: int(count) for region, count in rows if region}
    return base_registry.public_catalog(loaded_counts)


@router.post(
    # No trailing slash - sibling endpoints (``/vector/load-github/{db_id}``,
    # ``/vector/restore-snapshot/{db_id}``) are also slash-less, and the
    # frontend calls this one without a slash too. Prior version had a stray
    # trailing slash that caused 404 Not Found on every region click.
    "/load-cwicr/{db_id}",
    # Any authenticated user can load a CWICR regional database. The data
    # is public reference content (no confidentiality), and gating it to
    # editor+ would block viewers from completing onboarding. Permission
    # ``costs.read`` (VIEWER level) is used instead of ``costs.create``.
    dependencies=[Depends(RequirePermission("costs.read"))],
)
async def load_cwicr_database(
    db_id: str,
    session: SessionDep,
    _user_id: CurrentUserId,
) -> dict:
    """Load a CWICR regional database from local DDC Toolkit files.

    Thin HTTP wrapper around :func:`load_cwicr_region`. The actual import work
    lives in that module-level helper so the partner-pack one-click installer
    can run the same load path without going through HTTP.
    """
    return await load_cwicr_region(db_id, session)


async def load_cwicr_region(db_id: str, session: AsyncSession) -> dict:
    """Load one CWICR regional cost database into the relational store.

    Optimized: reads Parquet, deduplicates by rate_code (55K unique items
    from 900K total rows), then bulk-inserts into SQLite.
    Typical time: 10-30 seconds.

    For databases not available locally (e.g. UK_GBP, USA_USD), automatically
    downloads from GitHub and caches at ~/.openestimator/cache/.

    Reusable building block shared by the ``POST /load-cwicr/{db_id}`` route and
    the partner-pack ``full-install`` orchestrator. Raises ``HTTPException`` on
    a missing file (404) or an import failure (500); callers that need fail-soft
    behaviour must catch it. Returns the same body dict the route returns.
    """
    import time

    import pandas as pd
    from sqlalchemy import func, select

    start = time.monotonic()

    # Quick check: if this region is already loaded, return immediately
    from app.modules.costs.models import CostItem

    existing_count_stmt = (
        select(func.count()).select_from(CostItem).where(CostItem.region == db_id, CostItem.is_active.is_(True))
    )
    existing_count = (await session.execute(existing_count_stmt)).scalar_one()
    if existing_count > 10:
        duration = round(time.monotonic() - start, 1)
        # Count the resource components already persisted for this region so the
        # partner-pack installer reports the embedded resource database on
        # re-activation, not 0. Each work item carries its labour/material/
        # equipment breakdown in the ``components`` JSON array; the true count is
        # the sum of those array lengths, mirroring the fresh-import branch's
        # ``resource_components = sum(len(v) for v in resources_by_code.values())``.
        #
        # Dialect-portable JSON array length: SQLite has ``json_array_length`` and
        # PostgreSQL (where ``components`` is physically JSONB) has
        # ``jsonb_array_length``. Detect the dialect from THIS session's bind, not
        # the global engine, so the PG test lane (global engine still SQLite) does
        # not emit ``json_array_length(jsonb)`` which PG rejects. Same coalesce as
        # the costs repository's "has components" predicate.
        dialect_name = session.bind.dialect.name if session.bind else "sqlite"
        if dialect_name == "sqlite":
            _comp_len = func.coalesce(func.json_array_length(CostItem.components), 0)
        else:
            _comp_len = func.coalesce(func.jsonb_array_length(CostItem.components), 0)
        resource_components_stmt = (
            select(func.coalesce(func.sum(_comp_len), 0))
            .select_from(CostItem)
            .where(CostItem.region == db_id, CostItem.is_active.is_(True))
        )
        resource_components = int((await session.execute(resource_components_stmt)).scalar_one() or 0)
        logger.info(
            "CWICR %s already loaded (%d items, %d resource components), skipping",
            db_id,
            existing_count,
            resource_components,
        )
        return {
            "imported": 0,
            "skipped": existing_count,
            "region": db_id,
            "total_items": existing_count,
            "resource_components": resource_components,
            "status": "already_loaded",
            "message": f"Database '{db_id}' is already loaded with {existing_count:,} items. "
            f"To reload, delete the region first.",
            "duration_seconds": duration,
        }

    # Find the file (async - GitHub download runs in thread pool)
    cwicr_path = await _find_cwicr_file(db_id)
    if not cwicr_path:
        # Surface the most-specific download failure so the user knows
        # whether it's a backend version, a network issue, or a 0-byte
        # cache stuck on disk - instead of a generic "not found".
        last_error = _LAST_DOWNLOAD_ERROR.get(db_id)
        detail = f"CWICR database '{db_id}' not found."
        if last_error:
            detail = f"{detail} {last_error}"
        else:
            detail = (
                f"{detail} No local copy was found (DDC_Toolkit checkout, "
                f"{_CWICR_CACHE_DIR / (db_id + '.parquet')}, or bundled package data) "
                f"and the GitHub download did not succeed. Check that this server can "
                f"reach https://github.com/datadrivenconstruction/"
                f"OpenConstructionEstimate-DDC-CWICR, or place the regional parquet "
                f"in the cache directory and retry."
            )
        raise HTTPException(status_code=404, detail=detail)

    logger.info("Loading CWICR from %s", cwicr_path)

    # Read file in thread pool to avoid blocking event loop
    import asyncio

    _path = cwicr_path

    def _read_file() -> pd.DataFrame:
        if _path.suffix == ".parquet":
            return pd.read_parquet(_path)
        return pd.read_excel(_path, engine="openpyxl")

    df = await asyncio.to_thread(_read_file)

    total_rows = len(df)
    logger.info("Raw data: %d rows", total_rows)

    from app.config import get_settings

    settings = get_settings()
    # The worker loads into PostgreSQL via a short-lived sync SQLAlchemy engine
    # built from ``database_sync_url`` - never a stray local SQLite file.
    # Prefer the live process env: embedded PG (v6 default) sets
    # DATABASE_URL/DATABASE_SYNC_URL there after the Settings cache is built, so
    # the cached pydantic values can be stale/empty (mirrors auto_migrate +
    # seed_demo_v2, which also read os.environ directly).
    import os as _os

    target = _os.environ.get("DATABASE_SYNC_URL") or settings.database_sync_url

    # Run in thread to avoid blocking the event loop during heavy pandas + DB work.
    try:
        result_data = await asyncio.to_thread(_process_and_insert_cwicr, str(cwicr_path), db_id, target)
    except Exception:
        logger.exception("CWICR import failed for %s", db_id)
        raise HTTPException(
            status_code=500,
            detail=f"Failed to import CWICR database '{db_id}'. Check server logs.",
        )
    # Note: a parquet downloaded into ~/.openestimator/cache is deliberately
    # KEPT (persistent cache). Reloading the region - or reinstalling into a
    # fresh database - then works without the network. The download helper
    # replaces stale/partial leftovers before each fresh fetch.

    duration = round(time.monotonic() - start, 1)
    result_data["duration_seconds"] = duration
    result_data["source_file"] = cwicr_path.name
    logger.info(
        "CWICR %s: %d imported, %d skipped in %.1fs",
        db_id,
        result_data.get("imported", 0),
        result_data.get("skipped", 0),
        duration,
    )

    # Seed the resource price sheet from the freshly imported work items. This is
    # what lets a coefficient base (Vietnam Dinh Muc, Indonesia AHSP - norm
    # quantities, no prices) be priced locally: every distinct resource gets an
    # editable row (0 for a coefficient base, the observed price for a priced
    # one). Idempotent and fail-soft: a seeding error must never fail the import
    # (the cost items are already committed), so it is logged and swallowed.
    if result_data.get("imported", 0) > 0:
        try:
            seed = await ResourcePriceService(session).seed_region(db_id)
            result_data["resource_prices"] = seed.as_dict()
        except Exception:
            logger.exception("Resource price seeding failed for %s (non-fatal)", db_id)

    _invalidate_cost_cache()
    # A new CWICR parquet may have been written alongside the SQL import
    # (or the import itself writes a parquet artefact). Clear the polars
    # LazyFrame + path-resolution lru_caches so the next /qdrant-search
    # call opens the new file rather than reading the old mmap snapshot.
    try:
        from app.modules.costs.parquet_lookup import clear_parquet_caches

        clear_parquet_caches()
    except Exception:
        logger.debug("parquet cache clear failed (non-fatal)", exc_info=True)

    # Schema-level failures (e.g. parquet missing the required ``rate_code``
    # column) must surface as 422 Unprocessable Entity - the file was
    # uploaded fine, the server understood it, but the payload doesn't
    # carry the columns this endpoint needs. The previous silent 200 made
    # the failure invisible to monitoring + client retry logic. The body
    # shape is preserved so existing UIs that read ``error`` still work.
    if result_data.get("error") == "no rate_code column":
        return JSONResponse(content=result_data, status_code=422)
    return result_data


def _pg_bulk_insert_cost_rows(sync_url: str, rows: list[tuple]) -> int:
    """Bulk-load CWICR cost rows into PostgreSQL, idempotent on (code, region).

    Uses PostgreSQL's native ``COPY`` into a staging table cloned from the target
    (no indexes), then a single ``INSERT ... SELECT ... ON CONFLICT (code, region)
    DO NOTHING`` into the live table. A 55K-row regional database loads in a few
    seconds, matching the SQLite ``INSERT OR IGNORE`` fast path.

    The previous implementation issued one 1000-row multi-VALUES ``INSERT`` per
    batch. Each statement paid a large SQLAlchemy Core compile (≈16K bind
    parameters) plus a server-side parse of the same, so a single regional import
    took roughly three minutes on embedded PostgreSQL versus ~5 seconds on
    SQLite. The synchronous HTTP request then timed out and the client saw no
    response -- the user-visible "crash" in issue #171. ``COPY`` skips per-row SQL
    parsing entirely and closes that gap.

    Duplicate ``(code, region)`` pairs are silently skipped via the
    ``uq_costs_code_region`` constraint, exactly like the SQLite path. Runs on a
    short-lived sync engine built from ``database_sync_url`` (falling back to the
    live ``DATABASE_SYNC_URL`` env var, which embedded PG wires in after the
    Settings cache is built) so the load always reaches the real cluster.

    Args:
        sync_url: Sync SQLAlchemy URL (e.g. ``postgresql+psycopg2://...``).
        rows: Positional tuples in the SQLite column order
            ``(id, code, description, unit, rate, currency, source,
            classification, tags, components, descriptions, is_active,
            region, metadata)``. The five JSON columns carry pre-serialized JSON
            text and stream straight into the ``json`` columns -- no decode /
            re-encode round-trip.

    Returns:
        Number of rows actually inserted (conflicts excluded).
    """
    import csv
    import io
    import os as _os

    from sqlalchemy import create_engine

    if not rows:
        return 0

    # Embedded PostgreSQL (the v6 default) wires DATABASE_SYNC_URL into the
    # process env *after* the pydantic Settings cache is built, so a caller may
    # hand us an empty/stale sync URL. Fall back to the live env var (the
    # authoritative source, same as auto_migrate_legacy_sqlite + seed_demo_v2)
    # so the bulk load reaches the real embedded cluster instead of raising
    # "Could not parse SQLAlchemy URL from given URL string".
    if not sync_url:
        sync_url = _os.environ.get("DATABASE_SYNC_URL", "")

    table = CostItem.__table__.name  # "oe_costs_item"
    col_list = (
        "id, code, description, unit, rate, currency, source, "
        "classification, tags, components, descriptions, is_active, region, metadata"
    )

    def _row_for_copy(row: tuple) -> list[object]:
        # ``is_active`` (index 11) is an int flag in the tuple; COPY needs a
        # boolean literal. Everything else is already text (the JSON columns
        # carry JSON text, streamed verbatim into the ``json`` columns).
        out = list(row)
        out[11] = "true" if row[11] else "false"
        return out

    engine = create_engine(sync_url)
    inserted = 0
    copy_chunk = 5000  # bound peak memory of the CSV buffer for large JSON rows
    raw = engine.raw_connection()
    try:
        cur = raw.cursor()
        # Staging table cloned from the target: same column types and defaults,
        # but no indexes or unique constraint, so COPY is maximally fast.
        # ``INCLUDING DEFAULTS`` lets COPY fill the omitted created_at/updated_at
        # (NOT NULL) columns with now() instead of failing on a null. The temp
        # table is dropped automatically when the transaction commits.
        cur.execute(f"CREATE TEMP TABLE _cwicr_stage (LIKE {table} INCLUDING DEFAULTS) ON COMMIT DROP")  # noqa: S608
        copy_sql = f"COPY _cwicr_stage ({col_list}) FROM STDIN WITH (FORMAT csv)"  # noqa: S608
        for i in range(0, len(rows), copy_chunk):
            buf = io.StringIO()
            writer = csv.writer(buf, quoting=csv.QUOTE_ALL)
            for r in rows[i : i + copy_chunk]:
                writer.writerow(_row_for_copy(r))
            buf.seek(0)
            cur.copy_expert(copy_sql, buf)
        # One idempotent upsert from staging into the indexed live table. The
        # target's server_default fills created_at/updated_at for the omitted
        # columns; duplicate (code, region) rows are skipped.
        cur.execute(  # noqa: S608
            f"INSERT INTO {table} ({col_list}) "
            f"SELECT {col_list} FROM _cwicr_stage "
            f"ON CONFLICT (code, region) DO NOTHING"
        )
        inserted = cur.rowcount if cur.rowcount and cur.rowcount > 0 else 0
        raw.commit()
    except Exception:
        try:
            raw.rollback()
        except Exception:  # noqa: BLE001
            pass
        raise
    finally:
        raw.close()
        engine.dispose()
    return inserted


def _process_and_insert_cwicr(parquet_path: str, db_id: str, db_file: str) -> dict[str, Any]:
    """Process CWICR parquet + insert into PostgreSQL. Runs in a thread.

    Uses vectorized pandas (no iterrows!) and delegates the load to
    ``_pg_bulk_insert_cost_rows`` (PostgreSQL ``COPY`` into a staging table +
    ``INSERT ... ON CONFLICT (code, region) DO NOTHING``). ``db_file`` carries
    the sync SQLAlchemy URL (``postgresql+psycopg2://...``) of the target
    cluster.
    """
    import json as _json
    import logging
    import math
    import time

    import pandas as pd

    _log = logging.getLogger("cwicr_import")
    start = time.monotonic()

    # 1. Read parquet
    df = pd.read_parquet(parquet_path)
    total_rows = len(df)
    df.columns = [str(c).strip().lower() for c in df.columns]

    if "rate_code" not in df.columns:
        return {"imported": 0, "skipped": 0, "total_rows": total_rows, "error": "no rate_code column"}

    # 2. Vectorized processing - use groupby.first() instead of iterrows
    if "rate_original_name" in df.columns and "rate_final_name" in df.columns:
        df["_desc"] = (
            df["rate_original_name"].fillna("").astype(str) + " " + df["rate_final_name"].fillna("").astype(str)
        ).str.strip()
    elif "rate_original_name" in df.columns:
        df["_desc"] = df["rate_original_name"].fillna("").astype(str)
    else:
        df["_desc"] = ""

    # Aggregate: take first row's values per rate_code (vectorized, no iteration)
    agg_cols = {}
    for col in [
        "_desc",
        "rate_unit",
        "total_cost_per_position",
        "collection_name",
        "department_name",
        "section_name",
        "subsection_name",
        "category_type",
        "cost_of_working_hours",
        "total_value_machinery_equipment",
        "total_material_cost_per_position",
        "total_labor_hours_all_personnel",
        "count_total_people_per_unit",
    ]:
        if col in df.columns:
            agg_cols[col] = "first"

    # Abstract-resource rows carry per-variant price options; preserve them so the UI can offer a picker.
    # Column names follow the actual CWICR parquet schema (variable_parts / est_price_all_values),
    # not the legacy aliases. position_count is a single per-rate_code total, not per-variant.
    _ABSTRACT_COLS = (
        "row_type",
        "price_abstract_resource_variable_parts",
        "price_abstract_resource_est_price_all_values",
        "price_abstract_resource_position_count",
        "price_abstract_resource_est_price_min",
        "price_abstract_resource_est_price_max",
        "price_abstract_resource_est_price_mean",
        "price_abstract_resource_est_price_median",
        "price_abstract_resource_unit",
        "price_abstract_resource_group_per_unit",
        "price_abstract_resource_variable_parts_per_unit",
        "price_abstract_resource_est_price_all_values_per_unit",
    )
    for col in _ABSTRACT_COLS:
        if col in df.columns:
            agg_cols[col] = "first"

    grouped = df.groupby("rate_code", sort=False).agg(agg_cols)
    _log.info("Grouped %d unique items from %d rows in %.1fs", len(grouped), total_rows, time.monotonic() - start)

    # 3. Build insert tuples (vectorized - no Python loop over rows)
    def _safe_float(v: object) -> float:
        if v is None:
            return 0.0
        try:
            f = float(v)  # type: ignore[arg-type]
            return 0.0 if math.isnan(f) else f
        except (ValueError, TypeError):
            return 0.0

    def _safe_str(v: object) -> str:
        if v is None or (isinstance(v, float) and pd.isna(v)):
            return ""
        return str(v).strip()

    def _split_bul(value: object) -> list[str]:
        if value is None or (isinstance(value, float) and pd.isna(value)):
            return []
        return [p.strip() for p in str(value).split("\u2022") if p.strip()]

    # 4. Pre-build resource components per rate_code using vectorized pandas
    # Filter out empty rows, then group resources by rate_code
    _LABOR_UNITS = {"hrs", "h", "person-hour", "person-hours", "man-hours"}

    # Handle column name variants across different CWICR regional databases:
    # Most databases use: resource_cost, resource_price_per_unit_current
    # ENG_TORONTO uses:   resource_cost_eur, resource_price_per_unit_eur_current
    _cost_col = "resource_cost" if "resource_cost" in df.columns else "resource_cost_eur"
    _price_col = (
        "resource_price_per_unit_current"
        if "resource_price_per_unit_current" in df.columns
        else "resource_price_per_unit_eur_current"
    )

    res_cols = [
        "rate_code",
        "resource_name",
        "resource_code",
        "resource_unit",
        "resource_quantity",
        _price_col,
        _cost_col,
        "row_type",
        "is_machine",
        "is_material",
        "is_labor",
    ]
    available_res_cols = [c for c in res_cols if c in df.columns]

    # ── Per-component variants index (from Abstract resource rows) ──
    # Each rate_code can have several "Абстрактный ресурс" / "Abstract
    # resource" rows - one per variable component (e.g. formwork type +
    # board type + crane type). Each row carries its own
    # ``price_abstract_resource_variable_parts`` list. We index by
    # (rate_code, resource_code) so we can stamp the variant catalog onto
    # the matching component below - replacing the previous "first row
    # wins, dump on the cost item" behaviour that lost 2 of 3 variant
    # slots on KANE_RINE_KAKARI_KARI and similar rates.
    def _strip_unit_prefix(tok: str) -> str:
        # First per-unit token can be prefixed with a unit marker
        # (e.g. ``м3=20688.85``) - strip it so we can parse the number.
        if "=" in tok:
            return tok.split("=", 1)[1].strip()
        return tok

    abstract_variants_by_pair: dict[tuple[str, str], dict[str, Any]] = {}
    if "price_abstract_resource_variable_parts" in df.columns and "resource_code" in df.columns:
        abs_mask = df["price_abstract_resource_variable_parts"].fillna("").astype(str).str.len() > 0
        # Narrow to just the columns this loop reads before iterating. Without
        # it, ``df[abs_mask].iterrows()`` materializes a Series across all ~85
        # parquet columns per row; selecting the dozen columns actually used
        # mirrors the scope-of-work index below and trims the per-row overhead.
        # Same row set, same values - only the unused columns are dropped.
        _abs_cols = [
            c
            for c in (
                "rate_code",
                "resource_code",
                "price_abstract_resource_variable_parts",
                "price_abstract_resource_est_price_all_values",
                "price_abstract_resource_est_price_all_values_per_unit",
                "price_abstract_resource_common_start",
                "price_abstract_resource_est_price_min",
                "price_abstract_resource_est_price_max",
                "price_abstract_resource_est_price_mean",
                "price_abstract_resource_est_price_median",
                "price_abstract_resource_unit",
                "price_abstract_resource_group_per_unit",
            )
            if c in df.columns
        ]
        for _, r in df.loc[abs_mask, _abs_cols].iterrows():
            rc = _safe_str(r.get("rate_code", ""))
            rescode = _safe_str(r.get("resource_code", ""))
            if not rc or not rescode:
                continue
            labels = _split_bul(r.get("price_abstract_resource_variable_parts"))
            values = _split_bul(r.get("price_abstract_resource_est_price_all_values"))
            pu_vals_raw = _split_bul(r.get("price_abstract_resource_est_price_all_values_per_unit"))
            pu_vals = [_strip_unit_prefix(t) for t in pu_vals_raw]
            if not labels or len(labels) < 2:
                continue
            # Many rate rows (e.g. KANE_RINE_KAKARI_KARI's three variant
            # slots) have an empty ``..._all_values`` column and only the
            # ``_per_unit`` series populated. Fall back to per-unit so we
            # still build the variant catalog instead of dropping it.
            if len(values) != len(labels) and len(pu_vals) == len(labels):
                values = pu_vals
            if len(values) != len(labels):
                continue
            # ``common_start`` is the shared abstract-resource base name
            # (e.g. "Beton, Sortenliste C") that prefixes every variant's
            # variable_part. Read it BEFORE building variants so each row's
            # ``full_label`` = ``common_start + label`` - what the BOQ
            # resource row + Resource Summary display after a pick. The
            # picker still renders ``label`` (variable part only) in its
            # accordion rows because ``stats.common_start`` shows the base
            # once as a header; the BOQ side has no header and needs the
            # full composed name on each entry.
            common_start = _safe_str(r.get("price_abstract_resource_common_start"))[:240]
            variants_l: list[dict] = []
            for i, (lbl, val) in enumerate(zip(labels, values, strict=False)):
                v = _safe_float(val)
                if v <= 0:
                    continue
                variable_part = lbl[:200]
                full_label = (f"{common_start} {variable_part}".strip() if common_start else variable_part)[:400]
                variants_l.append(
                    {
                        "index": i,
                        "label": variable_part,
                        "full_label": full_label,
                        "price": round(v, 2),
                        "price_per_unit": round(_safe_float(pu_vals[i]), 4) if i < len(pu_vals) else None,
                    }
                )
            if not variants_l:
                continue
            abstract_variants_by_pair[(rc, rescode)] = {
                "variants": variants_l,
                "variant_stats": {
                    "min": round(_safe_float(r.get("price_abstract_resource_est_price_min")), 2),
                    "max": round(_safe_float(r.get("price_abstract_resource_est_price_max")), 2),
                    "mean": round(_safe_float(r.get("price_abstract_resource_est_price_mean")), 2),
                    "median": round(_safe_float(r.get("price_abstract_resource_est_price_median")), 2),
                    "unit": _safe_str(r.get("price_abstract_resource_unit"))[:20],
                    "group": _safe_str(r.get("price_abstract_resource_group_per_unit"))[:120],
                    "count": len(variants_l),
                    "common_start": common_start,
                },
            }
    _log.info("Indexed %d per-component variant catalogs", len(abstract_variants_by_pair))

    # ── Scope-of-work index ──
    # ``work_composition_text`` carries the ordered steps describing HOW a
    # position is performed (e.g. "Установка телескопических стоек." /
    # "Bodenbearbeitung nach Maß." / "Préparation du sol."). The companion
    # ``is_scope`` flag is set in EN/RU exports but stays False in DE/FR
    # exports, so we don't gate on it - instead we treat any row with a
    # non-empty ``work_composition_text`` AND an empty ``resource_name`` as
    # a scope step. Verified universal across all 16 cached regional
    # parquets (168 120 scope rows in each, 0 overlaps with resource rows).
    scope_by_code: dict[str, list[str]] = {}
    if "work_composition_text" in df.columns:
        wct_str = df["work_composition_text"].fillna("").astype(str)
        rname_str = df["resource_name"].fillna("").astype(str) if "resource_name" in df.columns else None
        scope_mask = wct_str.str.len() > 0
        if rname_str is not None:
            scope_mask = scope_mask & (rname_str.str.len() == 0)
        scope_sub = df[scope_mask][["rate_code", "work_composition_text"]]
        for _, r in scope_sub.iterrows():
            rc = _safe_str(r.get("rate_code", ""))
            text = _safe_str(r.get("work_composition_text", ""))
            if not rc or not text or text == "nan":
                continue
            scope_by_code.setdefault(rc, []).append(text[:500])
    _log.info("Indexed scope_of_work for %d rate_codes", len(scope_by_code))

    resources_by_code: dict[str, list[dict]] = {}
    if "resource_name" in df.columns and _cost_col in df.columns:
        # Filter rows that have resource data (non-empty name, non-zero cost)
        res_df = df[
            available_res_cols
            + (
                ["price_abstract_resource_variable_parts"]
                if "price_abstract_resource_variable_parts" in df.columns
                else []
            )
        ].copy()
        res_df = res_df[res_df["resource_name"].fillna("").str.len() > 0]
        if _cost_col in res_df.columns:
            # A row is a genuine resource component when it carries EITHER a
            # cost OR a norm quantity. Coefficient bases (Vietnam Dinh Muc,
            # Indonesia AHSP) publish the full labour / material / machine
            # breakdown with norm quantities but no prices (priced regionally),
            # so gating on cost alone dropped every one of their resources.
            # Keeping rows that carry a quantity fixes that; priced bases are
            # unaffected (their resource rows already carry a cost) beyond also
            # retaining a few legitimately unpriced-but-quantified lines.
            _keep = res_df[_cost_col].fillna(0).astype(float).abs() > 0.001
            if "resource_quantity" in res_df.columns:
                _keep = _keep | (pd.to_numeric(res_df["resource_quantity"], errors="coerce").fillna(0.0).abs() > 1e-9)
            # Abstract-resource rows are variant slots the user picks from - keep
            # them even when both cost and quantity are zero. Without this
            # carve-out KAME-LI-MENE-KAPU and similar amortisation / option rows
            # get silently dropped and the user loses one of their variant picks.
            if "price_abstract_resource_variable_parts" in res_df.columns:
                _keep = _keep | (res_df["price_abstract_resource_variable_parts"].fillna("").astype(str).str.len() > 0)
            res_df = res_df[_keep]
        if "row_type" in res_df.columns:
            res_df = res_df[res_df["row_type"].fillna("") != "Scope of work"]

        # FULLY VECTORIZED: build component dicts via column operations, then
        # group by rate_code once using a dict accumulator. This replaces the
        # previous iterrows() loop which was O(N) Python interpreter overhead
        # (~5min for 900K rows → now ~5s).
        if len(res_df) > 0:
            # Normalize types
            res_df["_rc"] = res_df["rate_code"].astype(str)
            res_df["_name"] = res_df["resource_name"].fillna("").astype(str).str.slice(0, 200)
            res_df["_code"] = (
                res_df["resource_code"].fillna("").astype(str).str.slice(0, 50)
                if "resource_code" in res_df.columns
                else ""
            )
            res_df["_unit"] = (
                res_df["resource_unit"].fillna("").astype(str).str.slice(0, 20)
                if "resource_unit" in res_df.columns
                else ""
            )
            res_df["_qty"] = (
                pd.to_numeric(res_df["resource_quantity"], errors="coerce").fillna(0.0).round(4)
                if "resource_quantity" in res_df.columns
                else 0.0
            )
            res_df["_rate"] = (
                pd.to_numeric(res_df[_price_col], errors="coerce").fillna(0.0).round(2)
                if _price_col in res_df.columns
                else 0.0
            )
            res_df["_cost_v"] = pd.to_numeric(res_df[_cost_col], errors="coerce").fillna(0.0).round(2)

            # Compute ctype vectorized
            _row_type = res_df.get("row_type", pd.Series([""] * len(res_df), index=res_df.index)).fillna("").astype(str)
            _is_mach = (
                res_df.get("is_machine", pd.Series([False] * len(res_df), index=res_df.index))
                .fillna(False)
                .astype(bool)
            )
            _is_mat = (
                res_df.get("is_material", pd.Series([False] * len(res_df), index=res_df.index))
                .fillna(False)
                .astype(bool)
            )
            _unit_lc = res_df["_unit"].str.lower()
            _is_labor_unit = _unit_lc.isin(_LABOR_UNITS)

            # Default
            ctype_arr = pd.Series(["other"] * len(res_df), index=res_df.index, dtype=object)
            # Material via row_type == Abstract resource
            ctype_arr = ctype_arr.mask(_row_type == "Abstract resource", "material")
            # is_material branch
            ctype_arr = ctype_arr.mask(_is_mat & ~_is_labor_unit, "material")
            ctype_arr = ctype_arr.mask(_is_mat & _is_labor_unit, "labor")
            # is_machine branch (overrides is_material)
            ctype_arr = ctype_arr.mask(_is_mach, "equipment")
            ctype_arr = ctype_arr.mask(_is_mach & (_row_type == "Machinist"), "operator")
            ctype_arr = ctype_arr.mask(_is_mach & (_row_type == "Electricity"), "electricity")
            # Explicit labour flag. Most bases tag the labour line with is_labor
            # rather than putting a labor unit on a material row, so fill any row
            # still left "other" that carries is_labor. Guarded on == "other" so
            # it never overrides a material / equipment / operator classification
            # already set above. Without this, coefficient bases (Vietnam Dinh
            # Muc, Indonesia AHSP) and even the priced bases' plain labour lines
            # fall through unclassified instead of typing as labor.
            _is_labor = (
                res_df.get("is_labor", pd.Series([False] * len(res_df), index=res_df.index)).fillna(False).astype(bool)
            )
            ctype_arr = ctype_arr.mask(_is_labor & (ctype_arr == "other"), "labor")
            res_df["_type"] = ctype_arr

            # Build records via zip over numpy arrays - much faster than iterrows
            rc_arr = res_df["_rc"].to_numpy()
            name_arr = res_df["_name"].to_numpy()
            code_arr = res_df["_code"].to_numpy()
            unit_arr = res_df["_unit"].to_numpy()
            qty_arr = res_df["_qty"].to_numpy()
            rate_arr = res_df["_rate"].to_numpy()
            cost_arr = res_df["_cost_v"].to_numpy()
            type_arr = res_df["_type"].to_numpy()

            # strict=True surfaces array length drift immediately instead of
            # silently truncating component rows mid-import - important for a
            # cost-data pipeline where a missing column would otherwise corrupt
            # the assembly composition without leaving any audit trail.
            for rc, nm, cd, un, qt, rt, cs, tp in zip(
                rc_arr,
                name_arr,
                code_arr,
                unit_arr,
                qty_arr,
                rate_arr,
                cost_arr,
                type_arr,
                strict=True,
            ):
                comps = resources_by_code.get(rc)
                if comps is None:
                    comps = []
                    resources_by_code[rc] = comps
                comp: dict[str, Any] = {
                    "name": nm,
                    "code": cd,
                    "unit": un,
                    "quantity": float(qt),
                    "unit_rate": float(rt),
                    "cost": float(cs),
                    "type": tp,
                }
                # Stamp per-component variant catalog if this resource is one
                # of the abstract-resource (variant) slots for this rate.
                v_data = abstract_variants_by_pair.get((rc, cd))
                if v_data is not None:
                    comp["available_variants"] = v_data["variants"]
                    comp["available_variant_stats"] = v_data["variant_stats"]
                comps.append(comp)

        _log.info("Built resources for %d rate_codes in %.1fs", len(resources_by_code), time.monotonic() - start)

    # 5. Build the insert rows. ``db_file`` carries the sync SQLAlchemy URL
    # (postgresql://...) of the target cluster - see the caller. Every row is
    # accumulated and handed to ``_pg_bulk_insert_cost_rows`` (COPY into a
    # staging table + ON CONFLICT DO NOTHING) below.

    # CWICR parquet carries no currency column - every rate is denominated in
    # the region's local currency. Resolve it ONCE from ``db_id`` (constant for
    # the whole import) so each row persists its true ISO currency instead of
    # the empty string that read-side fallbacks then had to paper over.
    resolved_currency = _resolve_currency(None, db_id)

    skipped_count = 0
    batch: list[tuple] = []

    for rate_code, row in grouped.iterrows():
        desc = _safe_str(row.get("_desc", ""))
        if len(desc) < 3:
            desc = _safe_str(row.get("subsection_name", ""))
        if len(desc) < 3:
            skipped_count += 1
            continue

        code = _safe_str(rate_code)[:100]
        if not code:
            skipped_count += 1
            continue

        unit = _safe_str(row.get("rate_unit", "m2"))[:20] or "m2"
        rate = round(_safe_float(row.get("total_cost_per_position", 0)), 2)

        classification: dict[str, str] = {}
        for key in ("collection_name", "department_name", "section_name", "subsection_name"):
            val = _safe_str(row.get(key, ""))
            if val:
                classification[key.replace("_name", "")] = val
        cat = _safe_str(row.get("category_type", ""))
        if cat:
            classification["category"] = cat

        metadata: dict[str, Any] = {}
        for mkey, col in [
            ("labor_cost", "cost_of_working_hours"),
            ("equipment_cost", "total_value_machinery_equipment"),
            ("material_cost", "total_material_cost_per_position"),
            ("labor_hours", "total_labor_hours_all_personnel"),
            ("workers_per_unit", "count_total_people_per_unit"),
        ]:
            v = _safe_float(row.get(col, 0))
            if v > 0:
                metadata[mkey] = round(v, 2)

        # ── Scope of work - ordered steps describing HOW the position is
        # performed (e.g. "Установка телескопических стоек."). Sourced from
        # rows flagged ``is_scope=True`` and pre-indexed above.
        steps = scope_by_code.get(code)
        if steps:
            metadata["scope_of_work"] = steps

        labels = _split_bul(row.get("price_abstract_resource_variable_parts"))
        values = _split_bul(row.get("price_abstract_resource_est_price_all_values"))
        pu_vals_raw = _split_bul(row.get("price_abstract_resource_est_price_all_values_per_unit"))
        pu_vals = [_strip_unit_prefix(t) for t in pu_vals_raw]
        # Some rate rows have an empty ``..._all_values`` series and only
        # the per-unit one is populated - fall back so the legacy picker
        # still gets a catalog instead of dropping it on import.
        if labels and len(values) != len(labels) and len(pu_vals) == len(labels):
            values = pu_vals
        # ``common_start`` is the shared base name for the abstract resource
        # (e.g. "Ready-mix concrete"); each ``variable_parts[i]`` is the
        # distinguishing tail (e.g. "C25/30 delivered"). The picker renders
        # ``common_start`` once as a header and the rows show only the
        # variable tails. ``full_label`` = ``common_start + variable_part``
        # is what the BOQ resource row displays after a pick - replacing
        # the position's default description so the user sees the actual
        # concrete material chosen.
        common_start = _safe_str(row.get("price_abstract_resource_common_start"))[:240]
        # position_count is a single per-rate_code total in the parquet, not per-variant.
        total_position_count = int(_safe_float(row.get("price_abstract_resource_position_count")))
        if labels and len(labels) > 1 and len(values) == len(labels):
            variants = []
            for i, (lbl, val) in enumerate(zip(labels, values, strict=False)):
                v = _safe_float(val)
                if v <= 0:
                    continue
                variable_part = lbl[:200]
                full_label = (f"{common_start} {variable_part}".strip() if common_start else variable_part)[:400]
                variants.append(
                    {
                        "index": i,
                        "label": variable_part,
                        "full_label": full_label,
                        "price": round(v, 2),
                        "price_per_unit": round(_safe_float(pu_vals[i]), 4) if i < len(pu_vals) else None,
                    }
                )
            if variants:
                metadata["variants"] = variants
                metadata["variant_stats"] = {
                    "min": round(_safe_float(row.get("price_abstract_resource_est_price_min")), 2),
                    "max": round(_safe_float(row.get("price_abstract_resource_est_price_max")), 2),
                    "mean": round(_safe_float(row.get("price_abstract_resource_est_price_mean")), 2),
                    "median": round(_safe_float(row.get("price_abstract_resource_est_price_median")), 2),
                    "unit": _safe_str(row.get("price_abstract_resource_unit"))[:20],
                    "group": _safe_str(row.get("price_abstract_resource_group_per_unit"))[:120],
                    "count": len(variants),
                    "position_count": total_position_count,
                    "common_start": common_start,
                }

        # Get full resource components for this rate_code
        components = resources_by_code.get(code, [])

        batch.append(
            (
                str(uuid.uuid4()),
                code,
                desc[:500],
                unit,
                str(rate),
                resolved_currency,
                "cwicr",
                _json.dumps(classification),
                "[]",
                _json.dumps(components),
                "{}",
                1,
                db_id,
                _json.dumps(metadata),
            )
        )

    # PostgreSQL: idempotent bulk insert via ON CONFLICT (code, region)
    # DO NOTHING. ``batch`` holds every accumulated row.
    imported = _pg_bulk_insert_cost_rows(db_file, batch)

    elapsed = round(time.monotonic() - start, 1)
    _log.info("CWICR %s: %d imported, %d skipped in %.1fs", db_id, imported, skipped_count, elapsed)

    # Total resource components carried by the imported work items. Each CWICR
    # work item (rate_code) bundles a labour/material/equipment breakdown in its
    # ``components`` array (the parquet is literally
    # ``..._workitems_costs_resources_...``); surfacing the aggregate count lets
    # the partner-pack installer report the embedded resource database it just
    # loaded alongside the work catalog.
    resource_components = sum(len(v) for v in resources_by_code.values())

    return {
        "imported": imported,
        "skipped": skipped_count,
        "total_rows": total_rows,
        "unique_items": len(grouped),
        "resource_components": resource_components,
        "database": db_id,
    }


def _build_cwicr_items(df: pd.DataFrame, db_id: str) -> list[dict[str, Any]]:  # noqa: F821
    """Legacy - kept for reference but no longer called."""
    import math

    import pandas as pd_local

    # Normalize column names
    df = df.copy()
    df.columns = [str(c).strip().lower() for c in df.columns]

    def _safe(val: object) -> float:
        if val is None:
            return 0.0
        try:
            f = float(val)  # type: ignore[arg-type]
            return 0.0 if math.isnan(f) else f
        except (ValueError, TypeError):
            return 0.0

    def _str(val: object) -> str:
        if val is None or (isinstance(val, float) and pd_local.isna(val)):
            return ""
        return str(val).strip()

    # Build description column
    if "rate_original_name" in df.columns and "rate_final_name" in df.columns:
        df.loc[:, "_full_desc"] = (
            df["rate_original_name"].fillna("").astype(str) + " " + df["rate_final_name"].fillna("").astype(str)
        ).str.strip()

    if "rate_code" not in df.columns:
        return []

    grouped = df.groupby("rate_code", sort=False)
    total_rows = len(df)
    logger.info("Grouped into %d unique rate_codes from %d rows", len(grouped), total_rows)

    result_items: list[dict[str, Any]] = []
    item_count = 0

    for rate_code, group in grouped:
        first = group.iloc[0]

        desc = _str(first.get("_full_desc", "")) if "_full_desc" in first.index else ""
        if len(desc) < 3:
            desc = _str(first.get("rate_original_name", ""))
        if len(desc) < 3:
            desc = _str(first.get("subsection_name", ""))
        if len(desc) < 3:
            continue

        code = _str(rate_code)[:100] or f"CWICR-{db_id}-{item_count:06d}"
        unit = _str(first.get("rate_unit", "m2"))[:20] or "m2"
        rate = _safe(first.get("total_cost_per_position", 0))

        classification: dict[str, str] = {}
        for key in ("collection_name", "department_name", "section_name", "subsection_name"):
            val = _str(first.get(key, ""))
            if val:
                classification[key.replace("_name", "")] = val
        cat_type = _str(first.get("category_type", ""))
        if cat_type:
            classification["category"] = cat_type

        # Extract summary metadata from first row only (skip per-row component iteration for speed)
        labor_total = _safe(first.get("cost_of_working_hours", 0))
        equipment_total = _safe(first.get("total_value_machinery_equipment", 0))
        material_total = _safe(first.get("total_material_cost_per_position", 0))
        labor_hours = _safe(first.get("total_labor_hours_all_personnel", 0))

        metadata: dict[str, Any] = {}
        if labor_total > 0:
            metadata["labor_cost"] = round(labor_total, 2)
        if equipment_total > 0:
            metadata["equipment_cost"] = round(equipment_total, 2)
        if material_total > 0:
            metadata["material_cost"] = round(material_total, 2)
        if labor_hours > 0:
            metadata["labor_hours"] = round(labor_hours, 2)
        workers = _safe(first.get("count_total_people_per_unit", 0))
        if workers > 0:
            metadata["workers_per_unit"] = round(workers, 1)

        result_items.append(
            {
                "code": code,
                "description": desc[:500],
                "unit": unit,
                "rate": str(round(rate, 2)),
                # CWICR parquets don't carry a currency column - every rate
                # is denominated in the region's local currency. Resolve
                # via the central region map so the picker shows the right
                # ISO code (e.g. RU_STPETERSBURG → RUB, not USD fallback).
                "currency": _resolve_currency(None, db_id),
                "source": "cwicr",
                "classification": classification,
                "tags": [],
                "components": [],
                "descriptions": {},
                "is_active": True,
                "region": db_id,
                "metadata": metadata,
            }
        )
        item_count += 1

    return result_items

    # Old processing code removed - now handled by _build_cwicr_items() + batch insert above
    pass  # unreachable - function returns above


async def _bulk_insert_costs(session: AsyncSession, items: list[dict]) -> int:
    """Async wrapper: runs the PostgreSQL bulk insert in a worker thread.

    Routes to the ON CONFLICT DO NOTHING bulk helper on a short-lived sync
    engine so the heavy load never blocks the async session pool.
    """
    import asyncio

    from app.config import get_settings

    settings = get_settings()

    # PostgreSQL: build positional rows matching the CWICR column order and run
    # the idempotent ON CONFLICT (code, region) DO NOTHING bulk insert.
    import json as _json

    rows: list[tuple] = [
        (
            str(uuid.uuid4()),
            item["code"],
            item["description"][:500],
            item["unit"][:20],
            item["rate"],
            # Resolve the region currency when the item has none, so PG rows
            # never land with an empty currency (matches the SQLite path).
            item.get("currency") or _resolve_currency(None, item.get("region", "")),
            item.get("source", "cwicr"),
            _json.dumps(item.get("classification", {})),
            "[]",
            "[]",
            "{}",
            1,
            item.get("region", ""),
            _json.dumps(item.get("metadata", {})),
        )
        for item in items
    ]
    return await asyncio.to_thread(_pg_bulk_insert_cost_rows, settings.database_sync_url, rows)


# ── Delete CWICR database ───────────────────────────────────────────────────


@router.delete(
    "/actions/clear-database/",
    dependencies=[Depends(RequireRole("admin"))],
)
async def clear_cost_database(
    session: SessionDep,
    _user_id: CurrentUserId,
    source: str = Query(
        default="",
        description="Filter by source (e.g. 'cwicr'). Empty = delete ALL.",
    ),
) -> dict:
    """Delete cost items. Optionally filter by source."""
    from sqlalchemy import delete as sql_delete

    from app.modules.costs.models import CostItem

    if source:
        stmt = sql_delete(CostItem).where(CostItem.source == source)
    else:
        stmt = sql_delete(CostItem)

    result = await session.execute(stmt)
    await session.commit()
    count = result.rowcount  # type: ignore[union-attr]

    _invalidate_cost_cache()
    return {"deleted": count, "source_filter": source or "all"}


# ── Export cost database as Excel ────────────────────────────────────────────


def _excel_safe(value: object) -> object:
    """Neutralize spreadsheet formula injection in exported text cells.

    Imported catalogues are user-supplied, so a cell starting with ``=``,
    ``+``, ``-``, ``@``, tab or CR would execute as a formula (or DDE
    payload) when another user opens the export in Excel/LibreOffice.
    Prefixing with an apostrophe makes Excel render it as literal text.
    """
    if isinstance(value, str) and value[:1] in ("=", "+", "-", "@", "\t", "\r"):
        return "'" + value
    return value


@router.get(
    "/actions/export-excel/",
    dependencies=[Depends(RequirePermission("costs.list"))],
)
async def export_cost_database(
    session: SessionDep,
    _user_id: CurrentUserId,
) -> StreamingResponse:
    """Export all cost items as Excel file.

    Uses openpyxl write_only mode and batched DB fetching (1000 rows)
    to keep memory usage constant regardless of dataset size.
    """
    from openpyxl import Workbook
    from sqlalchemy import select

    from app.modules.costs.models import CostItem

    wb = Workbook(write_only=True)
    ws = wb.create_sheet(title="Cost Database")

    # Header row
    ws.append(["Code", "Description", "Unit", "Rate", "Currency", "Source", "Region"])

    # Fetch in batches to avoid loading 50K+ rows into memory at once
    batch_size = 1000
    offset = 0
    base_stmt = select(CostItem).where(CostItem.is_active.is_(True)).order_by(CostItem.code)

    while True:
        result = await session.execute(base_stmt.offset(offset).limit(batch_size))
        items = result.scalars().all()
        if not items:
            break

        for item in items:
            try:
                rate_val = float(item.rate)
            except (ValueError, TypeError):
                rate_val = 0
            ws.append(
                [
                    _excel_safe(item.code),
                    _excel_safe(item.description),
                    _excel_safe(item.unit),
                    rate_val,
                    _excel_safe(item.currency),
                    _excel_safe(item.source),
                    _excel_safe(getattr(item, "region", "")),
                ]
            )

        if len(items) < batch_size:
            break
        offset += batch_size

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="cost_database.xlsx"'},
    )


# ── BIM-element cost suggestions ─────────────────────────────────────────────


@router.post(
    "/suggest-for-element/",
    response_model=list[CostSuggestion],
    dependencies=[Depends(RequirePermission("costs.read"))],
)
async def suggest_costs_for_element(
    request: SuggestCostsForElementRequest,
    _user_id: CurrentUserId,
    service: CostItemService = Depends(_get_service),
) -> list[CostSuggestion]:
    """Rank cost items that match a BIM element (body-only variant).

    The frontend already has the element loaded in the viewer, so the
    cheapest path is to pass the fields inline and avoid a second DB
    round-trip.  Returns at most ``request.limit`` suggestions sorted by
    relevance score (0..1).
    """
    return await service.suggest_for_bim_element(
        element_type=request.element_type,
        name=request.name,
        discipline=request.discipline,
        properties=request.properties,
        quantities=request.quantities,
        classification=request.classification,
        limit=request.limit,
        region=request.region,
    )


@router.post(
    "/suggest-for-element/{bim_element_id}/",
    response_model=list[CostSuggestion],
    dependencies=[Depends(RequirePermission("costs.read"))],
)
async def suggest_costs_for_element_by_id(
    bim_element_id: uuid.UUID,
    session: SessionDep,
    _user_id: CurrentUserId,
    limit: int = Query(default=5, ge=1, le=50),
    region: str | None = Query(default=None),
    service: CostItemService = Depends(_get_service),
) -> list[CostSuggestion]:
    """Convenience: load a ``BIMElement`` by ID and rank cost suggestions.

    Raises 404 if the element does not exist.  Classification is pulled
    from ``element.metadata_['classification']`` when present (BIM elements
    do not have a dedicated classification column).
    """
    # Local import to avoid a hard dependency loop between costs and bim_hub.
    from app.modules.bim_hub.service import BIMHubService

    bim_service = BIMHubService(session)
    try:
        element = await bim_service.get_element(bim_element_id)
    except HTTPException:
        raise
    except Exception as exc:  # pragma: no cover - defensive
        logger.exception("suggest_costs_for_element_by_id: failed to load element")
        raise HTTPException(status_code=500, detail="Failed to load BIM element") from exc

    # BIMElement has no `classification` column; pull from metadata if present.
    classification: dict[str, str] | None = None
    meta = getattr(element, "metadata_", None)
    if isinstance(meta, dict):
        candidate = meta.get("classification")
        if isinstance(candidate, dict):
            classification = {k: str(v) for k, v in candidate.items() if isinstance(v, (str, int))}

    # Quantities may contain non-float entries in practice; coerce safely.
    quantities_raw = getattr(element, "quantities", None) or {}
    quantities: dict[str, float] = {}
    if isinstance(quantities_raw, dict):
        for key, val in quantities_raw.items():
            try:
                quantities[str(key)] = float(val)
            except (TypeError, ValueError):
                continue

    return await service.suggest_for_bim_element(
        element_type=getattr(element, "element_type", None),
        name=getattr(element, "name", None),
        discipline=getattr(element, "discipline", None),
        properties=getattr(element, "properties", None),
        quantities=quantities,
        classification=classification,
        limit=limit,
        region=region,
    )


# ── CWICR Matcher (T12) ───────────────────────────────────────────────────


@router.post("/match/", response_model=list[MatchResult])
async def match_cwicr(
    request: CwicrMatchRequest,
    session: SessionDep,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
) -> list[MatchResult]:
    """Rank CWICR cost items for a free-form BOQ description.

    The endpoint is read-only against the cost database and is therefore
    public (matches the existing autocomplete + search endpoints). The
    optional ``mode`` selector chooses between ``lexical`` (always
    available), ``semantic`` (requires the ``[semantic]`` extra), and
    ``hybrid`` (blends both, falls back to lexical when deps absent).
    """
    _ = user_id  # accept anonymous - matches /autocomplete + /search
    return await match_cwicr_items(
        session,
        request.query,
        unit=request.unit,
        lang=request.lang,
        top_k=request.top_k,
        mode=request.mode,
        region=request.region,
    )


@router.post("/match-from-position/", response_model=list[MatchResult])
async def match_cwicr_from_position(
    request: CwicrMatchFromPositionRequest,
    session: SessionDep,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
) -> list[MatchResult]:
    """Resolve a Position by id and run the CWICR matcher on its description.

    Returns 404 if the position does not exist.  Empty list is returned
    (200) when the position has no description - that's the BOQ editor's
    "scroll past empty rows" UX path.
    """
    _ = user_id
    try:
        return await match_cwicr_for_position(
            session,
            request.position_id,
            top_k=request.top_k,
            mode=request.mode,
            lang=request.lang,
            region=request.region,
        )
    except LookupError as exc:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=str(exc)) from exc


# ── Cost Intelligence (v3.12.0 - Stream B) ────────────────────────────────


@router.get("/regional-adjust/", response_model=RegionalAdjustResponse)
async def regional_adjust(
    session: SessionDep,
    user: OptionalUserPayload,
    region: str = Query(..., min_length=2, max_length=64, description="Region code, e.g. DE_BERLIN"),
    category: str = Query(..., min_length=2, max_length=64, description="Category key"),
    # Round-7: ``Decimal`` for money - FastAPI parses the query string
    # without going through ``float``. The response model serialises
    # the values back out as strings so JS clients keep exact precision.
    base_rate: Decimal = Query(..., ge=0, description="Unit rate in the catalogue's currency"),
    subcategory: str | None = Query(
        default=None,
        max_length=64,
        description="Optional finer slice - falls back to the whole-category row when absent.",
    ),
) -> RegionalAdjustResponse:
    """Preview the same rate in a different region.

    Regional city cost index lookup - multiplies ``base_rate`` by
    the most recent ``factor`` on file for ``(region, category)``.
    When no factor exists, returns a 1:1 passthrough so the frontend
    can render the row without branching on null.

    Read-only and public (parity with autocomplete / search). The
    estimator is expected to confirm before applying the adjusted rate
    onto a BOQ position - no auto-apply.
    """
    _ = user  # accept anonymous
    svc = RegionalIndexService(session)
    adjusted, factor, source, effective = await svc.adjust(region, category, base_rate, subcategory=subcategory)
    return RegionalAdjustResponse(
        region=region.strip().upper(),
        category=category.strip().lower(),
        base_rate=base_rate,
        factor_applied=factor,
        adjusted_rate=adjusted,
        source=source,
        effective_date=effective,
    )


@router.get(
    "/regional-indices/",
    response_model=list[RegionalIndexResponse],
)
async def list_regional_indices(
    session: SessionDep,
    user: OptionalUserPayload,
    region: str = Query(..., min_length=2, max_length=64),
) -> list[RegionalIndexResponse]:
    """List every cost-index row for ``region``.

    Used by the Regional Adjust panel to populate the category picker
    and show historical effective dates. Ordered by category then
    effective_date desc so the freshest entries surface first.
    """
    _ = user
    svc = RegionalIndexService(session)
    rows = await svc.list_for_region(region)
    return [RegionalIndexResponse.model_validate(row) for row in rows]


@router.post("/certainty/batch/", response_model=list[CertaintyBadge])
async def get_cost_item_certainty_batch(
    body: CertaintyBatchRequest,
    session: SessionDep,
    user: OptionalUserPayload,
) -> list[CertaintyBadge]:
    """Return certainty badges for many cost items in a single round-trip.

    The list view renders one badge per visible row; fetching them
    individually fires N HTTP requests per page (one per row), which is
    a per-keystroke N+1 against the usage ledger. This endpoint folds
    the whole visible page into two grouped queries - ``count(*)`` and
    ``max(used_at)`` keyed by ``cost_item_id`` - then classifies each
    band in Python, so the page costs one request regardless of row
    count.

    Unknown ids are silently dropped (the badge is decorative - a
    missing row simply renders nothing on the client). Duplicate ids in
    the request collapse to one result. Public, mirroring the
    single-item endpoint.
    """
    _ = user

    # De-duplicate while preserving the caller's order so the response is
    # deterministic; cap at the request schema's max_length (validated on
    # ``ids``) so a hostile payload can't fan out into an unbounded IN().
    seen: set[uuid.UUID] = set()
    ordered_ids: list[uuid.UUID] = []
    for raw in body.ids:
        if raw not in seen:
            seen.add(raw)
            ordered_ids.append(raw)
    if not ordered_ids:
        return []

    from datetime import UTC, datetime

    from sqlalchemy import func

    from app.modules.costs.intelligence import (
        NEVER_USED_AGE_DAYS,
        classify_certainty,
    )
    from app.modules.costs.models import CostItemUsage
    from app.modules.costs.service import price_freshness

    # Only items that actually exist get a badge - resolve their ``source``,
    # price date and rate in one pass so the band carries the correct
    # provenance label and the price-freshness merge below has its inputs.
    item_rows = await session.execute(
        select(CostItem.id, CostItem.source, CostItem.price_as_of, CostItem.rate).where(CostItem.id.in_(ordered_ids))
    )
    source_by_id: dict[uuid.UUID, str] = {}
    price_info_by_id: dict[uuid.UUID, tuple[object, object]] = {}
    for row in item_rows.all():
        source_by_id[row[0]] = row[1] or "manual"
        price_info_by_id[row[0]] = (row[2], row[3])

    # Two grouped aggregates over the usage ledger - frequency + last use -
    # instead of one query per id. The composite index on
    # ``(cost_item_id, used_at)`` covers both.
    usage_rows = await session.execute(
        select(
            CostItemUsage.cost_item_id,
            func.count(CostItemUsage.id).label("freq"),
            func.max(CostItemUsage.used_at).label("last_used"),
        )
        .where(CostItemUsage.cost_item_id.in_(ordered_ids))
        .group_by(CostItemUsage.cost_item_id)
    )
    usage_by_id: dict[uuid.UUID, tuple[int, datetime | None]] = {
        row[0]: (int(row[1] or 0), row[2]) for row in usage_rows.all()
    }

    now = datetime.now(UTC)
    out: list[CertaintyBadge] = []
    for item_id in ordered_ids:
        if item_id not in source_by_id:
            continue
        frequency, last_used = usage_by_id.get(item_id, (0, None))
        if last_used is None:
            age_days = NEVER_USED_AGE_DAYS
            last_used_iso: datetime | None = None
        else:
            # SQLite returns naive datetimes; normalise to UTC-aware so the
            # diff matches the single-item endpoint's behaviour.
            if last_used.tzinfo is None:
                last_used = last_used.replace(tzinfo=UTC)
            age_days = max(0, int((now - last_used).total_seconds() // 86400))
            last_used_iso = last_used
        price_as_of, rate = price_info_by_id.get(item_id, (None, None))
        badge_data: dict[str, object] = {
            "cost_item_id": item_id,
            "frequency": frequency,
            "age_days": age_days,
            "source": source_by_id[item_id],
            "confidence_badge": classify_certainty(frequency, age_days),
            "last_used_at": last_used_iso,
        }
        # Same freshness merge as the single-item endpoint so the list badge
        # and the drill-down badge agree on the re-price-due state.
        badge_data.update(price_freshness(price_as_of, rate))
        out.append(CertaintyBadge.model_validate(badge_data))
    return out


@router.post("/usage-counts/")
async def get_cost_item_usage_counts(
    body: UsageCountsRequest,
    session: SessionDep,
    user: OptionalUserPayload,
) -> dict[str, int]:
    """Return ``{cost_item_id: usage_count}`` for the requested ids.

    Powers the Cost Database "used in N estimates" indicator. One grouped
    ``count(*)`` over the usage ledger keyed by ``cost_item_id`` resolves
    the whole visible page in a single round-trip - never one request per
    row. Ids with zero recorded uses are omitted from the response (the
    client treats a missing id as count 0), so the payload only carries
    the rows that actually have usage. Public, mirroring the certainty
    batch endpoint.
    """
    _ = user

    # De-duplicate, preserving order; the schema already caps ``ids`` length
    # so the grouped IN() can't fan out unboundedly.
    seen: set[uuid.UUID] = set()
    ordered_ids: list[uuid.UUID] = []
    for raw in body.ids:
        if raw not in seen:
            seen.add(raw)
            ordered_ids.append(raw)
    if not ordered_ids:
        return {}

    from sqlalchemy import func

    from app.modules.costs.models import CostItemUsage

    rows = await session.execute(
        select(
            CostItemUsage.cost_item_id,
            func.count(CostItemUsage.id).label("cnt"),
        )
        .where(CostItemUsage.cost_item_id.in_(ordered_ids))
        .group_by(CostItemUsage.cost_item_id)
    )
    # Only non-zero rows survive the GROUP BY; emit them as string keys so
    # the JSON object matches the frontend's ``Record<string, number>``.
    return {str(row[0]): int(row[1] or 0) for row in rows.all()}


@router.get("/{item_id}/certainty/", response_model=CertaintyBadge)
async def get_cost_item_certainty(
    item_id: uuid.UUID,
    session: SessionDep,
    user: OptionalUserPayload,
) -> CertaintyBadge:
    """Return the green / yellow / red certainty badge for one cost item.

    Aggregates the usage ledger into:

    * ``frequency`` - total recorded uses across all projects.
    * ``age_days`` - days since the most recent use (``999999`` when
      the item has never been used).
    * ``confidence_badge`` - bucketed band per the rules documented on
      ``schemas.CertaintyBadge``.

    Returns 404 when ``item_id`` does not exist in ``oe_costs_item``.
    """
    _ = user
    svc = CostCertaintyService(session)
    try:
        data = await svc.compute(item_id)
    except LookupError as exc:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=str(exc)) from exc
    return CertaintyBadge.model_validate(data)


@router.post(
    "/{item_id}/record-usage/",
    status_code=status.HTTP_201_CREATED,
    dependencies=[Depends(RequirePermission("costs.read"))],
)
async def record_cost_item_usage(
    item_id: uuid.UUID,
    body: RecordUsageRequest,
    session: SessionDep,
    user: CurrentUserPayload,
) -> dict[str, object]:
    """Append one row to the usage ledger.

    Called from the BOQ apply-rate path so the next user of the same
    rate sees an up-to-date certainty badge. Body intentionally small:
    the timestamp is server-stamped, the cost-item id rides on the
    URL.

    Returns the new usage row's id + the refreshed certainty band so
    the frontend can update its badge cache in one round-trip.

    Authenticated only. The usage ledger feeds the shared, cross-tenant
    certainty badge (``CostCertaintyService.compute`` counts every row for
    an item with no tenant scoping), so an anonymous writer could forge a
    rate's "proven" status and inflate any project's frequency. We require
    a usable subject and verify project access for EVERY caller.
    """
    # Verify cost item exists so we can give a precise 404 rather than
    # letting the FK CASCADE constraint do it at commit time.
    item_check = await session.execute(select(CostItem).where(CostItem.id == item_id).limit(1))
    if item_check.scalar_one_or_none() is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"CostItem {item_id} not found",
        )

    recorder = CostUsageRecorder(session)
    sub = (user or {}).get("sub")
    if not sub:
        # Authenticated route, but defend against a token without a subject.
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Not authenticated")
    try:
        used_by = uuid.UUID(str(sub))
    except (TypeError, ValueError) as exc:
        # A non-UUID subject (e.g. malformed token) cannot be attributed or
        # access-checked, so it must not be allowed to write the shared ledger.
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Not authenticated",
        ) from exc

    # IDOR guard: verify the caller can see the target project before we
    # record a usage row attributed to it. Without this check any user could
    # attribute apply-events to any project UUID they happen to know and bump
    # the item's shared certainty frequency. verify_project_access raises 404
    # for both missing and inaccessible projects.
    await verify_project_access(body.project_id, str(used_by), session)

    row = await recorder.record(
        item_id,
        project_id=body.project_id,
        unit_rate_at_use=body.unit_rate_at_use,
        context=body.context,
        used_by=used_by,
    )
    await session.commit()

    certainty = await CostCertaintyService(session).compute(item_id)
    return {
        "id": str(row.id),
        "cost_item_id": str(item_id),
        "used_at": row.used_at.isoformat() if row.used_at else None,
        "certainty": CertaintyBadge.model_validate(certainty).model_dump(mode="json"),
    }


# ── Cost benchmarks: own-portfolio distribution (Phase 2) ──────────────────


@router.post(
    "/benchmark/",
    response_model=BenchmarkResponse,
    dependencies=[Depends(RequirePermission("costs.read"))],
)
async def cost_benchmark_portfolio(
    body: BenchmarkRequest,
    session: SessionDep,
    user: CurrentUserPayload,
) -> BenchmarkResponse:
    """Benchmark a chosen metric against the tenant's OWN real projects.

    The default ``metric`` is ``cost_per_m2``: each of the caller's projects
    contributes its BOQ grand total divided by its recorded gross floor area.
    ``overrun_pct`` (priced scope over approved budget) and ``recovery_rate``
    (recovered share of chargeable cost) benchmark the same projects on two
    dimensionless ratios, giving a firm regional context for its own overrun and
    recovery when paired with the ``region`` filter. The endpoint returns the
    min / p25 / median / p75 / max of the distribution plus, when a value is
    supplied, where it sits within it.

    Industry reference ranges are NOT returned here; the client owns the
    richer static benchmark table and computes the industry percentile
    locally. This endpoint adds only the tenant-specific portfolio the
    client cannot compute, keeping it thin.

    Degrades gracefully: when the caller has no project with both a cost
    and an area, ``own_portfolio`` and ``percentile_vs_own`` are null and
    the response is still 200 so the client falls back to industry-only.
    """
    sub = (user or {}).get("sub")
    if not sub:
        # Authenticated route, but defend against a token without a subject.
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Not authenticated")
    owner_id = uuid.UUID(str(sub))
    is_admin = (user or {}).get("role") == "admin"

    result = await CostBenchmarkService(session).portfolio_distribution(
        owner_id=owner_id,
        is_admin=is_admin,
        building_type=body.building_type,
        region=body.region,
        currency=body.currency,
        cost_per_m2=body.cost_per_m2,
        metric=body.metric,
    )
    return BenchmarkResponse.model_validate(result)
