"""Design Options appraisal workbook builder.

Turns the side-by-side comparison the module already computes
(:class:`~app.modules.design_options.schemas.DesignOptionComparisonResponse`)
into one downloadable .xlsx an estimator or a client can open and audit. The
comparison is the single source of truth: the JSON endpoint and this spreadsheet
render the same numbers, in the same one comparison currency, so the download can
never disagree with what the screen shows.

The workbook has two sheets:

    Option Appraisal - the per-option matrix (direct cost, markups, grand total,
        delta versus the baseline, cost per m2, gross floor area, element and
        position counts, validation status), followed by the transparent
        recommendation and the set-level fairness banner. Options are alternative
        designs for the same project, so they are never summed: there is no TOTAL
        row across options, which would imply building all of them.

    By Trade - one row per trade bucket (a DIN 276 cost group, a MasterFormat
        division or a free-form trade tag), the baseline scope for that trade, and
        one cost column per option. Summing an option's column reconciles it back
        to that option's direct cost on the first sheet, so the headline numbers
        are auditable to the trade level.

Purity and safety
-----------------
This module is a pure projection of the comparison response: no database, no
network, no clock beyond the optional generated-on stamp. Money, quantity and
ratio values arrive as plain decimal strings (the platform Decimal-as-string
contract); they are parsed to :class:`decimal.Decimal` and handed to openpyxl as
real numbers so Excel stores them sortable and summable without a lossy float
round-trip, never blended across currencies because the comparison already
resolved one display currency. Every user-influenced text cell (set name, option
name, trade label) is routed through
:func:`app.core.csv_safety.neutralise_formula` so a value like ``=cmd|...`` can
never execute when the file is opened.

Offline i18n
-----------
A downloaded file carries no i18n runtime, so the recommendation reason and the
fairness notices, which the API emits as stable i18n keys, are rendered here into
honest English sentences (the same honest-English-default pattern the service
uses for its trade labels). An unknown key degrades to a readable de-camel-cased
label rather than leaking a raw key.
"""

from __future__ import annotations

from datetime import UTC, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from app.core.csv_safety import neutralise_formula
from app.modules.design_options.schemas import (
    DesignOptionComparisonResponse,
    DesignOptionFairnessWarning,
)

__all__ = [
    "XLSX_MEDIA_TYPE",
    "build_option_appraisal_workbook",
    "option_appraisal_filename",
]

# The single spreadsheet MIME type the download endpoint advertises, matching the
# BOQ / COBie exports so the browser saves a real file instead of rendering it.
XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

# ── Number formats ────────────────────────────────────────────────────────────
_MONEY_FORMAT = "#,##0.00"
_QTY_FORMAT = "#,##0.000"
_PCT_FORMAT = "#,##0.00"  # the value is already in percent units; the header says "%"
_RATIO_FORMAT = "0.00"
_COUNT_FORMAT = "#,##0"

# ── Styles (reused across cells; never mutated in place) ──────────────────────
_TITLE_FONT = Font(bold=True, size=14)
_META_FONT = Font(italic=True, color="FF6B7280")
_SECTION_FONT = Font(bold=True, size=11)
_HEADER_FONT = Font(bold=True, color="FFFFFFFF")
_BOLD_FONT = Font(bold=True)
_HEADER_FILL = PatternFill(start_color="FF1A1A2E", end_color="FF1A1A2E", fill_type="solid")
_SECTION_FILL = PatternFill(start_color="FFE8E8EE", end_color="FFE8E8EE", fill_type="solid")
_BASELINE_FILL = PatternFill(start_color="FFEFF6FF", end_color="FFEFF6FF", fill_type="solid")
_TOTAL_FILL = PatternFill(start_color="FFF3F4F6", end_color="FFF3F4F6", fill_type="solid")
_OK_FILL = PatternFill(start_color="FFDCFCE7", end_color="FFDCFCE7", fill_type="solid")
_WARN_FILL = PatternFill(start_color="FFFEF3C7", end_color="FFFEF3C7", fill_type="solid")
_ERROR_FILL = PatternFill(start_color="FFFEE2E2", end_color="FFFEE2E2", fill_type="solid")
_WRAP = Alignment(wrap_text=True, vertical="top")
_LEFT = Alignment(horizontal="left", vertical="center")

# ── Offline renderings of the comparison i18n keys ────────────────────────────
# The API emits stable i18n keys; a downloaded file has no i18n runtime, so the
# text is rendered here in honest English. Keep in step with the keys the
# comparator emits (recommendation reasons + fairness notices). An unknown key
# falls back to a de-camel-cased label so nothing ever leaks a raw dotted key.
_REASON_TEXT: dict[str, str] = {
    "designOptions.recommendation.lowestCostPerM2": "Lowest cost per m2 among the priced options.",
    "designOptions.recommendation.lowestTotal": "Lowest grand total among the priced options.",
    "designOptions.recommendation.onlyOption": "The only priced option in the set.",
    "designOptions.recommendation.none": "No option could be priced, so none is recommended yet.",
}

_FAIRNESS_TEXT: dict[str, str] = {
    "designOptions.fairness.singleOption": ("Only one option in the set, so there is nothing to compare against."),
    "designOptions.fairness.noBaseline": ("No baseline option was chosen, so the deltas are not shown."),
    "designOptions.fairness.unpricedOptions": ("{count} option(s) are not priced yet and read as zero."),
    "designOptions.fairness.mixedCurrencyOption": (
        "{count} option(s) mix currencies in their own bill; each total was converted "
        "to the comparison currency before it was summed, never blended."
    ),
    "designOptions.fairness.comparisonCurrencyUnavailable": (
        "The requested comparison currency {requested} could not be applied; the figures are shown in {used}."
    ),
    "designOptions.fairness.missingGfa": (
        "At least one priced option has no gross floor area, so its cost per m2 is not meaningful."
    ),
    "designOptions.fairness.mixedGfa": ("Options use different gross floor areas, so read the cost per m2 with care."),
    "designOptions.fairness.validationPending": "No option has been validated yet.",
}

_CLASSIFICATION_LABEL: dict[str, str] = {
    "din276": "DIN 276",
    "masterformat": "MasterFormat",
    "trade": "Trade",
    "none": "Unclassified",
}

_STATUS_LABEL: dict[str, str] = {"ok": "OK", "warnings": "Warnings", "error": "Attention needed"}
_STATUS_FILL: dict[str, PatternFill] = {"ok": _OK_FILL, "warnings": _WARN_FILL, "error": _ERROR_FILL}
_SEVERITY_RANK: dict[str, int] = {"error": 0, "warning": 1, "info": 2}
_SEVERITY_LABEL: dict[str, str] = {"error": "Error", "warning": "Warning", "info": "Info"}


# ── Value helpers ─────────────────────────────────────────────────────────────


def _dec(value: Any) -> Decimal | None:
    """Parse a decimal-string / number into a finite Decimal, else ``None``.

    Returns ``None`` for blank, non-numeric or non-finite input so the caller can
    leave the cell empty rather than write a bogus number.
    """
    if value is None:
        return None
    if isinstance(value, Decimal):
        return value if value.is_finite() else None
    try:
        parsed = Decimal(str(value).strip())
    except (InvalidOperation, ValueError, TypeError):
        return None
    return parsed if parsed.is_finite() else None


def _num_cell(ws: Worksheet, row: int, col: int, raw: Any, number_format: str) -> None:
    """Write a numeric cell from a decimal-string source.

    A parseable value is stored as a real :class:`Decimal` (sortable / summable in
    Excel, no float drift) with the given number format; anything unparseable is
    written as neutralised text so the cell is never silently wrong.
    """
    parsed = _dec(raw)
    if parsed is None:
        _text_cell(ws, row, col, "" if raw is None else str(raw))
        return
    cell = ws.cell(row=row, column=col, value=parsed)
    cell.number_format = number_format


def _text_cell(ws: Worksheet, row: int, col: int, value: Any) -> Any:
    """Write a formula-injection-safe text cell."""
    return ws.cell(row=row, column=col, value=neutralise_formula("" if value is None else str(value)))


def _header_cell(ws: Worksheet, row: int, col: int, label: str) -> None:
    """Write one dark, bold, white header cell (data-derived labels are neutralised)."""
    cell = ws.cell(row=row, column=col, value=neutralise_formula(label))
    cell.font = _HEADER_FONT
    cell.fill = _HEADER_FILL
    cell.alignment = _WRAP


def _section_heading(ws: Worksheet, row: int, text: str, span: int) -> None:
    """Write a shaded section heading spanning ``span`` columns."""
    cell = ws.cell(row=row, column=1, value=neutralise_formula(text))
    cell.font = _SECTION_FONT
    cell.fill = _SECTION_FILL
    for col in range(2, span + 1):
        ws.cell(row=row, column=col).fill = _SECTION_FILL


def _fmt_dt(value: datetime) -> str:
    """UTC ISO-second stamp for the generated-on line."""
    if value.tzinfo is not None:
        value = value.astimezone(UTC).replace(tzinfo=None)
    return value.strftime("%Y-%m-%dT%H:%M:%S")


def _decamel(text: str) -> str:
    """Turn a camelCase key tail into a readable label (fallback for unknown keys)."""
    out: list[str] = []
    for i, ch in enumerate(text):
        if ch.isupper() and i > 0 and not text[i - 1].isupper():
            out.append(" ")
        out.append(ch)
    spaced = "".join(out).strip()
    return (spaced[:1].upper() + spaced[1:]) if spaced else text


def _render_reason(reason_key: str) -> str:
    """Human-readable recommendation basis for the given i18n reason key."""
    if not reason_key:
        return ""
    text = _REASON_TEXT.get(reason_key)
    return text if text is not None else _decamel(reason_key.rsplit(".", 1)[-1])


def _render_fairness(warning: DesignOptionFairnessWarning) -> str:
    """Human-readable fairness notice, interpolating its context values."""
    template = _FAIRNESS_TEXT.get(warning.key)
    if template is None:
        return _decamel(warning.key.rsplit(".", 1)[-1])
    try:
        return template.format(**(warning.context or {}))
    except (KeyError, IndexError, ValueError):
        return template


def _autosize(ws: Worksheet, *, min_row: int = 1) -> None:
    """Best-effort column widths so the sheet opens readable, capped so one long
    label cannot blow a column out to the full row."""
    for column_cells in ws.columns:
        longest = 0
        letter = None
        for cell in column_cells:
            if cell.row < min_row:
                continue
            letter = cell.column_letter
            value = cell.value
            if value is not None:
                longest = max(longest, len(str(value)))
        if letter is not None:
            ws.column_dimensions[letter].width = min(max(longest + 2, 10), 48)


# ── Sheet 1: option appraisal matrix + recommendation + fairness ──────────────


def _write_matrix_sheet(ws: Worksheet, comparison: DesignOptionComparisonResponse, now: datetime) -> None:
    options = list(comparison.options)
    ccy = (comparison.comparison_currency or "").strip()
    money_suffix = f" ({ccy})" if ccy else ""
    span = 13

    names = {col.option_id: (col.name or "") for col in options}
    baseline_id = comparison.baseline_option_id
    recommended_id = comparison.recommendation.option_id
    baseline_name = names.get(baseline_id, "-") if baseline_id is not None else "-"

    ws.cell(
        row=1, column=1, value=neutralise_formula(f"Design Option Appraisal - {comparison.set_name or ''}")
    ).font = _TITLE_FONT
    meta = (
        f"Comparison currency: {ccy or '-'}    Baseline: {baseline_name or '-'}    "
        f"Options: {len(options)}    Generated: {_fmt_dt(now)}"
    )
    ws.cell(row=2, column=1, value=neutralise_formula(meta)).font = _META_FONT

    row = 4
    _section_heading(ws, row, "Options", span)
    row += 1

    headers = [
        "#",
        "Option",
        f"Direct cost{money_suffix}",
        f"Markups{money_suffix}",
        f"Grand total{money_suffix}",
        f"Delta vs baseline{money_suffix}",
        "Delta %",
        f"Cost per m2{money_suffix}",
        "GFA (m2)",
        "Elements",
        "Positions",
        "Validation",
        "Note",
    ]
    for idx, label in enumerate(headers, start=1):
        _header_cell(ws, row, idx, label)
    row += 1

    for line_no, col in enumerate(options, start=1):
        is_baseline = baseline_id is not None and col.option_id == baseline_id
        roles: list[str] = []
        if is_baseline:
            roles.append("Baseline")
        if recommended_id is not None and col.option_id == recommended_id:
            roles.append("Recommended")

        ws.cell(row=row, column=1, value=line_no).number_format = _COUNT_FORMAT
        name_cell = _text_cell(ws, row, 2, col.name or "")
        if is_baseline:
            name_cell.font = _BOLD_FONT
        _num_cell(ws, row, 3, col.direct_cost, _MONEY_FORMAT)
        _num_cell(ws, row, 4, col.markups_total, _MONEY_FORMAT)
        _num_cell(ws, row, 5, col.grand_total, _MONEY_FORMAT)
        _num_cell(ws, row, 6, col.delta_vs_baseline, _MONEY_FORMAT)
        if col.delta_pct is not None:
            _num_cell(ws, row, 7, col.delta_pct, _PCT_FORMAT)
        _num_cell(ws, row, 8, col.cost_per_m2, _MONEY_FORMAT)
        _num_cell(ws, row, 9, col.gfa, _QTY_FORMAT)
        ws.cell(row=row, column=10, value=int(col.element_count or 0)).number_format = _COUNT_FORMAT
        ws.cell(row=row, column=11, value=int(col.position_count or 0)).number_format = _COUNT_FORMAT
        _text_cell(ws, row, 12, col.validation_status or "pending")
        _text_cell(ws, row, 13, ", ".join(roles))

        if is_baseline:
            for c in range(1, span + 1):
                ws.cell(row=row, column=c).fill = _BASELINE_FILL
        row += 1

    if not options:
        _text_cell(ws, row, 1, "No options in this set yet.")
        row += 1

    # ── Recommendation ────────────────────────────────────────────────────────
    row += 1
    _section_heading(ws, row, "Recommendation", span)
    row += 1
    recommended_name = names.get(recommended_id, str(recommended_id)) if recommended_id is not None else "None"
    _text_cell(ws, row, 1, "Recommended option").font = _BOLD_FONT
    _text_cell(ws, row, 2, recommended_name)
    row += 1
    _text_cell(ws, row, 1, "Confidence (0-1)").font = _BOLD_FONT
    _num_cell(ws, row, 2, comparison.recommendation.confidence, _RATIO_FORMAT)
    row += 1
    _text_cell(ws, row, 1, "Basis").font = _BOLD_FONT
    basis_cell = _text_cell(ws, row, 2, _render_reason(comparison.recommendation.reason_key))
    basis_cell.alignment = _WRAP
    row += 2

    # ── Fairness banner ───────────────────────────────────────────────────────
    fairness = comparison.fairness
    status = (fairness.status or "ok").strip().lower()
    heading = ws.cell(
        row=row,
        column=1,
        value=neutralise_formula(f"Fairness check: {_STATUS_LABEL.get(status, status)}"),
    )
    heading.font = _SECTION_FONT
    status_fill = _STATUS_FILL.get(status, _SECTION_FILL)
    for c in range(1, span + 1):
        ws.cell(row=row, column=c).fill = status_fill
    row += 1

    _header_cell(ws, row, 1, "Severity")
    _header_cell(ws, row, 2, "Check")
    row += 1

    ordered = sorted(fairness.warnings, key=lambda w: _SEVERITY_RANK.get(w.severity, 3))
    if ordered:
        for warning in ordered:
            _text_cell(ws, row, 1, _SEVERITY_LABEL.get(warning.severity, warning.severity or "info"))
            note_cell = _text_cell(ws, row, 2, _render_fairness(warning))
            note_cell.alignment = _WRAP
            row += 1
    else:
        _text_cell(ws, row, 1, "OK")
        _text_cell(ws, row, 2, "No fairness concerns were raised.")
        row += 1

    _autosize(ws, min_row=4)
    # Keep the option matrix header visible while scrolling the blocks below it.
    ws.freeze_panes = "A6"


# ── Sheet 2: by-trade cost breakdown ──────────────────────────────────────────


def _write_by_trade_sheet(ws: Worksheet, comparison: DesignOptionComparisonResponse, now: datetime) -> None:
    options = list(comparison.options)
    ccy = (comparison.comparison_currency or "").strip()
    money_suffix = f" ({ccy})" if ccy else ""
    baseline_id = comparison.baseline_option_id

    front_cols = 5  # #, Trade, Classification, Baseline qty, Baseline unit
    span = front_cols + len(options)

    ws.cell(
        row=1,
        column=1,
        value=neutralise_formula(f"Cost by trade - {comparison.set_name or ''}"),
    ).font = _TITLE_FONT
    meta = f"Comparison currency: {ccy or '-'}    Trades: {len(comparison.by_trade)}    Generated: {_fmt_dt(now)}"
    ws.cell(row=2, column=1, value=neutralise_formula(meta)).font = _META_FONT

    row = 4
    _section_heading(ws, row, "Cost by trade (one column per option)", max(span, 1))
    row += 1

    headers = ["#", "Trade", "Classification", "Baseline qty", "Baseline unit"]
    for idx, label in enumerate(headers, start=1):
        _header_cell(ws, row, idx, label)
    for offset, col in enumerate(options):
        _header_cell(ws, row, front_cols + 1 + offset, f"{col.name or ''}{money_suffix}")
    row += 1

    # Per-option running totals so the sheet reconciles to each option's direct cost.
    totals: dict[Any, Decimal] = {col.option_id: Decimal("0") for col in options}

    if comparison.by_trade:
        for line_no, trade in enumerate(comparison.by_trade, start=1):
            costs = {cell.option_id: cell.cost for cell in trade.per_option}
            baseline_unit = next(
                (cell.unit for cell in trade.per_option if cell.option_id == baseline_id),
                "",
            )
            ws.cell(row=row, column=1, value=line_no).number_format = _COUNT_FORMAT
            label_cell = _text_cell(ws, row, 2, trade.label or trade.key or "Unclassified")
            label_cell.alignment = _WRAP
            _text_cell(ws, row, 3, _CLASSIFICATION_LABEL.get(trade.classification_system, trade.classification_system))
            _num_cell(ws, row, 4, trade.baseline_quantity, _QTY_FORMAT)
            _text_cell(ws, row, 5, baseline_unit)
            for offset, col in enumerate(options):
                raw_cost = costs.get(col.option_id, "0")
                _num_cell(ws, row, front_cols + 1 + offset, raw_cost, _MONEY_FORMAT)
                parsed = _dec(raw_cost)
                if parsed is not None:
                    totals[col.option_id] += parsed
            row += 1

        # TOTAL row: summing an option's column returns that option's direct cost.
        total_label = ws.cell(row=row, column=1, value="TOTAL")
        total_label.font = _BOLD_FONT
        for c in range(1, front_cols + 1):
            ws.cell(row=row, column=c).fill = _TOTAL_FILL
        for offset, col in enumerate(options):
            cell = ws.cell(row=row, column=front_cols + 1 + offset, value=totals[col.option_id])
            cell.number_format = _MONEY_FORMAT
            cell.font = _BOLD_FONT
            cell.fill = _TOTAL_FILL
        row += 1
    else:
        _text_cell(ws, row, 1, "No priced positions in any option yet.")
        row += 1

    _autosize(ws, min_row=4)
    if span >= 1:
        ws.freeze_panes = ws.cell(row=6, column=front_cols + 1).coordinate


# ── Public API ────────────────────────────────────────────────────────────────


def build_option_appraisal_workbook(
    comparison: DesignOptionComparisonResponse,
    *,
    generated_at: datetime | None = None,
) -> bytes:
    """Build the option-appraisal .xlsx for a computed comparison.

    Args:
        comparison: The side-by-side comparison to render (already resolved to one
            comparison currency; every money / quantity / ratio value is a plain
            decimal string).
        generated_at: Optional generated-on stamp; defaults to now (UTC). Pass a
            fixed value for deterministic snapshot tests.

    Returns:
        The workbook as .xlsx bytes.
    """
    now = generated_at or datetime.now(UTC)

    wb = Workbook()
    matrix = wb.active
    if matrix is not None:
        matrix.title = "Option Appraisal"
    _write_matrix_sheet(wb["Option Appraisal"], comparison, now)
    _write_by_trade_sheet(wb.create_sheet("By Trade"), comparison, now)

    # Stamp authorship into docProps so a downloaded appraisal carries its origin
    # even when the UI strings are localised away (mirrors the BOQ / COBie export).
    try:
        wb.properties.creator = "OpenConstructionERP · DataDrivenConstruction"
        wb.properties.lastModifiedBy = "OpenConstructionERP"
        wb.properties.title = f"Design Option Appraisal - {comparison.set_name or ''}"
        wb.properties.description = "Generated by OpenConstructionERP (https://openconstructionerp.com)"
    except Exception:  # noqa: BLE001 - best-effort metadata stamp
        pass

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def option_appraisal_filename(set_name: str) -> str:
    """Build a safe download filename for a set's appraisal workbook.

    Control characters (including CR / LF, which would let a crafted set name
    inject an HTTP response header) are stripped while ordinary Unicode is kept,
    so :func:`app.core.http_headers.content_disposition_attachment` can still
    surface the original name through its RFC 6266 ``filename*`` parameter.
    """
    base = "".join(ch for ch in (set_name or "") if ch.isprintable()).strip()
    base = base.replace("/", "-").replace("\\", "-")
    base = base or "appraisal"
    return f"Design Option Appraisal - {base}.xlsx"
